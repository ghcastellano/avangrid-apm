import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import io
import difflib

# --- PAGE CONFIG ---
st.set_page_config(page_title="APM Strategy Generator", page_icon="📊", layout="wide")

st.title("📊 APM Strategy Generator")
st.markdown("**McKinsey-Style Portfolio Analysis Engine**")
st.markdown("Upload your completed `Application Questionnaire` to automatically generate the Strategic Assessment, Dashboard, and Roadmap.")

# --- CONFIG DATA ---
synergy_blocks = {
    "Strategic Fit": {"Type": "Business", "Weight": 30, "Defs": {1: "Completely misaligned", 2: "Partially aligned", 3: "Neutral", 4: "Well-aligned", 5: "Strategic driver"}},
    "Business Efficiency": {"Type": "Business", "Weight": 30, "Defs": {1: "Manual", 2: "Low efficiency", 3: "Average", 4: "High", 5: "Optimized"}},
    "User Value": {"Type": "Business", "Weight": 20, "Defs": {1: "Rejected", 2: "Low satisfaction", 3: "Acceptable", 4: "Good", 5: "Delightful"}},
    "Financial Value": {"Type": "Business", "Weight": 20, "Defs": {1: "Negative", 2: "Poor", 3: "Neutral", 4: "Positive", 5: "Exceptional"}},
    "Architecture": {"Type": "Tech", "Weight": 30, "Defs": {1: "Obsolete", 2: "Aging", 3: "Stable", 4: "Modern", 5: "Future-proof"}},
    "Operational Risk": {"Type": "Tech", "Weight": 30, "Defs": {1: "Critical", 2: "High", 3: "Managed", 4: "Low", 5: "Fortified"}},
    "Maintainability": {"Type": "Tech", "Weight": 25, "Defs": {1: "Impossible", 2: "Hard", 3: "Standard", 4: "Good", 5: "Excellent"}},
    "Support Quality": {"Type": "Tech", "Weight": 15, "Defs": {1: "Non-existent", 2: "Reactive", 3: "Defined", 4: "Proactive", 5: "World-class"}}
}

MASTER_QUESTIONS = {
    "Strategic Fit": ["What is the name of the application?", "What is the primary business purpose of the application?", "Which OPCOs use the application?", "Which utility domain(s) is the application used for? (Electric, Gas, or Both)", "Which Business Unit(s) use or own the application?", "Is the application IT-owned, business-owned, or jointly governed?", "Is this application considered business-critical, important, or supportive? Provide an explanation of the statement", "Does the application align with the current and future Mobility strategy?", "Are there important capabilities missing that limit business effectiveness?", "Is the application expected to be used in the next 3–5 years?", "Are there planned upgrades, migrations, or replacements?", "Could this application be replaced or consolidated with another platform?"],
    "Business Efficiency": ["What key business processes does the application support?", "What core functionalities does the application provide?", "Are any processes partially supported or handled outside the application (manual workarounds, spreadsheets, etc.)?", "Does the application overlap functionally with other systems?", "Could this application absorb business processes currently supported by another platform or executed through manual workarounds? If yes, which processes and which platform(s)"],
    "User Value": ["Which user roles or personas use the application (e.g., field technician, dispatcher, supervisor)?", "How many active users does the application have (daily/monthly)?", "Is application usage growing, stable, or declining?", "Is usage mandatory or optional for users?", "What is the overall level of user satisfaction?", "Are there known usability or mobility experience issues?"],
    "Financial Value": ["What business value does the application deliver today?", "What would be the business impact if the application were unavailable?", "What are the main cost components (licenses, infrastructure, support)?, and what is the total cost of ownership?", "Is the cost reasonable compared to the business value delivered?", "Are there upcoming license renewals or contract milestones?", "Are there opportunities for cost reduction through consolidation or modernization?"],
    "Architecture": ["Is the application a custom-built solution or a market (COTS/SaaS) product?", "What platforms does the application run on (mobile OS, web, backend)?", "What technologies, frameworks, or programming languages are used?", "What version of the application is currently deployed?", "Is the application deployed on-premises, in the cloud, or in a hybrid model?", "If cloud-based, which hyperscaler or cloud provider is used (e.g., AWS, Azure, GCP)?", "Which systems does the application integrate with?", "Are integrations real-time, batch-based, or manual?", "What limits future evolution or innovation?"],
    "Operational Risk": ["Does the application support regulatory and compliance requirements?", "How critical are these integrations to business operations?", "What type of data does the application create, consume, or update?", "Are there known integration or data quality issues?", "Does the application handle sensitive, personal, or regulated data?", "Is the application governed by corporate security and IT policies?", "Is identity and access management (IAM) integrated with corporate IAM solutions?", "Are security controls (authentication, authorization, logging) centrally managed or application-specific?", "Are there known security risks, audit findings, or compliance gaps?"],
    "Maintainability": ["How complex is ongoing maintenance and support?", "How frequently are incidents or defects reported?", "How easy is it to implement enhancements or changes?", "What are the main business challenges with the application?", "What are the main technical challenges or limitations?", "Are there scalability, performance, or reliability concerns?", "Are stakeholders requesting changes or replacement?"],
    "Support Quality": ["Is vendor or technology support still available and active?", "Who provides application support (internal IT, vendor, third party)?", "Is the application proactively monitored (e.g., APM, Dynatrace), or is downtime primarily reported by users?", "Are alerts integrated with ITSM tools (e.g., ServiceNow) for auto-ticketing, or are they email-based/manual?"]
}

ALL_QUESTIONS_FLAT = []
for k, q_list in MASTER_QUESTIONS.items():
    ALL_QUESTIONS_FLAT.extend(q_list)

MANUAL_APPS_DATA = {
    "Kaffa": {
        "What is the name of the application?": "Kaffa Platform",
        "What is the primary business purpose of the application?": "Asset Lifecycle Management (ALM) & Mobile GIS. Unifies Network Design, Construction, and Maintenance.",
        "Which OPCOs use the application?": "NY, CMP and UI in the future",
        "Is this application considered business-critical, important, or supportive? Provide an explanation of the statement": "High (Tier 2/1). Manages network construction and 'As-Built' asset registration.",
        "Does the application align with the current and future Mobility strategy?": "High. Construction Digitalization and Data Quality are strategic pillars.",
        "What business value does the application deliver today?": "Asset Base Integrity (RAB). Ensures that what is in the field matches the system.",
        "What key business processes does the application support?": "1. Network Design Projects. 2. Construction Monitoring. 3. Asset Commissioning.",
        "Are any processes partially supported or handled outside the application (manual workarounds, spreadsheets, etc.)?": "Yes. It usually replaces paper, but complex SAP integration failures can lead to manual steps.",
        "Does the application overlap functionally with other systems?": "High with SAP & GE Smallworld/Esri. Kaffa acts as the Middleware.",
        "Could this application absorb business processes currently supported by another platform or executed through manual workarounds? If yes, which processes and which platform(s)": "Yes. Can absorb other smaller inspection and registration tools.",
        "What is the overall level of user satisfaction?": "Good (UX Focus). Marketed as 'Easy to use'.",
        "Is application usage growing, stable, or declining?": "Growing (New Implementation). Since 2022.",
        "What are the main cost components (licenses, infrastructure, support)?, and what is the total cost of ownership?": "Medium. Licensing + Integration Services.",
        "Is the cost reasonable compared to the business value delivered?": "Positive. Revenue recovery via correct asset registration pays the bill.",
        "What technologies, frameworks, or programming languages are used?": "Java / .NET (Backend) + Mobile Native. Model Driven Architecture.",
        "Which systems does the application integrate with?": "GIS (Esri/Smallworld) and ERP (SAP).",
        "Are integrations real-time, batch-based, or manual?": "API / Services (Bus). Complex bidirectional.",
        "Does the application handle sensitive, personal, or regulated data?": "Yes (CEII). Detailed electric network maps.",
        "Are there known security risks, audit findings, or compliance gaps?": "Vendor Supply Chain (International).",
        "Is identity and access management (IAM) integrated with corporate IAM solutions?": "Yes. Supports Enterprise integration.",
        "How complex is ongoing maintenance and support?": "Medium/High. Keeping engineering rules synchronized requires effort.",
        "How easy is it to implement enhancements or changes?": "High (Configurable).",
        "Who provides application support (internal IT, vendor, third party)?": "Vendor (Codex) + Internal IT.",
        "Is vendor or technology support still available and active?": "Active."
    },
    "Mapping Computer": {
        "What is the name of the application?": "Mapping Computer (Ad-hoc Process)",
        "What is the primary business purpose of the application?": "Provide offline updated maps to field workers. Copies maps from shared folder to hard drives.",
        "Is this application considered business-critical, important, or supportive? Provide an explanation of the statement": "Critical",
        "Does the application align with the current and future Mobility strategy?": "No",
        "What key business processes does the application support?": "Share maps",
        "What core functionalities does the application provide?": "Create and update maps, download maps",
        "Are any processes partially supported or handled outside the application (manual workarounds, spreadsheets, etc.)?": "Scripts to be executed manually",
        "Does the application overlap functionally with other systems?": "Yes",
        "Could this application be replaced or consolidated with another platform?": "Yes, documentviewer or any application to share documents",
        "Is usage mandatory or optional for users?": "Mandatory",
        "Is application usage growing, stable, or declining?": "Stable",
        "Is the application deployed on-premises, in the cloud, or in a hybrid model?": "On the devices (Local)",
        "What limits future evolution or innovation?": "Bad architecture",
        "Does the application handle sensitive, personal, or regulated data?": "Gas and electric maps",
        "Is the application governed by corporate security and IT policies?": "No",
        "Is identity and access management (IAM) integrated with corporate IAM solutions?": "No",
        "Are there known security risks, audit findings, or compliance gaps?": "Yes, security and architectural gaps found",
        "Are there known integration or data quality issues?": "Yes",
        "Is the application a custom-built solution or a market (COTS/SaaS) product?": "Custom",
        "How complex is ongoing maintenance and support?": "Easy",
        "Are stakeholders requesting changes or replacement?": "Yes",
        "Who provides application support (internal IT, vendor, third party)?": "Internal IT",
        "Is vendor or technology support still available and active?": "No",
        "Is the application proactively monitored (e.g., APM, Dynatrace), or is downtime primarily reported by users?": "No"
    }
}

KEYWORDS = {
    "high": ["automation", "optimized", "cloud", "modern", "integrated", "innovative", "strategic", "secure", "critical", "differentiator", "gold standard", "mandatory", "essential", "growing", "stable", "lifecycle", "unified", "mobile", "configurable", "active", "yes. supports", "middleware"],
    "med": ["standard", "functional", "managed", "defined", "accepted", "on premise", "market", "medium"],
    "low": ["manual", "legacy", "obsolete", "error", "poor", "risk", "unsupported", "silo", "redundant", "costly", "one person", "custom built", "unknown", "no", "scripts", "gaps", "bad architecture", "custom", "na"]
}

MANUAL_LAYOUT = {"PoleForeman": 'l', "Aspen OneLiner": 'b', "SCAL-360 N": 'r', "Mapping Computers": 't', "Mapping Computer": 't', "Standard Tracking": 'b', "ARCOS": 't', "Cathodic": 'r', "Bentley View": 'l', "Bentley - ProjectWise": 'r'} 

# --- HELPERS ---
def style(cell, bold=False, bg=None, color=None, border=True, align='center', wrap=False, italic=False, size=None, underline=None):
    font_args = {'bold': bold, 'italic': italic}
    if color: font_args['color'] = color
    if size: font_args['size'] = size
    if underline: font_args['underline'] = underline
    cell.font = Font(**font_args)
    if bg: cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    if border:
        bd = Side(style='thin', color='000000')
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    al_args = {'horizontal': align, 'vertical': 'top'}
    if wrap: al_args['wrap_text'] = True
    cell.alignment = Alignment(**al_args)

def evaluate_answer_text(text):
    if not text or len(text.strip()) < 2: return 0
    t = text.lower()
    score = 3
    pos_hits = sum(1 for w in KEYWORDS["high"] if w in t)
    neg_hits = sum(1 for w in KEYWORDS["low"] if w in t)
    if "no" == t.strip(): return 1
    if "yes" == t.strip(): return 4
    if pos_hits > neg_hits: score = 4
    if pos_hits > 2 and neg_hits == 0: score = 5
    if neg_hits > pos_hits: score = 2
    if neg_hits > 1: score = 1
    if ("security" in t or "iam" in t) and "no" in t: score = 1
    return score

def smart_wrap(text, limit=12):
    if len(text) <= limit: return text
    return "\n".join([text[i:i+limit] for i in range(0, len(text), limit)])

def smart_merge(base_answers, new_answers):
    # Update base_answers with new_answers ONLY if new_answer is valid
    for q, data in new_answers.items():
        if data.get("a") and len(str(data["a"]).strip()) > 1:
            base_answers[q] = data
    return base_answers

def parse_wb(wb):
    apps = []
    start_found = False
    target_start_name = "Bentley - AssetWise"
    
    # Track which manual apps we found to avoid duplicates if needed, 
    # though we are forcing injection so we should strictly skip the sheets.
    
    for sheet_name in wb.sheetnames:
        # st.write(f"Checking sheet: {sheet_name}") # Debug
        
        # Start Condition - RELAXED for specific apps
        # If Kaffa/Mapping/Switching are BEFORE Bentley, they must be processed.
        bypass = False
        if any(x in sheet_name.lower() for x in ["kaffa", "mapping", "switching"]): bypass = True
        
        if not start_found and not bypass:
            if target_start_name.lower() in sheet_name.lower(): start_found = True
        
        if not start_found and not bypass: continue
            
        ws = wb[sheet_name]
        
        # Ignore List (Junk Tabs)
        ignore_list = ['Meetings', 'OPCOS', 'To Delete', 'Template', 'Questions Template', 'Sheet20', '_MASTER_TEMPLATE_', 'Sheet1']
        if any(x.lower() in sheet_name.lower() for x in ignore_list): continue

        # Header Scan
        q_col=a_col=s_col=-1; h_row=-1; found_headers = False
        for r in range(1, 25): # Increased scan depth
            row_vals = [str(ws.cell(r,c).value or "").lower() for c in range(1, 30)]
            for c, v in enumerate(row_vals, 1):
                if ("question" in v or "pergunta" in v or "questão" in v or "item" in v) and "number" not in v: q_col=c
                if ("response" in v or "answer" in v or "resposta" in v) and "score" not in v: a_col=c
                if ("score" in v or "evaluation" in v or "pontua" in v or "nota" in v): s_col=c
            if q_col > 0 and a_col > 0: h_row=r; found_headers = True; break
        
        if "kaffa" in sheet_name.lower():
             pass # st.write(f"Kaffa Headers: Q={q_col}, A={a_col}, Row={h_row}")
        
        if not found_headers: continue
            
        clean = sheet_name[:30].strip(); safe=clean
        for ch in "[]:*?/\\'": safe = safe.replace(ch, "")
        is_green = False
        try: 
            if ws.sheet_properties.tabColor and ("00FF00" in str(ws.sheet_properties.tabColor.rgb) or "00B050" in str(ws.sheet_properties.tabColor.rgb)): is_green=True
        except: pass
        
        
        current_answers = {}
        current_q = None
        
        for r in range(h_row+1, ws.max_row+1):
            q_val = ws.cell(r, q_col).value
            a_val = ws.cell(r, a_col).value
            s_val = ws.cell(r, s_col).value if s_col > 0 else None
            q_str = str(q_val or "").strip()
            a_str = str(a_val or "").strip()
            
            if not q_str and not a_str: continue

            # Robust Match: 0.75 cutoff
            match = difflib.get_close_matches(q_str, ALL_QUESTIONS_FLAT, n=1, cutoff=0.75)
            if match:
                current_q = match[0]
                if a_str: current_answers[current_q] = {"a": a_str, "s": s_val}
            else:
                if current_q and not q_str.endswith("?"):
                     current_answers[current_q] = {"a": q_str, "s": s_val}
        
        # MERGE: Base (Manual) <- Overwritten by Excel (Current)
        # final_answers = smart_merge(base_answers, current_answers)
        final_answers = current_answers

        app = {"name": clean, "safe_name": safe, "is_green": is_green, "answers": final_answers}
        apps.append(calculate_score(app))
    
    return apps

def calculate_score(app):
    app["blocks"] = {k: {"qa":[], "raw":0, "cnt":0, "score":1} for k in synergy_blocks}
    
    # Helper for robust lookup
    def get_ans(q_text):
        # 1. Exact
        if q_text in app["answers"]: return app["answers"][q_text]
        # 2. Key Fuzzy Match (in case Manual Keys differ slightly from Master Keys)
        keys = list(app["answers"].keys())
        matches = difflib.get_close_matches(q_text, keys, n=1, cutoff=0.85)
        if matches: return app["answers"][matches[0]]
        return {}

    for category, q_list in MASTER_QUESTIONS.items():
        for q in q_list:
            ans_obj = get_ans(q)
            
            if isinstance(ans_obj, str): ans_obj = {"a": ans_obj, "s": None}
            a_txt = ans_obj.get("a", "")
            s_val = ans_obj.get("s", None)
            
            score = 0
            t_score = evaluate_answer_text(a_txt)
            if t_score <= 2: score = t_score 
            else:
                 try: 
                    if s_val and float(s_val) >= 1: score = float(s_val)
                    else: score = t_score
                 except: score = t_score
            if len(a_txt) < 2: score = 0
            
            app["blocks"][category]["qa"].append({"q": q, "a": a_txt, "s": score})
            if score > 0:
                app["blocks"][category]["raw"] += score; app["blocks"][category]["cnt"] += 1

    for k, b in app["blocks"].items():
        if b["cnt"] > 0:
            sc = int(round(b["raw"] / b["cnt"]))
            if sc < 1: sc=1; 
            if sc > 5: sc=5; 
            b["score"] = sc
        else: b["score"] = 1
    return app

def generate_report(wb, apps):
    # 1. DELETE ALL EXISTING TABS from uploaded file (they will be recreated)
    # Keep only the first sheet temporarily to avoid empty workbook error
    all_sheets = list(wb.sheetnames)
    for sheet_name in all_sheets[1:]:  # Keep first sheet temporarily
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    
    # 2. Copy Reference Tabs from v52 Template (WITH IMAGES)
    template_path = "/Users/gustavohenriquecastellano/Downloads/Gerador Excel Avandrig/Avangrid_Application_Portfolio_Management_v52.xlsx"
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from copy import copy
        wb_template = load_workbook(template_path)
        
        # Copy Index, Methodology, Introduction tabs
        for idx, tab_name in enumerate(["Index", "Introduction", "Methodology"]):
            if tab_name in wb_template.sheetnames:
                source_ws = wb_template[tab_name]
                
                # Create new sheet at specific position
                target_ws = wb.create_sheet(tab_name, idx)
                
                # Copy all cells and styles
                for row in source_ws.iter_rows():
                    for cell in row:
                        target_cell = target_ws[cell.coordinate]
                        target_cell.value = cell.value
                        if cell.has_style:
                            target_cell.font = copy(cell.font)
                            target_cell.border = copy(cell.border)
                            target_cell.fill = copy(cell.fill)
                            target_cell.number_format = copy(cell.number_format)
                            target_cell.protection = copy(cell.protection)
                            target_cell.alignment = copy(cell.alignment)
                
                # Copy merged cells
                for merged_cell_range in source_ws.merged_cells.ranges:
                    target_ws.merge_cells(str(merged_cell_range))
                
                # Copy column dimensions
                for col_letter, col_dim in source_ws.column_dimensions.items():
                    target_ws.column_dimensions[col_letter].width = col_dim.width
                
                # Copy row dimensions
                for row_num, row_dim in source_ws.row_dimensions.items():
                    target_ws.row_dimensions[row_num].height = row_dim.height
                
                # Copy images (if any)
                try:
                    if hasattr(source_ws, '_images'):
                        for img in source_ws._images:
                            # Create a new image from the same image data
                            from openpyxl.drawing.image import Image as XLImage
                            new_image = XLImage(img.ref)
                            # Copy the anchor position
                            if hasattr(img, 'anchor'):
                                new_image.anchor = img.anchor
                            target_ws.add_image(new_image)
                except Exception as img_err:
                    st.warning(f"Could not copy images from {tab_name}: {img_err}")
        
        # Now delete the temporary first sheet
        if all_sheets[0] in wb.sheetnames:
            del wb[all_sheets[0]]
        
        st.success("✅ Copied Index, Methodology, Introduction from v52 template (with images)")
    except Exception as e:
        st.error(f"❌ Error copying template tabs: {e}")
        import traceback
        st.code(traceback.format_exc())

    C_PRIM, C_SEC, C_NEU, C_W = 'E87722', '0066B3', '444444', 'FFFFFF'
    ordered_keys = ["Strategic Fit", "Business Efficiency", "User Value", "Financial Value", "Architecture", "Operational Risk", "Maintainability", "Support Quality"]

    # 2. Calculator
    ws_calc = wb.create_sheet("Calculator"); ws_calc.freeze_panes = "B4"
    ws_calc.column_dimensions['A'].width = 30
    ws_calc.cell(1,1).value="Application Name"; style(ws_calc.cell(1,1), bold=True, bg=C_NEU, color=C_W)
    ws_calc.cell(2,1).value="Weight"; style(ws_calc.cell(2,1), italic=True, align='right')
    col=2; biz_blocks=[b for b,v in synergy_blocks.items() if v["Type"]=="Business"]; tech_blocks=[b for b,v in synergy_blocks.items() if v["Type"]=="Tech"]
    for b_list, color in [(biz_blocks, C_PRIM), (tech_blocks, C_SEC)]:
        for b in b_list:
            c = ws_calc.cell(1, col); c.value=b; style(c, bold=True, bg=color, color=C_W, wrap=True); ws_calc.cell(2, col).value = synergy_blocks[b]["Weight"]; col += 1
    c_bvi=col; ws_calc.cell(1, c_bvi).value="Business Value Index (BVI)"; ws_calc.column_dimensions[get_column_letter(c_bvi)].width=25; style(ws_calc.cell(1, c_bvi), bold=True, wrap=True)
    c_thi=col+1; ws_calc.cell(1, c_thi).value="Technical Health Index (THI)"; ws_calc.column_dimensions[get_column_letter(c_thi)].width=25; style(ws_calc.cell(1, c_thi), bold=True, wrap=True)
    c_dec=col+2; ws_calc.cell(1, c_dec).value="RECOMMENDATION"; style(ws_calc.cell(1, c_dec), bold=True)
    MAX = 100
    for i in range(MAX):
        r = i + 4
        if i < len(apps): app = apps[i]; ws_calc.cell(r,1).value = app["safe_name"] 
        style(ws_calc.cell(r,1), align='left', bg='FAFAFA'); cc = 2
        for b in biz_blocks + tech_blocks:
            b_idx = ordered_keys.index(b); fixed_row = 4 + b_idx; ws_calc.cell(r,cc).value = f'=IF($A{r}="","",IFERROR(INDIRECT("\'"&$A{r}&"\'!$B${fixed_row}"),0))'; style(ws_calc.cell(r,cc), bg='FAFAFA'); cc += 1
        rb_s, rb_e = get_column_letter(2), get_column_letter(1+len(biz_blocks)); rt_s, rt_e = get_column_letter(2+len(biz_blocks)), get_column_letter(1+len(biz_blocks)+len(tech_blocks))
        sum_rb = f"SUM({rb_s}{r}:{rb_e}{r})"; ws_calc.cell(r, c_bvi).value = f'=IF(OR($A{r}="", {sum_rb}=0),NA(),IFERROR(SUMPRODUCT({rb_s}{r}:{rb_e}{r}, ${rb_s}$2:${rb_e}$2)/SUM(${rb_s}$2:${rb_e}$2)*20, 0))'
        sum_rt = f"SUM({rt_s}{r}:{rt_e}{r})"; ws_calc.cell(r, c_thi).value = f'=IF(OR($A{r}="", {sum_rt}=0),NA(),IFERROR(SUMPRODUCT({rt_s}{r}:{rt_e}{r}, ${rt_s}$2:${rt_e}$2)/SUM(${rt_s}$2:${rt_e}$2)*20, 0))'
        cb, ct = f"{get_column_letter(c_bvi)}{r}", f"{get_column_letter(c_thi)}{r}"
        ws_calc.cell(r, c_dec).value = f'=IF(ISNA({cb}),"",IF(AND({cb}>=60,{ct}>=60),"EVOLVE",IF(AND({cb}>=60,{ct}<60),"INVEST",IF(AND({cb}<60,{ct}>=60),"MAINTAIN","ELIMINATE"))))'
    rng_dec = f"{get_column_letter(c_dec)}4:{get_column_letter(c_dec)}{MAX+3}"
    ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="EVOLVE"'], fill=PatternFill(bgColor='C6EFCE')))
    ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="INVEST"'], fill=PatternFill(bgColor='FFEB9C')))
    ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="MAINTAIN"'], fill=PatternFill(bgColor='BDD7EE')))
    ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="ELIMINATE"'], fill=PatternFill(bgColor='FFC7CE')))
    
    # Color formatting for Affinity Block Scores (0-5 scale)
    # Apply to all block columns (Business + Tech)
    last_block_col = 1 + len(biz_blocks) + len(tech_blocks)
    for col in range(2, last_block_col + 1):
        rng_block = f"{get_column_letter(col)}4:{get_column_letter(col)}{MAX+3}"
        # Red for 0-2, Yellow for 3, Green for 4-5
        ws_calc.conditional_formatting.add(rng_block, FormulaRule(formula=[f'{get_column_letter(col)}4<=2'], fill=PatternFill(bgColor='FFC7CE')))
        ws_calc.conditional_formatting.add(rng_block, FormulaRule(formula=[f'{get_column_letter(col)}4=3'], fill=PatternFill(bgColor='FFEB9C')))
        ws_calc.conditional_formatting.add(rng_block, FormulaRule(formula=[f'{get_column_letter(col)}4>=4'], fill=PatternFill(bgColor='C6EFCE')))

    # Create Hidden Data for Quadrant Lines (at 60, 60)
    # H-Line: (0,60) -> (100,60)
    ws_calc.cell(1, 15).value = "Q_X1"; ws_calc.cell(1, 16).value = "Q_Y1" # O, P
    ws_calc.cell(2, 15).value = 0; ws_calc.cell(2, 16).value = 60
    ws_calc.cell(3, 15).value = 100; ws_calc.cell(3, 16).value = 60
    
    # V-Line: (60,0) -> (60,100)
    ws_calc.cell(1, 17).value = "Q_X2"; ws_calc.cell(1, 18).value = "Q_Y2" # Q, R
    ws_calc.cell(2, 17).value = 60; ws_calc.cell(2, 18).value = 0
    ws_calc.cell(3, 17).value = 60; ws_calc.cell(3, 18).value = 100
    
    # 3. Dashboard
    ws_dash = wb.create_sheet("Dashboard"); ch = ScatterChart(); ch.style=13
    ch.x_axis.title="Technical Health Index (THI)"; ch.y_axis.title="Business Value Index (BVI)"
    ch.x_axis.scaling.min=0; ch.x_axis.scaling.max=100
    ch.y_axis.scaling.min=0; ch.y_axis.scaling.max=100
    
    # QUADRANT LINES (Axes crossing at 60) - FIXED: Use Dummy Series instead of Axis Cross to keep Labels Visible
    sp_pr = GraphicalProperties(ln=LineProperties(prstDash='sysDot', w=9525, solidFill='D9D9D9'))
    ch.x_axis.majorGridlines = ChartLines(spPr=sp_pr); ch.y_axis.majorGridlines = ChartLines(spPr=sp_pr)
    
    # Hide axis tick labels (no numbers)
    ch.x_axis.tickLblPos = "none"
    ch.y_axis.tickLblPos = "none"
    ch.x_axis.delete = False  # Keep axis visible
    ch.y_axis.delete = False
    ch.x_axis.majorUnit = 20  # Gridlines every 20 units
    ch.y_axis.majorUnit = 20
    
    # Restore Main Axis Lines (Black)
    black_line = GraphicalProperties(ln=LineProperties(w=12700, solidFill='000000'))
    ch.x_axis.spPr = black_line; ch.y_axis.spPr = black_line
    
    # Add Quadrant Lines Series
    # H-Line
    s_h = Series(values=Reference(ws_calc, min_col=16, min_row=2, max_row=3), xvalues=Reference(ws_calc, min_col=15, min_row=2, max_row=3), title="ref_h")
    s_h.marker.symbol = "picture"; s_h.graphicalProperties.line.solidFill = "000000"; s_h.graphicalProperties.line.width = 12700 # 1pt
    ch.series.append(s_h)
    
    # V-Line
    s_v = Series(values=Reference(ws_calc, min_col=18, min_row=2, max_row=3), xvalues=Reference(ws_calc, min_col=17, min_row=2, max_row=3), title="ref_v")
    s_v.marker.symbol = "picture"; s_v.graphicalProperties.line.solidFill = "000000"; s_v.graphicalProperties.line.width = 12700
    ch.series.append(s_v)

    ch.legend = None; ch.width = 45; ch.height = 25
    colors = {'EVOLVE':('C6EFCE','006100'), 'INVEST':('FFEB9C','9C5700'), 'MAINTAIN':('BDD7EE','0066CC'), 'ELIMINATE':('FFC7CE','9C0006')}
    coord_map = {}; BUCKET = 3 
    for i, app in enumerate(apps):
        r = 4 + i
        rb = sum(app["blocks"][k]["score"] for k in ordered_keys if "Fit" in k or "Efficiency" in k or "User" in k or "Financial" in k) / 4.0 * 20
        rt = sum(app["blocks"][k]["score"] for k in ordered_keys if "Architecture" in k or "Risk" in k or "Maintain" in k or "Support" in k) / 4.0 * 20
        rec = "ELIMINATE"
        if rb >= 60 and rt >= 60: rec="EVOLVE"
        elif rb >= 60 and rt < 60: rec="INVEST"
        elif rb < 60 and rt >= 60: rec="MAINTAIN" 
        x_b = int(rt // BUCKET); y_b = int(rb // BUCKET); k_coord = (x_b, y_b); count = coord_map.get(k_coord, 0); coord_map[k_coord] = count + 1
        fill, line = colors[rec]
        if i < MAX:
            # Use calculator reference
            x_r = Reference(ws_calc, min_col=c_thi, min_row=r, max_row=r); y_r = Reference(ws_calc, min_col=c_bvi, min_row=r, max_row=r)
            s = Series(values=y_r, xvalues=x_r, title=app["safe_name"])
            s.marker.symbol = 'circle'; s.marker.size = 10
            s.marker.graphicalProperties.solidFill = fill
            s.marker.graphicalProperties.line.solidFill = line
            
            # Add data labels with smart positioning and smaller font
            s.dLbls = DataLabelList()
            s.dLbls.showSerName = True
            s.dLbls.showVal = False
            s.dLbls.showCatName = False
            s.dLbls.showLegendKey = False
            
            # Set smaller font size to reduce overlap
            from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
            from openpyxl.chart.text import RichText
            txt_props = CharacterProperties(sz=600)  # 6pt font (600 = 6pt * 100)
            s.dLbls.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=txt_props))])
            
            # Position cycling to reduce overlap
            pos = 't'
            if count > 0:
                cycle = ['b', 'r', 'l', 't']
                pos = cycle[count % 4]
            
            # Manual overrides for specific apps
            for k_man in MANUAL_LAYOUT:
                if k_man in app["safe_name"]:
                    pos = MANUAL_LAYOUT[k_man]
                    break
            
            s.dLbls.position = pos
            ch.series.append(s)
    
    ws_dash.add_chart(ch, "B2")

    # 4. Roadmap
    ws_strat = wb.create_sheet("Strategic Roadmap"); ws_strat.freeze_panes = "A2"
    heads = ["Application Name", "Business Value Index (BVI)", "Technical Health Index (THI)", "Recommendation", "Subcategory", "Quick Win?", "Priority Rule", "Rationale", "Comments"]
    for idx, h in enumerate(heads): ws_strat.cell(1, idx+1).value=h; style(ws_strat.cell(1, idx+1), bold=True, bg=C_NEU, color=C_W, wrap=True)
    if idx==1 or idx==2: ws_strat.column_dimensions[get_column_letter(idx+1)].width=25
    ws_strat.column_dimensions['A'].width=35; ws_strat.column_dimensions['E'].width=20
    c_cfg = 12; LISTS = {"ELIMINATE": ["Replace", "Retire", "Absorbed"], "INVEST": ["Absorb"], "EVOLVE": ["Modernize", "Enhance", "Migrate", "Refactor", "Upgrade"], "MAINTAIN": ["Internalize", "Maintain"]}
    for key, items in LISTS.items():
        ws_strat.cell(1, c_cfg).value = key; 
        for idx, item in enumerate(items): ws_strat.cell(2+idx, c_cfg).value = item
        defn = DefinedName(f"LIST_{key}", attr_text=f"'Strategic Roadmap'!${get_column_letter(c_cfg)}$2:${get_column_letter(c_cfg)}{1+len(items)}"); wb.defined_names.add(defn); c_cfg += 1
    ws_strat.cell(1, c_cfg).value = "QWIN"; ws_strat.cell(2, c_cfg).value="Yes"; ws_strat.cell(3, c_cfg).value="No"; ws_strat.cell(4, c_cfg).value="Review"; 
    defn_qw = DefinedName("LIST_QWIN", attr_text=f"'Strategic Roadmap'!${get_column_letter(c_cfg)}$2:${get_column_letter(c_cfg)}{4}"); wb.defined_names.add(defn_qw); c_cfg += 1
    c_prio = c_cfg + 1; ws_strat.cell(1, c_prio).value = "Matrix Config"; 
    cfg_rows = [("ELIMINATE", "Replace", "P1 - Critical", "High Risk / EOL."), ("ELIMINATE", "Retire", "P1 - Critical", "Decommission"), ("ELIMINATE", "Absorbed", "P2 - Tactical", "Consolidation"), ("INVEST", "Absorb", "P1 - Critical", "High Value Opportunity"), ("EVOLVE", "Modernize", "P1 - Critical", "Transformation"), ("EVOLVE", "Migrate", "P1 - Critical", "Platform shift"), ("EVOLVE", "Enhance", "P2 - Strategic", "Expansion"), ("EVOLVE", "Refactor", "P2 - Strategic", "Code quality"), ("EVOLVE", "Upgrade", "P2 - Strategic", "Version"), ("MAINTAIN", "Internalize", "P2 - Compliance", "Governance"), ("MAINTAIN", "Maintain", "P3 - Routine", "Keep lights on")]
    ws_strat.column_dimensions[get_column_letter(c_prio)].hidden = True
    for i, h in enumerate(["Key", "Decision", "Subcategory", "Priority Rule", "Rationale"]): style(ws_strat.cell(2, c_prio+i), bold=True, bg='333333', color='FFFFFF', border=True); ws_strat.cell(2, c_prio+i).value=h
    for i, (d, s, p, rat) in enumerate(cfg_rows):
        r=i+3; ws_strat.cell(r, c_prio).value = d+s; ws_strat.cell(r, c_prio+1).value = d; ws_strat.cell(r, c_prio+2).value = s; ws_strat.cell(r, c_prio+3).value = p; ws_strat.cell(r, c_prio+4).value = rat
    defn_mtx = DefinedName("MatrixConfig", attr_text=f"'Strategic Roadmap'!${get_column_letter(c_prio)}$3:${get_column_letter(c_prio+4)}{len(cfg_rows)+3}"); wb.defined_names.add(defn_mtx)
    
    # 5b. POPULATE ROADMAP
    for i, app in enumerate(apps):
        r = i + 2
        # A: Name
        ws_strat.cell(r, 1).value = app["safe_name"]
        # B: BVI =Calculator! (Col 10 = J)
        ws_strat.cell(r, 2).value = f"=IFERROR(VLOOKUP(A{r},Calculator!A:L,10,FALSE),0)" 
        # C: THI =Calculator! (Col 11 = K)
        ws_strat.cell(r, 3).value = f"=IFERROR(VLOOKUP(A{r},Calculator!A:L,11,FALSE),0)"
        # D: Rec =Calculator! (Col 12 = L)
        ws_strat.cell(r, 4).value = f"=IFERROR(VLOOKUP(A{r},Calculator!A:L,12,FALSE),\"\")"
        
        # Validation for Subcategory
        dv_sub = DataValidation(type="list", formula1=f"=INDIRECT(\"LIST_\"&D{r})", allow_blank=True)
        ws_strat.add_data_validation(dv_sub); dv_sub.add(ws_strat.cell(r, 5))
        
        # Validation for QWIN
        dv_qw = DataValidation(type="list", formula1="=LIST_QWIN", allow_blank=True)
        ws_strat.add_data_validation(dv_qw); dv_qw.add(ws_strat.cell(r, 6))

        # Priority & Rationale (Lookup from MatrixConfig based on Rec+Subcat)
        ws_strat.cell(r, 7).value = f"=IFERROR(VLOOKUP(D{r}&E{r},MatrixConfig,4,FALSE),\"\")"
        ws_strat.cell(r, 8).value = f"=IFERROR(VLOOKUP(D{r}&E{r},MatrixConfig,5,FALSE),\"\")"
    
    # Apply conditional formatting to Strategic Roadmap (Recommendation column only)
    max_row_strat = len(apps) + 1
    # Recommendation column (D)
    rng_rec_strat = f"D2:D{max_row_strat}"
    ws_strat.conditional_formatting.add(rng_rec_strat, FormulaRule(formula=['D2="EVOLVE"'], fill=PatternFill(bgColor='C6EFCE')))
    ws_strat.conditional_formatting.add(rng_rec_strat, FormulaRule(formula=['D2="INVEST"'], fill=PatternFill(bgColor='FFEB9C')))
    ws_strat.conditional_formatting.add(rng_rec_strat, FormulaRule(formula=['D2="MAINTAIN"'], fill=PatternFill(bgColor='BDD7EE')))
    ws_strat.conditional_formatting.add(rng_rec_strat, FormulaRule(formula=['D2="ELIMINATE"'], fill=PatternFill(bgColor='FFC7CE')))

    # 5. GENERATE APP TABS (LAST, so they appear at end of file)
    for app in apps:
        ws = wb.create_sheet(app["safe_name"]); ws.sheet_properties.tabColor = "00B050" if app["is_green"] else "D9D9D9"
        ws.column_dimensions['A'].width=50 
        ws.column_dimensions['B'].width=80 
        ws.merge_cells("A1:B1"); ws.cell(1,1).value=f"Assessment: {app['name']}"; style(ws.cell(1,1), bold=True, size=14, color=C_PRIM)
        ws.cell(3,1).value = "EXECUTIVE SCORECARD"; style(ws.cell(3,1), bold=True, bg='333333', color='FFFFFF')
        ws.cell(3,2).value = "SCORE (0-5)"; style(ws.cell(3,2), bold=True, bg='333333', color='FFFFFF')
        
        for i, k in enumerate(ordered_keys):
            r = 4 + i
            ws.cell(r, 1).value = k; style(ws.cell(r, 1), bold=True, align='right', bg='FAFAFA')
            sc_cell = ws.cell(r, 2); sc_cell.value = app["blocks"][k]["score"]; style(sc_cell, bg='FFFFFF', border=True)
            dv = DataValidation(type="list", formula1='"0,1,2,3,4,5"', allow_blank=True); ws.add_data_validation(dv); dv.add(sc_cell)
            
        curr = 14
        for b_key in ordered_keys:
            b = app["blocks"][b_key]
            ws.merge_cells(f"A{curr}:B{curr}"); ws.cell(curr,1).value=b_key.upper(); style(ws.cell(curr,1), bold=True, bg=C_NEU, color=C_W)
            curr+=1; ws.cell(curr,1).value="Definitions"; style(ws.cell(curr,1), bold=True); curr+=1
            for scr in range(1,6): ws.cell(curr,1).value=f"{scr} - {synergy_blocks[b_key]['Defs'].get(scr,'')}"; style(ws.cell(curr,1), align='left'); curr+=1
            ws.cell(curr,1).value="Q"; style(ws.cell(curr,1), bold=True, bg=C_NEU, color=C_W); ws.cell(curr,2).value="A"; style(ws.cell(curr,2), bold=True, bg=C_NEU, color=C_W); curr+=1
            for qa in b["qa"]: 
                ws.cell(curr,1).value=qa["q"]; style(ws.cell(curr,1), bg='FAFAFA', align='left', wrap=True)
                ws.cell(curr,2).value=qa["a"]; style(ws.cell(curr,2), bg='FAFAFA', align='left', wrap=True)
                curr+=1
            curr+=1

    # 6. Update INDEX - Find and populate existing structure
    ws_idx = None
    for s in wb.sheetnames:
        if "index" in s.lower():
            ws_idx = wb[s]; break
    
    if ws_idx:
        # Clean up obsolete tabs from Index (User Guide, Master Template)
        # Scan cols A, B, C to be sure
        for row in range(ws_idx.max_row, 1, -1):
            values = [str(ws_idx.cell(row, c).value or "").lower() for c in range(1, 4)]
            row_text = " ".join(values)
            if "user guide" in row_text or "master_template" in row_text or "master template" in row_text or "questions template" in row_text:
                # st.write(f"Deleting row {row}: {row_text}") # Debug
                ws_idx.delete_rows(row)

        # Find "Generated Tabs:" row - ROBUST SEARCH
        gen_tabs_row = None
        app_assess_row = None
        
        # Scan A-C for target headers
        for row in range(1, ws_idx.max_row + 1):
             # Check A, B, C
             row_vals = [str(ws_idx.cell(row, c).value or "").lower().strip() for c in range(1, 4)]
             row_str = " ".join(row_vals)
             
             if "generated tabs" in row_str or "relatórios gerados" in row_str:
                 gen_tabs_row = row
             if "application assessment" in row_str or "avaliação de aplicações" in row_str:
                 app_assess_row = row
        
        # Fallback: If not found, append to end
        if not gen_tabs_row:
             gen_tabs_row = ws_idx.max_row + 2
             ws_idx.cell(gen_tabs_row, 1).value = "Generated Tabs:"
             style(ws_idx.cell(gen_tabs_row, 1), bold=True, bg='333333', color='FFFFFF')

        if not app_assess_row:
             app_assess_row = gen_tabs_row + 5 # Space out
             ws_idx.cell(app_assess_row, 1).value = "Application Assessments:"
             style(ws_idx.cell(app_assess_row, 1), bold=True, bg='333333', color='FFFFFF')
             
        # Add links to Generated Tabs section
        r = gen_tabs_row + 1
        for tab_name in ["Calculator", "Dashboard", "Strategic Roadmap"]:
            ws_idx.cell(r, 1).value = tab_name
            ws_idx.cell(r, 1).hyperlink = f"#'{tab_name}'!A1"
            ws_idx.cell(r, 1).font = Font(color='0563C1', underline='single')
            ws_idx.cell(r, 2).value = "Strategic Analysis View" # Add description
            r += 1
        
        # Add links to Application Assessments section
        r = app_assess_row + 1
        for app in apps:
            ws_idx.cell(r, 1).value = app["safe_name"]
            ws_idx.cell(r, 1).hyperlink = f"#'{app['safe_name']}'!A1"
            ws_idx.cell(r, 1).font = Font(color='0563C1', underline='single')
            ws_idx.cell(r, 2).value = "Detailed App Scorecard"
            r += 1
    
    # 6. Generate "Application Groups" Tab (Formerly Analysis)
    # 7. Generate "Value Chain" Tab
    
    # --- HELPER: KEYWORD SCORING ---
    def get_category_scores(app_data, keywords_dict):
        # app_data is expected to be the full app object
        # We need to scan answers for context.
        # Concatenate all answers into big text blob
        full_text = app_data["safe_name"].lower() + " "
        if "answers" in app_data:
            for q, ans_obj in app_data["answers"].items():
                if isinstance(ans_obj, dict):
                    full_text += str(ans_obj.get("a", "")).lower() + " "
                else: 
                     full_text += str(ans_obj).lower() + " "
        
        scores = {}
        for cat, kws in keywords_dict.items():
            score = 0
            for k in kws:
                if k in full_text: score += 1
            scores[cat] = score
        return scores

    # --- TAB 1: APPLICATION GROUPS (Functional) ---
    ws_groups = wb.create_sheet("Application Groups")
    ws_groups.sheet_properties.tabColor = "0000FF"
    
    # Expanded Categories
    GROUP_CATEGORIES = {
        "Task Management / User Alignment": ["task", "schedule", "track", "assign", "user", "alignment", "collab", "arcos", "jums", "switching", "ppe"],
        "Maintenance & Asset Mgmt": ["maintain", "asset", "work order", "inspection", "repair", "lifecycle", "bentley", "cimplicity", "cathodic", "poleforeman", "esosr"],
        "Grid Operations / Engineering": ["scada", "dms", "oms", "real-time", "grid", "voltage", "design", "model", "calculate", "engineer", "kaffa", "scal", "dsd"],
        "Document / Info Management": ["document", "file", "map", "drawing", "view", "repository", "knowledge", "projectwise", "mapping"],
        "Corporate / Administrative": ["finance", "hr", "legal", "compliance", "security", "supply chain", "admin", "erp", "sap"]
    }
    
    # Assign apps to groups (Multi-select allowed)
    group_data = {k: [] for k in GROUP_CATEGORIES}
    group_data["Uncategorized"] = []
    
    for app in apps:
        scores = get_category_scores(app, GROUP_CATEGORIES)
        # Threshold: >0 matches. If Name matches, big boost.
        
        assigned = False
        # Special check: If name contains category keywords, force assign
        for cat, kws in GROUP_CATEGORIES.items():
            if any(k in app["safe_name"].lower() for k in kws):
                scores[cat] += 5 # Boost
        
        # Assign to all with score >= 1 (or top N? User said "considering all answers")
        # Let's assign to any category with score >= 1, but if no score, then Uncategorized.
        # To avoid noise, let's say score >= 2 unless it's a name match (score >= 5)
        
        best_score = max(scores.values()) if scores else 0
        
        if best_score > 0:
             for cat, sc in scores.items():
                 # Rule: Assign if score is high relative to others or absolute high
                 if sc >= 5 or (sc >= 1 and sc >= best_score * 0.5):
                     # Calculate color
                     rb = sum(app["blocks"][k]["score"] for k in ordered_keys if "Fit" in k or "Efficiency" in k or "User" in k or "Financial" in k) / 4.0 * 20
                     rt = sum(app["blocks"][k]["score"] for k in ordered_keys if "Architecture" in k or "Risk" in k or "Maintain" in k or "Support" in k) / 4.0 * 20
                     bg_c = "FFFFFF"
                     if rb >= 60 and rt >= 60: bg_c="C6EFCE" 
                     elif rb >= 60 and rt < 60: bg_c="FFEB9C"
                     elif rb < 60 and rt >= 60: bg_c="BDD7EE"
                     else: bg_c="FFC7CE"
                     
                     group_data[cat].append({"name": app["safe_name"], "color": bg_c})
                     assigned = True
        
        if not assigned:
            rb = sum(app["blocks"][k]["score"] for k in ordered_keys if "Fit" in k or "Efficiency" in k or "User" in k or "Financial" in k) / 4.0 * 20
            rt = sum(app["blocks"][k]["score"] for k in ordered_keys if "Architecture" in k or "Risk" in k or "Maintain" in k or "Support" in k) / 4.0 * 20
            bg_c = "FFFFFF"
            if rb >= 60 and rt >= 60: bg_c="C6EFCE" 
            elif rb >= 60 and rt < 60: bg_c="FFEB9C"
            elif rb < 60 and rt >= 60: bg_c="BDD7EE"
            else: bg_c="FFC7CE"
            group_data["Uncategorized"].append({"name": app["safe_name"], "color": bg_c})

    # Render Groups (Grid 2x3 or 3x2)
    # Layout Config: (Row, Col)
    LAYOUT = [
        ("Task Management / User Alignment", 3, 2, "E87722"),
        ("Maintenance & Asset Mgmt", 3, 7, "0066B3"),
        ("Grid Operations / Engineering", 3, 12, "00B050"),
        ("Document / Info Management", 25, 2, "7030A0"),
        ("Corporate / Administrative", 25, 7, "FFC000"),
        ("Uncategorized", 25, 12, "999999")
    ]
    
    ws_groups.column_dimensions['B'].width = 30; ws_groups.column_dimensions['G'].width = 30; ws_groups.column_dimensions['L'].width = 30
    
    for (name, r, c, color) in LAYOUT:
        if name not in group_data and name != "Uncategorized": continue
        items = group_data.get(name, [])
        
        hc = ws_groups.cell(r, c)
        ws_groups.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+3)
        hc.value = name
        style(hc, bold=True, size=12, bg=color, color="FFFFFF")
        
        curr = r + 1
        # Grey Box
        for rb in range(curr, curr+15):
            for cb in range(c, c+4): ws_groups.cell(rb, cb).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
        for item in items:
            cell = ws_groups.cell(curr, c)
            cell.value = item["name"]
            style(cell, size=11, bg=item["color"], border=True, align='left')
            ws_groups.merge_cells(start_row=curr, start_column=c, end_row=curr, end_column=c+3)
            curr += 1

    # --- TAB 2: VALUE CHAIN (Refined for Avangrid) ---
    ws_chain = wb.create_sheet("Value Chain")
    ws_chain.sheet_properties.tabColor = "FF9900" # Orange
    
    # Detailed Avangrid Chain
    CHAIN_STAGES = {
        "Generation (Renewables)": ["generation", "renewable", "wind", "solar", "hydro", "offshore", "onshore", "plant", "turbine", "energy source", "production"],
        "Transmission (Transport)": ["transmission", "high voltage", "substation", "interconnection", "control center", "ecc", "tcc", "line", "relay", "protection"],
        "Distribution (Delivery)": ["distribution", "medium voltage", "low voltage", "ami", "smart grid", "meter", "outage", "oms", "dms", "scada", "field", "pole", "circuit"],
        "Customer Solutions": ["customer", "billing", "crm", "call center", "service", "payment", "meter to cash", "der", "ev charging", "portal", "account"],
        "Corporate / Shared Services": ["finance", "hr", "legal", "compliance", "security", "supply chain", "admin", "it", "procurement", "cybersecurity", "ehs"]
    }
    
    chain_data = {k: [] for k in CHAIN_STAGES}
    chain_data["Cross-Cutting"] = []
    
    for app in apps:
        scores = get_category_scores(app, CHAIN_STAGES)
        best_cat = "Cross-Cutting"
        high_score = 0
        
        for cat, sc in scores.items():
            if sc > high_score:
                high_score = sc
                best_cat = cat
        
        # Color calculation
        rb = sum(app["blocks"][k]["score"] for k in ordered_keys if "Fit" in k or "Efficiency" in k or "User" in k or "Financial" in k) / 4.0 * 20
        rt = sum(app["blocks"][k]["score"] for k in ordered_keys if "Architecture" in k or "Risk" in k or "Maintain" in k or "Support" in k) / 4.0 * 20
        bg_c = "FFFFFF"
        if rb >= 60 and rt >= 60: bg_c="C6EFCE" 
        elif rb >= 60 and rt < 60: bg_c="FFEB9C"
        elif rb < 60 and rt >= 60: bg_c="BDD7EE"
        else: bg_c="FFC7CE"
        
        chain_data[best_cat].append({"name": app["safe_name"], "color": bg_c})
        
    # Render Value Chain - Linear Flow Layout
    # Layout: Gen (B) -> Trans (G) -> Dist (L) -> Cust (Q)
    # Corporate at Bottom
    
    CHAIN_LAYOUT = [
        ("Generation (Renewables)", 2, "00B0F0"),     # Blue
        ("Transmission (Transport)", 7, "0070C0"),    # Darker Blue
        ("Distribution (Delivery)", 12, "00B050"),    # Green
        ("Customer Solutions", 17, "FFC000")          # Orange
    ]
    
    ws_chain.merge_cells("B2:T2"); h = ws_chain["B2"]; h.value = "AVANGRID INTEGRATED UTILITY VALUE CHAIN"; style(h, bold=True, size=16, bg="333333", color="FFFFFF")
    
    # Process Flow Arrows (Headers)
    for i, (name, col, color) in enumerate(CHAIN_LAYOUT):
        # Header Box
        ws_chain.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+3)
        h = ws_chain.cell(4, col)
        h.value = name
        style(h, bold=True, size=12, bg=color, color="FFFFFF")
        
        # Data
        items = chain_data.get(name, [])
        curr = 5
        ws_chain.column_dimensions[get_column_letter(col)].width=25
        
        for item in items:
            cell = ws_chain.cell(curr, col)
            cell.value = item["name"]
            style(cell, size=11, bg=item["color"], border=True, align='center')
            ws_chain.merge_cells(start_row=curr, start_column=col, end_row=curr, end_column=col+3)
            curr += 1
            
        # Draw "Arrow" connector if not last
        if i < len(CHAIN_LAYOUT) - 1:
            # Simple right arrow in column between blocks usually works but we have gap (col+4 is empty)
            # Col F (6), K (11), P (16) are gaps
            gap_col = col + 4
            ws_chain.cell(4, gap_col).value = "➜"
            style(ws_chain.cell(4, gap_col), size=14, bold=True, align='center')

    # Corporate / Shared Services (Foundation)
    corp_name = "Corporate / Shared Services"
    ws_chain.merge_cells("B20:T20"); h = ws_chain["B20"]; h.value = corp_name; style(h, bold=True, bg="7030A0", color="FFFFFF")
    
    items = chain_data.get(corp_name, [])
    curr = 21
    # Distribute horizontally across the bottom
    c_corp = 2
    for item in items:
        if c_corp > 17: c_corp = 2; curr += 1
        cell = ws_chain.cell(curr, c_corp)
        cell.value = item["name"]
        style(cell, size=11, bg=item["color"], border=True)
        ws_chain.merge_cells(start_row=curr, start_column=c_corp, end_row=curr, end_column=c_corp+3)
        c_corp += 5

    # Cross Cutting / Enterprise (Below Corporate)
    if chain_data["Cross-Cutting"]:
        curr = max(curr, 25) + 2
        ws_chain.merge_cells(f"B{curr}:T{curr}"); h = ws_chain[f"B{curr}"]; h.value = "Enterprise Broad / Cross-Cutting"; style(h, bold=True, bg="999999", color="FFFFFF")
        curr += 1
        c_cc = 2
        for item in chain_data["Cross-Cutting"]:
            if c_cc > 17: c_cc = 2; curr += 1
            cell = ws_chain.cell(curr, c_cc)
            cell.value = item["name"]
            style(cell, size=11, bg=item["color"], border=True)
            ws_chain.merge_cells(start_row=curr, start_column=c_cc, end_row=curr, end_column=c_cc+3)
            c_cc += 5

    # 8. Update INDEX and REORDER TABS
    # We want exact order: 
    # [Start Tabs...] -> Dashboard -> Strategic Roadmap -> Application Groups -> Value Chain -> [App Tabs...]
    
    SPECIAL_TABS = ["Introduction", "Index", "Methodology", "User Guide", "_MASTER_TEMPLATE_", "Calculator", "Dashboard", "Strategic Roadmap", "Application Groups", "Value Chain"]
    
    # Reorder workbook tabs
    all_sheets = wb.sheetnames
    new_order = []
    
    # Add Special Tabs in order if they exist
    for t in SPECIAL_TABS:
        if t in all_sheets: new_order.append(wb[t])
        
    # Add remaining tabs (App tabs)
    for t in all_sheets:
        if t not in SPECIAL_TABS:
            new_order.append(wb[t])
            
    wb._sheets = new_order
    
    if ws_idx:
        # Find position to insert links in Index
        target_row = -1
        # Clear old links to be safe or append
        # Let's rebuild the links section starting from "Strategic Roadmap"
        
        # Search for header "Generated Tabs:"
        gen_row = -1
        for r in range(1, 100):
            if "Generated Tabs" in str(ws_idx.cell(r,1).value): gen_row = r; break
        
        if gen_row > 0:
            # Overwrite links below header
            r = gen_row + 1
            links = [
                ("Calculator", "Scoring Engine"),
                ("Dashboard", "Strategic Scatter Plot"),
                ("Strategic Roadmap", "Action Plan (P1/P2/P3)"),
                ("Application Groups", "Functional Categorization"),
                ("Value Chain", "Avangrid Integrated Value Chain")
            ]
            for name, desc in links:
                if name in wb.sheetnames:
                    ws_idx.cell(r, 1).value = name
                    ws_idx.cell(r, 1).hyperlink = f"#'{name}'!A1"
                    ws_idx.cell(r, 1).font = Font(color='0563C1', underline='single')
                    ws_idx.cell(r, 2).value = desc
                    r += 1

    return wb

def update_index_with_analysis(ws_idx):
    # Find "Generated Tabs:" again or just verify
    # Add Analysis link at the top of generated tabs list if possible, or append
    # Just append to the list of tabs row
    pass # Already handled by appending logic in previous steps if we added it to correct list, 
         # but actually we need to insert it specifically or just let user find it.
         # Let's add it explicitly to the "Generated Tabs" section we created/found.
    
    # We can search for "Calculator" and insert before/after
    for r in range(1, ws_idx.max_row+1):
        if str(ws_idx.cell(r,1).value) == "Calculator":
            # Insert row before
            ws_idx.insert_rows(r)
            ws_idx.cell(r, 1).value = "Analysis"
            ws_idx.cell(r, 1).hyperlink = "#'Analysis'!A1"
            ws_idx.cell(r, 1).font = Font(color='0563C1', underline='single')
            ws_idx.cell(r, 2).value = "Application Quadrant Analysis"
            break

# --- APP FLOW ---
uploaded_file = st.file_uploader("Upload 'Application questionnaire.xlsx'", type=["xlsx"])

if uploaded_file:
    st.success("File Uploaded! Analyzing tabs...")
    
    if st.button("🚀 Run AI Analysis"):
        with st.spinner("Parsing tabs, scoring answers, and generating dashboard..."):
            wb_in = openpyxl.load_workbook(uploaded_file, data_only=True)
            # COPY WB for Output (Preserve Tabs)
            wb_in.save("temp_source.xlsx")
            wb_out = openpyxl.load_workbook("temp_source.xlsx") 
            
            apps = parse_wb(wb_in)
            
            st.info(f"Processed {len(apps)} applications from upload.")
            
            # Generate (Pass wb_out to preserve tabs)
            wb_final = generate_report(wb_out, apps)
            
            # Save to Bytes
            out_buffer = io.BytesIO()
            wb_final.save(out_buffer)
            out_buffer.seek(0)
            
            st.success("Analysis Complete!")
            st.download_button(
                label="📥 Download Strategic Portfolio (.xlsx)",
                data=out_buffer,
                file_name="Avangrid_Application_Portfolio_Management_v58.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
