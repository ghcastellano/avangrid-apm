import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.workbook.defined_name import DefinedName
import os
import shutil
import math
import difflib
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.drawing.image import Image

# --- CONFIG ---
synergy_blocks = {
    "Strategic Fit": {"Type": "Business", "Weight": 30, "Defs": {1: "Completely misaligned", 2: "Partially aligned", 3: "Neutral", 4: "Well-aligned", 5: "Strategic driver"}},
    "Business Efficiency": {"Type": "Business", "Weight": 30, "Defs": {1: "Manual", 2: "Low efficiency", 3: "Average", 4: "High", 5: "Optimized"}},
    "User Value": {"Type": "Business", "Weight": 20, "Defs": {1: "Rejected", 2: "Low satisfaction", 3: "Acceptable", 4: "Good", 5: "Delightful"}},
    "Financial Value": {"Type": "Business", "Weight": 20, "Defs": {1: "Negative", 2: "Poor", 3: "Neutral", 4: "Positive", 5: "Exceptional"}},
    "Architecture": {"Type": "Tech", "Weight": 30, "Defs": {1: "Obsolete", 2: "Aging", 3: "Stable", 4: "Modern", 5: "Future-proof"}},
    "Operational Risk": {"Type": "Tech", "Weight": 30, "Defs": {1: "Critical", 2: "High", 3: "Managed", 4: "Low", 5: "Fortified"}},
    "Maintainability": {"Type": "Tech", "Weight": 25, "Defs": {1: "Impossible", 2: "Hard", 3: "Standard", 4: "Good", 5: "Excellent"}},
    "Support Quality": {"Type": "Tech", "Weight": 15, "Defs": {1: "Non-existent", 2: "Reactive", 3: "Defined", 4: "Proactive", 5: "World-class"}}
}

# USER PROVIDED EXACT LIST (Mapped to Blocks)
MASTER_QUESTIONS = {
    "Strategic Fit": [
        "What is the name of the application?",
        "What is the primary business purpose of the application?",
        "Which OPCOs use the application?",
        "Which utility domain(s) is the application used for? (Electric, Gas, or Both)",
        "Which Business Unit(s) use or own the application?",
        "Is the application IT-owned, business-owned, or jointly governed?",
        "Is this application considered business-critical, important, or supportive? Provide an explanation of the statement",
        "Does the application align with the current and future Mobility strategy?",
        "Are there important capabilities missing that limit business effectiveness?",
        "Is the application expected to be used in the next 3–5 years?", 
        "Are there planned upgrades, migrations, or replacements?", 
        "Could this application be replaced or consolidated with another platform?"
    ],
    "Business Efficiency": [
        "What key business processes does the application support?",
        "What core functionalities does the application provide?",
        "Are any processes partially supported or handled outside the application (manual workarounds, spreadsheets, etc.)?",
        "Does the application overlap functionally with other systems?",
        "Could this application absorb business processes currently supported by another platform or executed through manual workarounds? If yes, which processes and which platform(s)"
    ],
    "User Value": [
        "Which user roles or personas use the application (e.g., field technician, dispatcher, supervisor)?",
        "How many active users does the application have (daily/monthly)?",
        "Is application usage growing, stable, or declining?",
        "Is usage mandatory or optional for users?",
        "What is the overall level of user satisfaction?",
        "Are there known usability or mobility experience issues?"
    ],
    "Financial Value": [
        "What business value does the application deliver today?",
        "What would be the business impact if the application were unavailable?",
        "What are the main cost components (licenses, infrastructure, support)?, and what is the total cost of ownership?",
        "Is the cost reasonable compared to the business value delivered?",
        "Are there upcoming license renewals or contract milestones?",
        "Are there opportunities for cost reduction through consolidation or modernization?"
    ],
    "Architecture": [
        "Is the application a custom-built solution or a market (COTS/SaaS) product?",
        "What platforms does the application run on (mobile OS, web, backend)?",
        "What technologies, frameworks, or programming languages are used?",
        "What version of the application is currently deployed?",
        "Is the application deployed on-premises, in the cloud, or in a hybrid model?",
        "If cloud-based, which hyperscaler or cloud provider is used (e.g., AWS, Azure, GCP)?",
        "Which systems does the application integrate with?",
        "Are integrations real-time, batch-based, or manual?",
        "What limits future evolution or innovation?"
    ],
    "Operational Risk": [
        "Does the application support regulatory and compliance requirements?", 
        "How critical are these integrations to business operations?",
        "What type of data does the application create, consume, or update?",
        "Are there known integration or data quality issues?",
        "Does the application handle sensitive, personal, or regulated data?",
        "Is the application governed by corporate security and IT policies?",
        "Is identity and access management (IAM) integrated with corporate IAM solutions?",
        "Are security controls (authentication, authorization, logging) centrally managed or application-specific?",
        "Are there known security risks, audit findings, or compliance gaps?"
    ],
    "Maintainability": [
        "How complex is ongoing maintenance and support?",
        "How frequently are incidents or defects reported?",
        "How easy is it to implement enhancements or changes?",
        "What are the main business challenges with the application?",
        "What are the main technical challenges or limitations?", 
        "Are there scalability, performance, or reliability concerns?",
        "Are stakeholders requesting changes or replacement?"
    ],
    "Support Quality": [
        "Is vendor or technology support still available and active?",
        "Who provides application support (internal IT, vendor, third party)?",
        "Is the application proactively monitored (e.g., APM, Dynatrace), or is downtime primarily reported by users?",
        "Are alerts integrated with ITSM tools (e.g., ServiceNow) for auto-ticketing, or are they email-based/manual?"
    ]
}

# FLATTEN FOR MATCHING
ALL_QUESTIONS_FLAT = []
for k, q_list in MASTER_QUESTIONS.items():
    ALL_QUESTIONS_FLAT.extend(q_list)

MANUAL_LAYOUT = {
    "PoleForeman": 'l', 
    "Aspen OneLiner": 'b', 
    "SCAL-360 N": 'r', 
    "Mapping Computers": 't', 
    "Standard Tracking": 'b', 
    "ARCOS": 't', 
    "Cathodic": 'r', 
    "Bentley View": 'l', 
    "Bentley - ProjectWise": 'r',
}

# MANUAL APPS (Kaffa/Mapping) - Exact Keys Updated
MANUAL_APPS_DATA = {
    "Kaffa": {
        "What is the name of the application?": "Kaffa Platform",
        "What is the primary business purpose of the application?": "Asset Lifecycle Management (ALM) & Mobile GIS. Unifies Network Design, Construction, and Maintenance.",
        "Which OPCOs use the application?": "NY, CMP and UI in the future",
        "Is this application considered business-critical, important, or supportive? Provide an explanation of the statement": "High (Tier 2/1). Manages network construction and 'As-Built' asset registration.",
        "Does the application align with the current and future Mobility strategy?": "High. Construction Digitalization and Data Quality are strategic pillars.",
        "What business value does the application deliver today?": "Asset Base Integrity (RAB). Ensures that what is in the field matches the system.",
        "What key business processes does the application support?": "1. Network Design Projects. 2. Construction Monitoring. 3. Asset Commissioning.",
        "Are any processes partially supported or handled outside the application (manual workarounds, spreadsheets, etc.)?": "Yes. It usually replaces paper, but complex SAP integration failures can lead to manual steps.",
        "Does the application overlap functionally with other systems?": "High with SAP & GE Smallworld/Esri. Kaffa acts as the Middleware.",
        "Could this application absorb business processes currently supported by another platform or executed through manual workarounds? If yes, which processes and which platform(s)": "Yes. Can absorb other smaller inspection and registration tools.",
        "What is the overall level of user satisfaction?": "Good (UX Focus). Marketed as 'Easy to use'.",
        "Is application usage growing, stable, or declining?": "Growing (New Implementation). Since 2022.",
        "What are the main cost components (licenses, infrastructure, support)?, and what is the total cost of ownership?": "Medium. Licensing + Integration Services.",
        "Is the cost reasonable compared to the business value delivered?": "Positive. Revenue recovery via correct asset registration pays the bill.",
        "What technologies, frameworks, or programming languages are used?": "Java / .NET (Backend) + Mobile Native. Model Driven Architecture.",
        "Which systems does the application integrate with?": "GIS (Esri/Smallworld) and ERP (SAP).",
        "Are integrations real-time, batch-based, or manual?": "API / Services (Bus). Complex bidirectional.",
        "Does the application handle sensitive, personal, or regulated data?": "Yes (CEII). Detailed electric network maps.",
        "Are there known security risks, audit findings, or compliance gaps?": "Vendor Supply Chain (International).",
        "Is identity and access management (IAM) integrated with corporate IAM solutions?": "Yes. Supports Enterprise integration.",
        "How complex is ongoing maintenance and support?": "Medium/High. Keeping engineering rules synchronized requires effort.",
        "How easy is it to implement enhancements or changes?": "High (Configurable).",
        "Who provides application support (internal IT, vendor, third party)?": "Vendor (Codex) + Internal IT.",
        "Is vendor or technology support still available and active?": "Active."
    },
    "Mapping Computer": {
        "What is the name of the application?": "Mapping Computer (Ad-hoc Process)",
        "What is the primary business purpose of the application?": "Provide offline updated maps to field workers. Copies maps from shared folder to hard drives.",
        "Is this application considered business-critical, important, or supportive? Provide an explanation of the statement": "Critical",
        "Does the application align with the current and future Mobility strategy?": "No",
        "What key business processes does the application support?": "Share maps",
        "What core functionalities does the application provide?": "Create and update maps, download maps",
        "Are any processes partially supported or handled outside the application (manual workarounds, spreadsheets, etc.)?": "Scripts to be executed manually",
        "Does the application overlap functionally with other systems?": "Yes",
        "Could this application be replaced or consolidated with another platform?": "Yes, documentviewer or any application to share documents",
        "Is usage mandatory or optional for users?": "Mandatory",
        "Is application usage growing, stable, or declining?": "Stable",
        "Is the application deployed on-premises, in the cloud, or in a hybrid model?": "On the devices (Local)",
        "What limits future evolution or innovation?": "Bad architecture",
        "Does the application handle sensitive, personal, or regulated data?": "Gas and electric maps",
        "Is the application governed by corporate security and IT policies?": "No",
        "Is identity and access management (IAM) integrated with corporate IAM solutions?": "No",
        "Are there known security risks, audit findings, or compliance gaps?": "Yes, security and architectural gaps found",
        "Are there known integration or data quality issues?": "Yes",
        "Is the application a custom-built solution or a market (COTS/SaaS) product?": "Custom",
        "How complex is ongoing maintenance and support?": "Easy",
        "Are stakeholders requesting changes or replacement?": "Yes",
        "Who provides application support (internal IT, vendor, third party)?": "Internal IT",
        "Is vendor or technology support still available and active?": "No",
        "Is the application proactively monitored (e.g., APM, Dynatrace), or is downtime primarily reported by users?": "No"
    }
}

SOURCE_FILE = 'Application questionnaire (2).xlsx'
TEMPLATE_FILE = 'Avangrid_Application_Portfolio_Management - migrate first 3 tabs to our spreadsheet.xlsx'
OUTPUT_FILE = 'Avangrid_Application_Portfolio_Management_v54.xlsx'

def style(cell, bold=False, bg=None, color=None, border=True, align='center', wrap=False, italic=False, size=None, rotate=0):
    font_args = {'bold': bold, 'italic': italic}
    if color: font_args['color'] = color
    if size: font_args['size'] = size
    cell.font = Font(**font_args)
    if bg: cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    if border:
        bd = Side(style='thin', color='000000')
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    al_args = {'horizontal': align, 'vertical': 'top'}
    if wrap: al_args['wrap_text'] = True
    if rotate: al_args['text_rotation'] = rotate
    cell.alignment = Alignment(**al_args)

C_PRIM, C_SEC, C_NEU, C_W = 'E87722', '0066B3', '444444', 'FFFFFF'

KEYWORDS = {
    "high": ["automation", "optimized", "cloud", "modern", "integrated", "innovative", "strategic", "secure", "critical", "differentiator", "gold standard", "mandatory", "essential", "growing", "stable", "lifecycle", "unified", "mobile", "configurable", "active", "yes. supports", "middleware"],
    "med": ["standard", "functional", "managed", "defined", "accepted", "on premise", "market", "medium"],
    "low": ["manual", "legacy", "obsolete", "error", "poor", "risk", "unsupported", "silo", "redundant", "costly", "one person", "custom built", "unknown", "no", "scripts", "gaps", "bad architecture", "custom", "na"]
}

def evaluate_answer_text(text):
    if not text or len(text.strip()) < 2: return 0
    t = text.lower()
    score = 3
    pos_hits = sum(1 for w in KEYWORDS["high"] if w in t)
    neg_hits = sum(1 for w in KEYWORDS["low"] if w in t)
    if "no" == t.strip(): return 1
    if "yes" == t.strip(): return 4
    if pos_hits > neg_hits: score = 4
    if pos_hits > 2 and neg_hits == 0: score = 5
    if neg_hits > pos_hits: score = 2
    if neg_hits > 1: score = 1
    if ("security" in t or "iam" in t) and "no" in t: score = 1
    return score

def smart_wrap(text, limit=12):
    if len(text) <= limit: return text
    return "\n".join([text[i:i+limit] for i in range(0, len(text), limit)])

def parse_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames: return None
    ws = wb[sheet_name]
    q_col=a_col=s_col=-1; h_row=-1
    for r in range(1, 10):
        for c in range(1, 20):
            v = str(ws.cell(r,c).value).lower() if ws.cell(r,c).value else ""
            if "question" in v and "number" not in v: q_col=c
            if "response" in v or "answer" in v: a_col=c; 
            if "score" in v or "evaluation" in v: s_col=c
        if q_col>0 and a_col>0: h_row=r; break
    if h_row == -1: return None
    
    clean = sheet_name[:30].strip(); safe=clean
    for ch in "[]:*?/\\'": safe = safe.replace(ch, "")
    is_green = False
    try: 
        if ws.sheet_properties.tabColor and ("00FF00" in str(ws.sheet_properties.tabColor.rgb) or "00B050" in str(ws.sheet_properties.tabColor.rgb)): is_green=True
    except: pass
    
    app = {"name": clean, "safe_name": safe, "is_green": is_green, "answers": {}}
    current_q = None
    
    for r in range(h_row+1, ws.max_row+1):
        q_val = ws.cell(r, q_col).value
        a_val = ws.cell(r, a_col).value
        s_val = ws.cell(r, s_col).value if s_col > 0 else None
        q_str = str(q_val or "").strip()
        a_str = str(a_val or "").strip()
        
        if not q_str and not a_str: continue

        match = difflib.get_close_matches(q_str, ALL_QUESTIONS_FLAT, n=1, cutoff=0.85)
        
        if match:
            current_q = match[0]
            if a_str: app["answers"][current_q] = {"a": a_str, "s": s_val}
        else:
            if current_q and not q_str.endswith("?"):
                 app["answers"][current_q] = {"a": q_str, "s": s_val}
    return app

def calculate_score(app):
    app["blocks"] = {k: {"qa":[], "raw":0, "cnt":0, "score":1} for k in synergy_blocks}
    for category, q_list in MASTER_QUESTIONS.items():
        for q in q_list:
            ans_obj = app.get("answers", {}).get(q, {})
            if isinstance(ans_obj, str): ans_obj = {"a": ans_obj, "s": None}
            a_txt = ans_obj.get("a", "")
            s_val = ans_obj.get("s", None)
            score = 0
            t_score = evaluate_answer_text(a_txt)
            if t_score <= 2: score = t_score 
            else:
                 try: 
                    if s_val and float(s_val) >= 1: score = float(s_val)
                    else: score = t_score
                 except: score = t_score
            if len(a_txt) < 2: score = 0
            app["blocks"][category]["qa"].append({"q": q, "a": a_txt, "s": score})
            if score > 0:
                app["blocks"][category]["raw"] += score; app["blocks"][category]["cnt"] += 1

    for k, b in app["blocks"].items():
        if b["cnt"] > 0:
            sc = int(round(b["raw"] / b["cnt"]))
            if sc < 1: sc=1; 
            if sc > 5: sc=5; 
            b["score"] = sc
        else: b["score"] = 1
    return app

def get_apps():
    apps = []
    if os.path.exists(SOURCE_FILE): 
        try: wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
        except: wb = None
        if wb:
            ignore = ['Meetings', 'OPCOS', 'To Delete', 'Template', 'Questions Template', 'Sheet20', '_MASTER_TEMPLATE_']
            for sheet in wb.sheetnames:
                if sheet in ignore: continue
                is_manual = False
                for m_name in MANUAL_APPS_DATA:
                    if m_name.lower() in sheet.lower(): is_manual = True
                if not is_manual:
                    parsed_app = parse_sheet(wb, sheet)
                    if parsed_app:
                        apps.append(calculate_score(parsed_app))
                        
    for m_name, m_data in MANUAL_APPS_DATA.items():
        app = {"name": m_name, "safe_name": m_name, "is_green": True, "answers": m_data}
        apps.append(calculate_score(app))
    return apps

real_apps = get_apps()
if not real_apps: real_apps = []

print(f"Copying Template (v54)... Output: {OUTPUT_FILE}")
if os.path.exists(TEMPLATE_FILE):
    if os.path.exists(OUTPUT_FILE): os.remove(OUTPUT_FILE)
    shutil.copyfile(TEMPLATE_FILE, OUTPUT_FILE)
else: wb = openpyxl.Workbook(); wb.save(OUTPUT_FILE)

wb = openpyxl.load_workbook(OUTPUT_FILE)
wb.calculation.calcMode = 'auto'; wb.calculation.fullCalcOnLoad = True
for t in ["Calculator", "Dashboard", "Strategic Roadmap", "_MASTER_TEMPLATE_"]:
    if t in wb.sheetnames: del wb[t]
for app in real_apps:
    if app["safe_name"] in wb.sheetnames: del wb[app["safe_name"]]

# --- GENERATE TABS ---
print("Generating Content...")
ordered_keys = ["Strategic Fit", "Business Efficiency", "User Value", "Financial Value", "Architecture", "Operational Risk", "Maintainability", "Support Quality"]

def create_app_sheet(wb, app):
    ws = wb.create_sheet(app["safe_name"]); ws.sheet_properties.tabColor = "00B050" if app["is_green"] else "D9D9D9"
    ws.column_dimensions['A'].width=50 
    ws.column_dimensions['B'].width=80 
    ws.merge_cells("A1:B1"); ws.cell(1,1).value=f"Assessment: {app['name']}"; style(ws.cell(1,1), bold=True, size=14, color=C_PRIM)
    
    ws.cell(3,1).value = "EXECUTIVE SCORECARD"; style(ws.cell(3,1), bold=True, bg='333333', color='FFFFFF')
    ws.cell(3,2).value = "SCORE (0-5)"; style(ws.cell(3,2), bold=True, bg='333333', color='FFFFFF')
    
    for i, k in enumerate(ordered_keys):
        r = 4 + i
        ws.cell(r, 1).value = k
        style(ws.cell(r, 1), bold=True, align='right', bg='FAFAFA')
        sc_cell = ws.cell(r, 2)
        sc_cell.value = app["blocks"][k]["score"]
        style(sc_cell, bg='FFFFFF', border=True)
        dv = DataValidation(type="list", formula1='"0,1,2,3,4,5"', allow_blank=True); ws.add_data_validation(dv); dv.add(sc_cell)
        
    curr = 14
    for b_key in ordered_keys:
        b = app["blocks"][b_key]
        ws.merge_cells(f"A{curr}:B{curr}"); ws.cell(curr,1).value=b_key.upper(); style(ws.cell(curr,1), bold=True, bg=C_NEU, color=C_W)
        curr+=1; ws.cell(curr,1).value="Definitions"; style(ws.cell(curr,1), bold=True)
        curr+=1
        for scr in range(1,6): ws.cell(curr,1).value=f"{scr} - {synergy_blocks[b_key]['Defs'].get(scr,'')}"; style(ws.cell(curr,1), align='left'); curr+=1
        
        ws.cell(curr,1).value="Q"; style(ws.cell(curr,1), bold=True, bg=C_NEU, color=C_W); ws.cell(curr,2).value="A"; style(ws.cell(curr,2), bold=True, bg=C_NEU, color=C_W)
        curr+=1
        
        for qa in b["qa"]: 
            ws.cell(curr,1).value=qa["q"]; style(ws.cell(curr,1), bg='FAFAFA', align='left', wrap=True)
            ws.cell(curr,2).value=qa["a"]; style(ws.cell(curr,2), bg='FAFAFA', align='left', wrap=True)
            curr+=1
        curr+=1

for app in real_apps: create_app_sheet(wb, app)

# 2. Calculator
ws_calc = wb.create_sheet("Calculator"); ws_calc.freeze_panes = "B4"
ws_calc.column_dimensions['A'].width = 30
ws_calc.cell(1,1).value="Application Name"; style(ws_calc.cell(1,1), bold=True, bg=C_NEU, color=C_W)
ws_calc.cell(2,1).value="Weight"; style(ws_calc.cell(2,1), italic=True, align='right')
col=2; biz_blocks=[b for b,v in synergy_blocks.items() if v["Type"]=="Business"]; tech_blocks=[b for b,v in synergy_blocks.items() if v["Type"]=="Tech"]
for b_list, color in [(biz_blocks, C_PRIM), (tech_blocks, C_SEC)]:
    for b in b_list:
        c = ws_calc.cell(1, col); c.value=b; style(c, bold=True, bg=color, color=C_W, wrap=True); ws_calc.cell(2, col).value = synergy_blocks[b]["Weight"]; col += 1

c_bvi=col; ws_calc.cell(1, c_bvi).value="Business Value Index (BVI)"; ws_calc.column_dimensions[get_column_letter(c_bvi)].width=25; style(ws_calc.cell(1, c_bvi), bold=True, wrap=True)
c_thi=col+1; ws_calc.cell(1, c_thi).value="Technical Health Index (THI)"; ws_calc.column_dimensions[get_column_letter(c_thi)].width=25; style(ws_calc.cell(1, c_thi), bold=True, wrap=True)
c_dec=col+2; ws_calc.cell(1, c_dec).value="RECOMMENDATION"; style(ws_calc.cell(1, c_dec), bold=True)

MAX = 100
for i in range(MAX):
    r = i + 4
    if i < len(real_apps):
        app = real_apps[i]; ws_calc.cell(r,1).value = app["safe_name"] 
    style(ws_calc.cell(r,1), align='left', bg='FAFAFA'); cc = 2
    for b in biz_blocks + tech_blocks:
        b_idx = ordered_keys.index(b)
        fixed_row = 4 + b_idx
        ws_calc.cell(r,cc).value = f'=IF($A{r}="","",IFERROR(INDIRECT("\'"&$A{r}&"\'!$B${fixed_row}"),0))'
        style(ws_calc.cell(r,cc), bg='FAFAFA'); cc += 1
    rb_s, rb_e = get_column_letter(2), get_column_letter(1+len(biz_blocks)); rt_s, rt_e = get_column_letter(2+len(biz_blocks)), get_column_letter(1+len(biz_blocks)+len(tech_blocks))
    sum_rb = f"SUM({rb_s}{r}:{rb_e}{r})"; ws_calc.cell(r, c_bvi).value = f'=IF(OR($A{r}="", {sum_rb}=0),NA(),IFERROR(SUMPRODUCT({rb_s}{r}:{rb_e}{r}, ${rb_s}$2:${rb_e}$2)/SUM(${rb_s}$2:${rb_e}$2)*20, 0))'
    sum_rt = f"SUM({rt_s}{r}:{rt_e}{r})"; ws_calc.cell(r, c_thi).value = f'=IF(OR($A{r}="", {sum_rt}=0),NA(),IFERROR(SUMPRODUCT({rt_s}{r}:{rt_e}{r}, ${rt_s}$2:${rt_e}$2)/SUM(${rt_s}$2:${rt_e}$2)*20, 0))'
    cb, ct = f"{get_column_letter(c_bvi)}{r}", f"{get_column_letter(c_thi)}{r}"
    ws_calc.cell(r, c_dec).value = f'=IF(ISNA({cb}),"",IF(AND({cb}>=60,{ct}>=60),"EVOLVE",IF(AND({cb}>=60,{ct}<60),"INVEST",IF(AND({cb}<60,{ct}>=60),"MAINTAIN","ELIMINATE"))))'
    
rng_dec = f"{get_column_letter(c_dec)}4:{get_column_letter(c_dec)}{MAX+3}"
ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="EVOLVE"'], fill=PatternFill(bgColor='C6EFCE')))
ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="INVEST"'], fill=PatternFill(bgColor='FFEB9C')))
ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="MAINTAIN"'], fill=PatternFill(bgColor='BDD7EE')))
ws_calc.conditional_formatting.add(rng_dec, FormulaRule(formula=[f'{get_column_letter(c_dec)}4="ELIMINATE"'], fill=PatternFill(bgColor='FFC7CE')))

# 3. Dashboard
ws_dash = wb.create_sheet("Dashboard"); 
ch = ScatterChart(); ch.style=13; 
ch.x_axis.title="Technical Health Index (THI)"; 
ch.y_axis.title="Business Value Index (BVI)"; 
ch.x_axis.scaling.min=0; ch.x_axis.scaling.max=100; 
ch.y_axis.scaling.min=0; ch.y_axis.scaling.max=100; 
ch.display_blanks = 'gap'
ch.x_axis.crossesAt = 60; ch.y_axis.crossesAt = 60
sp_pr = GraphicalProperties(ln=LineProperties(prstDash='sysDot', w=9525, solidFill='D9D9D9')) 
ch.x_axis.majorGridlines = ChartLines(spPr=sp_pr); ch.y_axis.majorGridlines = ChartLines(spPr=sp_pr)
strong_line = GraphicalProperties(ln=LineProperties(w=19050, solidFill='000000'))
ch.x_axis.spPr = strong_line; ch.y_axis.spPr = strong_line
ch.legend = None; ch.width = 45; ch.height = 25
colors = {'EVOLVE':('C6EFCE','006100'), 'INVEST':('FFEB9C','9C5700'), 'MAINTAIN':('BDD7EE','0066CC'), 'ELIMINATE':('FFC7CE','9C0006')}
coord_map = {}; BUCKET = 3 
for i, app in enumerate(real_apps):
    r = 4 + i
    rb = sum(app["blocks"][k]["score"] for k in ordered_keys if "Fit" in k or "Efficiency" in k or "User" in k or "Financial" in k) / 4.0 * 20
    rt = sum(app["blocks"][k]["score"] for k in ordered_keys if "Architecture" in k or "Risk" in k or "Maintain" in k or "Support" in k) / 4.0 * 20
    rec = "ELIMINATE"
    if rb >= 60 and rt >= 60: rec="EVOLVE"
    elif rb >= 60 and rt < 60: rec="INVEST"
    elif rb < 60 and rt >= 60: rec="MAINTAIN" 
    x_b = int(rt // BUCKET); y_b = int(rb // BUCKET); k_coord = (x_b, y_b)
    count = coord_map.get(k_coord, 0); coord_map[k_coord] = count + 1
    fill, line = colors[rec]
    x_r = Reference(ws_calc, min_col=c_thi, min_row=r, max_row=r)
    y_r = Reference(ws_calc, min_col=c_bvi, min_row=r, max_row=r)
    wrapped_name = smart_wrap(app["safe_name"], 15)
    s = Series(values=y_r, xvalues=x_r, title=wrapped_name)
    s.marker.symbol = 'circle'; s.marker.size = 10
    s.marker.graphicalProperties.solidFill = fill
    s.marker.graphicalProperties.line.solidFill = line
    s.dLbls = DataLabelList(); s.dLbls.showSerName = True; s.dLbls.showVal = False
    pos = 't' 
    if count > 0: cycle = ['b', 'r', 'l', 't']; pos = cycle[count % 4]
    for k_man in MANUAL_LAYOUT:
        if k_man in app["safe_name"]: pos = MANUAL_LAYOUT[k_man]; break
    s.dLbls.position = pos
    ch.series.append(s)
ws_dash.add_chart(ch, "B2")

# 4. Roadmap
ws_strat = wb.create_sheet("Strategic Roadmap"); ws_strat.freeze_panes = "A2"
heads = ["Application Name", "Business Value Index (BVI)", "Technical Health Index (THI)", "Recommendation", "Subcategory", "Quick Win?", "Priority Rule", "Rationale", "Comments"]
for idx, h in enumerate(heads): 
    ws_strat.cell(1, idx+1).value=h
    if idx==1 or idx==2: ws_strat.column_dimensions[get_column_letter(idx+1)].width=25
    style(ws_strat.cell(1, idx+1), bold=True, bg=C_NEU, color=C_W, wrap=True)
ws_strat.column_dimensions['A'].width=35; ws_strat.column_dimensions['E'].width=20
c_cfg = 12; LISTS = {"ELIMINATE": ["Replace", "Retire", "Absorbed"], "INVEST": ["Absorb"], "EVOLVE": ["Modernize", "Enhance", "Migrate", "Refactor", "Upgrade"], "MAINTAIN": ["Internalize", "Maintain"]}
for key, items in LISTS.items():
    ws_strat.cell(1, c_cfg).value = key; 
    for idx, item in enumerate(items): ws_strat.cell(2+idx, c_cfg).value = item
    defn = DefinedName(f"LIST_{key}", attr_text=f"'Strategic Roadmap'!${get_column_letter(c_cfg)}$2:${get_column_letter(c_cfg)}{1+len(items)}"); wb.defined_names.add(defn); c_cfg += 1
ws_strat.cell(1, c_cfg).value = "QWIN"; ws_strat.cell(2, c_cfg).value="Yes"; ws_strat.cell(3, c_cfg).value="No"; ws_strat.cell(4, c_cfg).value="Review"; 
defn_qw = DefinedName("LIST_QWIN", attr_text=f"'Strategic Roadmap'!${get_column_letter(c_cfg)}$2:${get_column_letter(c_cfg)}{4}"); wb.defined_names.add(defn_qw); c_cfg += 1
c_prio = c_cfg + 1; ws_strat.cell(1, c_prio).value = "Matrix Config"; 
cfg_rows = [("ELIMINATE", "Replace", "P1 - Critical", "High Risk / EOL."), ("ELIMINATE", "Retire", "P1 - Critical", "Decommission"), ("ELIMINATE", "Absorbed", "P2 - Tactical", "Consolidation"), ("INVEST", "Absorb", "P1 - Critical", "High Value Opportunity"), ("EVOLVE", "Modernize", "P1 - Critical", "Transformation"), ("EVOLVE", "Migrate", "P1 - Critical", "Platform shift"), ("EVOLVE", "Enhance", "P2 - Strategic", "Expansion"), ("EVOLVE", "Refactor", "P2 - Strategic", "Code quality"), ("EVOLVE", "Upgrade", "P2 - Strategic", "Version"), ("MAINTAIN", "Internalize", "P2 - Compliance", "Governance"), ("MAINTAIN", "Maintain", "P3 - Routine", "Keep lights on")]
ws_strat.column_dimensions[get_column_letter(c_prio)].hidden = True
for i, h in enumerate(["Key", "Decision", "Subcategory", "Priority Rule", "Rationale"]): style(ws_strat.cell(2, c_prio+i), bold=True, bg='333333', color='FFFFFF', border=True); ws_strat.cell(2, c_prio+i).value=h
for i, (d, s, p, rat) in enumerate(cfg_rows):
    r=i+3; ws_strat.cell(r, c_prio).value = d+s; ws_strat.cell(r, c_prio+1).value = d; ws_strat.cell(r, c_prio+2).value = s; ws_strat.cell(r, c_prio+3).value = p; ws_strat.cell(r, c_prio+4).value = rat
defn_mtx = DefinedName("MatrixConfig", attr_text=f"'Strategic Roadmap'!${get_column_letter(c_prio)}$3:${get_column_letter(c_prio+4)}{len(cfg_rows)+3}"); wb.defined_names.add(defn_mtx)

wb.save(OUTPUT_FILE)
print("Done v54.")
