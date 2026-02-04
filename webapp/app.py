"""
Avangrid APM Platform - Main Application
Modern web application for Application Portfolio Management
"""

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from streamlit_option_menu import option_menu
import uuid
import io
import os
import sys
import json
from datetime import datetime, timezone
from typing import Dict
import difflib

# Ensure webapp directory is in Python path (needed for Streamlit Cloud)
_WEBAPP_DIR = os.path.dirname(os.path.abspath(__file__))
if _WEBAPP_DIR not in sys.path:
    sys.path.insert(0, _WEBAPP_DIR)

# Import local modules
from database import (
    get_session, close_session,
    Application, QuestionnaireAnswer, MeetingTranscript,
    TranscriptAnswer, SynergyScore, Insight, QAHistory, CustomWeight
)
from ai_processor import (
    extract_answers_from_transcript,
    suggest_scores,
    generate_insights,
    answer_question,
    calculate_bvi_thi,
    get_recommendation,
    get_subcategory_and_priority_detail,
    extract_dependencies_info,
    MASTER_QUESTIONS,
    SYNERGY_BLOCKS
)

# Import existing parsing logic
try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import ScatterChart, Series, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.chart.axis import ChartLines
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
    from PyPDF2 import PdfReader
    from docx import Document
except ImportError:
    pass

# ==================== PAGE CONFIG ====================
st.set_page_config(
    page_title="Avangrid APM Platform",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== WEIGHT PERSISTENCE HELPERS ====================
def load_weights_from_db():
    """Load custom weights from database into session_state"""
    if 'custom_weights' not in st.session_state:
        session = get_session()
        try:
            saved = session.query(CustomWeight).all()
            if saved:
                st.session_state.custom_weights = {w.block_name: w.weight for w in saved}
            else:
                st.session_state.custom_weights = {block: info['Weight'] for block, info in SYNERGY_BLOCKS.items()}
        finally:
            close_session(session)


def save_weights_to_db():
    """Persist current session_state weights to database"""
    if 'custom_weights' not in st.session_state:
        return
    session = get_session()
    try:
        for block_name, weight in st.session_state.custom_weights.items():
            existing = session.query(CustomWeight).filter_by(block_name=block_name).first()
            if existing:
                existing.weight = weight
                existing.updated_at = datetime.now(timezone.utc)
            else:
                session.add(CustomWeight(block_name=block_name, weight=weight, updated_at=datetime.now(timezone.utc)))
        session.commit()
    except Exception:
        session.rollback()
    finally:
        close_session(session)


def get_current_weights():
    """Get current weights from session_state or defaults"""
    load_weights_from_db()
    return st.session_state.get('custom_weights', {block: info['Weight'] for block, info in SYNERGY_BLOCKS.items()})


# ==================== ULTRA MODERN CSS - AVANGRID DESIGN SYSTEM ====================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

    /* === GLOBAL RESET === */
    * {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
        -webkit-font-smoothing: antialiased;
        -moz-osx-font-smoothing: grayscale;
    }

    #MainMenu, footer, header {visibility: hidden;}

    /* === MAIN LAYOUT === */
    .stApp {
        background: #F8F9FA;
    }

    .main .block-container {
        padding: 2rem 3rem;
        max-width: none;
        background: transparent;
    }

    /* === SIDEBAR - ULTRA CLEAN === */
    [data-testid="stSidebar"] {
        background: #FFFFFF;
        box-shadow: 2px 0 12px rgba(0, 0, 0, 0.04);
        border-right: 1px solid #E5E7EB;
        position: fixed !important;
        left: 0 !important;
        top: 0 !important;
        height: 100vh !important;
        overflow-y: auto !important;
        z-index: 999 !important;
    }

    /* Hide sidebar collapse button */
    [data-testid="collapsedControl"],
    button[kind="header"] {
        display: none !important;
    }

    [data-testid="stSidebar"] > div:first-child {
        padding: 2rem 1.5rem;
    }

    /* Sidebar Logo/Brand Area */
    [data-testid="stSidebar"] .element-container:first-child {
        background: linear-gradient(135deg, #E87722 0%, #FF8C42 100%);
        padding: 1.5rem;
        border-radius: 12px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 12px rgba(232, 119, 34, 0.15);
    }

    /* Sidebar Text Colors */
    [data-testid="stSidebar"] * {
        color: #374151 !important;
    }

    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] strong {
        color: #111827 !important;
    }

    /* === MAIN CONTENT AREA === */
    .main {
        background: #F8F9FA;
    }

    /* === TITLES - MODERN & CLEAN === */
    h1 {
        font-size: 2.5rem !important;
        font-weight: 800 !important;
        color: #111827 !important;
        margin-bottom: 0.5rem !important;
        letter-spacing: -0.025em;
        line-height: 1.2 !important;
    }

    h2 {
        font-size: 1.5rem !important;
        font-weight: 700 !important;
        color: #111827 !important;
        margin-top: 2rem !important;
        margin-bottom: 1rem !important;
        letter-spacing: -0.02em;
    }

    h3 {
        font-size: 1.125rem !important;
        font-weight: 600 !important;
        color: #374151 !important;
        margin-top: 1.5rem !important;
    }

    h4 {
        font-size: 0.875rem !important;
        font-weight: 600 !important;
        color: #6B7280 !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 0.5rem !important;
    }

    /* === METRIC CARDS - ULTRA MODERN & CLEAN === */
    [data-testid="stMetric"] {
        background: #FFFFFF;
        padding: 1.75rem;
        border-radius: 12px;
        border: 1px solid #E5E7EB;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
        transition: all 0.2s ease;
    }

    [data-testid="stMetric"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.08);
        border-color: #E87722;
    }

    [data-testid="stMetricValue"] {
        font-size: 2.25rem !important;
        font-weight: 800 !important;
        color: #E87722 !important;
        line-height: 1 !important;
    }

    [data-testid="stMetricLabel"] {
        color: #6B7280 !important;
        font-weight: 600 !important;
        font-size: 0.875rem !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 0.75rem;
    }

    /* === BUTTONS - MODERN & CLEAN === */
    .stButton > button {
        background: #E87722;
        color: white !important;
        border: none;
        border-radius: 8px;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 0.875rem;
        text-transform: none;
        letter-spacing: 0;
        transition: all 0.2s ease;
        box-shadow: 0 1px 3px rgba(232, 119, 34, 0.3);
    }

    .stButton > button:hover {
        background: #D66A1A;
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(232, 119, 34, 0.4);
    }

    .stButton > button:active {
        transform: translateY(0);
    }

    .stButton > button[kind="primary"] {
        background: #0066B3;
        box-shadow: 0 1px 3px rgba(0, 102, 179, 0.3);
    }

    .stButton > button[kind="primary"]:hover {
        background: #005499;
        box-shadow: 0 4px 12px rgba(0, 102, 179, 0.4);
    }

    /* === MODERN FEATURE CARDS === */
    .feature-card {
        background: #FFFFFF;
        border-radius: 12px;
        padding: 2rem;
        border: 1px solid #E5E7EB;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
        transition: all 0.2s ease;
        height: 100%;
        display: flex;
        flex-direction: column;
    }

    .feature-card:hover {
        transform: translateY(-4px);
        box-shadow: 0 12px 24px rgba(0, 0, 0, 0.08);
        border-color: #E87722;
    }

    .feature-icon {
        width: 48px;
        height: 48px;
        border-radius: 10px;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.5rem;
        margin-bottom: 1rem;
        flex-shrink: 0;
    }

    .feature-icon.orange {
        background: linear-gradient(135deg, #FEF3E7 0%, #FDEBD0 100%);
        color: #E87722;
    }

    .feature-icon.blue {
        background: linear-gradient(135deg, #E8F4FB 0%, #D1E7F5 100%);
        color: #0066B3;
    }

    .feature-icon.green {
        background: linear-gradient(135deg, #ECFDF5 0%, #D1FAE5 100%);
        color: #10B981;
    }

    .feature-title {
        font-size: 0.875rem;
        font-weight: 600;
        color: #111827;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        margin-bottom: 0.5rem;
    }

    .feature-items {
        list-style: none;
        padding: 0;
        margin: 0;
        flex-grow: 1;
    }

    .feature-items li {
        color: #6B7280;
        font-size: 0.9375rem;
        line-height: 1.8;
        padding-left: 1.25rem;
        position: relative;
    }

    .feature-items li:before {
        content: "•";
        color: #E87722;
        font-weight: bold;
        position: absolute;
        left: 0;
    }

    /* === METRIC HIGHLIGHTS === */
    .metric-highlight {
        text-align: center;
        padding: 2rem 0;
    }

    .metric-highlight-value {
        font-size: 3rem;
        font-weight: 900;
        background: linear-gradient(135deg, #E87722 0%, #FF8C42 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        line-height: 1;
        margin-bottom: 0.5rem;
    }

    .metric-highlight-label {
        color: #6B7280;
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.1em;
    }

    /* === FILE UPLOADER === */
    [data-testid="stFileUploader"] {
        background: #FFFFFF;
        border-radius: 12px;
        padding: 2.5rem;
        border: 2px dashed #D1D5DB;
        transition: all 0.2s ease;
    }

    [data-testid="stFileUploader"]:hover {
        border-color: #E87722;
        background: #FEFEFE;
    }

    /* === ALERTS === */
    .stSuccess, .stWarning, .stError, .stInfo {
        border-radius: 8px;
        padding: 1rem 1.25rem;
        border-left: 4px solid;
        font-weight: 500;
    }

    .stSuccess {
        background: #F0FDF4;
        border-left-color: #10B981;
        color: #065F46 !important;
    }

    .stWarning {
        background: #FFFBEB;
        border-left-color: #F59E0B;
        color: #92400E !important;
    }

    .stError {
        background: #FEF2F2;
        border-left-color: #EF4444;
        color: #991B1B !important;
    }

    .stInfo {
        background: #EFF6FF;
        border-left-color: #3B82F6;
        color: #1E40AF !important;
    }

    /* === TABS === */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.5rem;
        background: transparent;
        border-bottom: 1px solid #E5E7EB;
    }

    .stTabs [data-baseweb="tab"] {
        background: transparent;
        border-radius: 6px 6px 0 0;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        border: none;
        color: #6B7280;
        transition: all 0.2s ease;
    }

    .stTabs [data-baseweb="tab"]:hover {
        background: #F9FAFB;
        color: #E87722;
    }

    .stTabs [aria-selected="true"] {
        background: transparent;
        color: #E87722 !important;
        border-bottom: 2px solid #E87722;
    }

    /* === DATAFRAME === */
    [data-testid="stDataFrame"] {
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid #E5E7EB;
        background: #FFFFFF;
    }

    /* === INPUTS === */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea {
        background: #FFFFFF !important;
        border: 1px solid #D1D5DB !important;
        border-radius: 8px;
        padding: 0.625rem 0.875rem;
        color: #111827 !important;
        font-size: 0.9375rem;
        transition: all 0.2s ease;
    }

    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: #E87722 !important;
        box-shadow: 0 0 0 3px rgba(232, 119, 34, 0.1) !important;
    }

    /* === SELECTBOX === */
    .stSelectbox > div > div {
        background: #FFFFFF !important;
        border: 1px solid #D1D5DB;
        border-radius: 8px;
        color: #111827 !important;
    }

    /* === EXPANDER === */
    .streamlit-expanderHeader {
        background: #FFFFFF;
        border-radius: 8px;
        border: 1px solid #E5E7EB;
        padding: 1rem 1.25rem;
        font-weight: 600;
        color: #111827 !important;
        transition: all 0.2s ease;
    }

    .streamlit-expanderHeader:hover {
        border-color: #E87722;
        background: #FEFEFE;
    }

    /* === SPINNER & PROGRESS === */
    .stSpinner > div {
        border-top-color: #E87722 !important;
    }

    .stProgress > div > div > div {
        background: linear-gradient(90deg, #E87722 0%, #FF8C42 100%);
        border-radius: 8px;
    }

    /* === MARKDOWN TEXT === */
    .main .stMarkdown {
        color: #374151;
    }

    .main .stMarkdown strong {
        color: #111827;
    }

    .main .stMarkdown a {
        color: #E87722;
        text-decoration: none;
        transition: color 0.2s;
    }

    .main .stMarkdown a:hover {
        color: #D66A1A;
        text-decoration: underline;
    }

    /* === SCROLLBAR === */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }

    ::-webkit-scrollbar-track {
        background: #F3F4F6;
    }

    ::-webkit-scrollbar-thumb {
        background: #D1D5DB;
        border-radius: 4px;
    }

    ::-webkit-scrollbar-thumb:hover {
        background: #9CA3AF;
    }

    /* === ANIMATIONS === */
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }

    .animate-fade-in {
        animation: fadeInUp 0.4s ease-out;
    }
</style>
""", unsafe_allow_html=True)

# ==================== SESSION STATE INITIALIZATION ====================
if 'current_page' not in st.session_state:
    st.session_state.current_page = 'Introduction'
if 'selected_application' not in st.session_state:
    st.session_state.selected_application = None
if 'applications_data' not in st.session_state:
    st.session_state.applications_data = []
if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []

# ==================== HELPER FUNCTIONS ====================

def _parse_meetings_sheet(wb):
    """Parse the Meetings sheet to extract Business Owner and IT Owner per application.
    Returns dict: {normalized_app_name: {'business_owner': str, 'it_owner': str, 'raw_name': str}}
    """
    meetings_data = {}
    if 'Meetings' not in wb.sheetnames:
        return meetings_data

    ws = wb['Meetings']
    # Find header row and column indices
    header_row = None
    col_app_name = None
    col_business_owner = None
    col_it_owner = None

    for row in ws.iter_rows(max_row=5):
        for idx, cell in enumerate(row):
            val = str(cell.value).lower().strip() if cell.value else ""
            if 'application name' in val or val == 'application':
                col_app_name = idx
                header_row = cell.row
            elif 'business owner' in val:
                col_business_owner = idx
            elif 'it owner' in val:
                col_it_owner = idx

    if col_app_name is None or header_row is None:
        return meetings_data

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        app_name_cell = row[col_app_name].value if col_app_name < len(row) else None
        if not app_name_cell:
            continue
        app_name = str(app_name_cell).strip()
        if not app_name:
            continue

        bo = ""
        it_o = ""
        if col_business_owner is not None and col_business_owner < len(row):
            bo = str(row[col_business_owner].value or "").strip()
        if col_it_owner is not None and col_it_owner < len(row):
            it_o = str(row[col_it_owner].value or "").strip()

        normalized = normalize_app_name(app_name)
        meetings_data[normalized] = {
            'business_owner': bo,
            'it_owner': it_o,
            'raw_name': app_name
        }

    return meetings_data


def parse_questionnaire_excel(uploaded_file):
    """Parse uploaded questionnaire Excel file (from existing code)"""
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        applications = []

        # Parse Meetings sheet for Business Owner / IT Owner data
        meetings_data = _parse_meetings_sheet(wb)

        for sheet_name in wb.sheetnames:
            # Skip template/metadata sheets
            if sheet_name.lower() in ['index', 'introduction', 'methodology', 'user guide',
                                       'calculator', 'dashboard', 'strategic roadmap',
                                       'application groups', 'value chain', 'sheet1',
                                       'meetings', 'opcos', 'to delete', 'questions template']:
                continue

            ws = wb[sheet_name]

            # Find columns
            header_row = None
            for row in ws.iter_rows(max_row=10):
                cells = [str(cell.value).lower() if cell.value else "" for cell in row]
                if any('question' in c for c in cells):
                    header_row = row
                    break

            if not header_row:
                continue

            # Map columns
            col_map = {}
            for idx, cell in enumerate(header_row):
                val = str(cell.value).lower() if cell.value else ""
                if 'question' in val:
                    col_map['question'] = idx
                elif 'answer' in val or 'response' in val:
                    col_map['answer'] = idx
                elif 'score' in val:
                    col_map['score'] = idx

            if 'question' not in col_map or 'answer' not in col_map:
                continue

            # Extract Q&A
            app_data = {
                'name': sheet_name.strip(),
                'safe_name': sheet_name[:31].strip(),
                'is_green': ws.sheet_properties.tabColor is not None and
                           hasattr(ws.sheet_properties.tabColor, 'rgb') and
                           ws.sheet_properties.tabColor.rgb and
                           'FF00FF00' in str(ws.sheet_properties.tabColor.rgb).upper(),
                'answers': {}
            }

            for row in ws.iter_rows(min_row=header_row[0].row + 1):
                question = row[col_map['question']].value
                answer = row[col_map['answer']].value
                score = row[col_map.get('score', -1)].value if 'score' in col_map else None

                if question and str(question).strip():
                    question_text = str(question).strip()
                    answer_text = str(answer).strip() if answer else ""

                    # Fuzzy match to master questions
                    best_match = None
                    best_ratio = 0
                    for block, questions in MASTER_QUESTIONS.items():
                        for mq in questions:
                            ratio = difflib.SequenceMatcher(None, question_text.lower(), mq.lower()).ratio()
                            if ratio > best_ratio:
                                best_ratio = ratio
                                best_match = (mq, block)

                    if best_ratio > 0.75 and best_match:
                        matched_question, synergy_block = best_match
                        app_data['answers'][matched_question] = {
                            'a': answer_text,
                            's': score if score else None,
                            'block': synergy_block
                        }

            # Inject Business Owner / IT Owner from Meetings sheet
            if meetings_data:
                norm_sheet = normalize_app_name(sheet_name)
                # Try exact normalized match first, then fuzzy
                owner_info = meetings_data.get(norm_sheet)
                if not owner_info:
                    # Fuzzy match against Meetings app names
                    best_ratio = 0
                    best_key = None
                    for mk in meetings_data:
                        ratio = difflib.SequenceMatcher(None, norm_sheet, mk).ratio()
                        if ratio > best_ratio:
                            best_ratio = ratio
                            best_key = mk
                    if best_ratio > 0.75 and best_key:
                        owner_info = meetings_data[best_key]

                if owner_info:
                    if owner_info['business_owner']:
                        app_data['answers']["Who is the Business Owner of this application?"] = {
                            'a': owner_info['business_owner'],
                            's': None,
                            'block': 'Strategic Fit'
                        }
                    if owner_info['it_owner']:
                        app_data['answers']["Who is the IT Owner of this application?"] = {
                            'a': owner_info['it_owner'],
                            's': None,
                            'block': 'Strategic Fit'
                        }

            if app_data['answers']:
                applications.append(app_data)

        return applications

    except Exception as e:
        st.error(f"Error parsing questionnaire: {e}")
        return []


def normalize_app_name(name: str) -> str:
    """Normalize application name for matching"""
    import re

    # Remove extra spaces, convert to lowercase
    normalized = ' '.join(name.strip().lower().split())

    # Remove parentheses and their content (e.g., "(SCG & CNG)" → "")
    normalized = re.sub(r'\([^)]*\)', '', normalized)

    # Remove common words that don't affect app identity
    noise_words = ['remote', 'local', 'the', 'a', 'an']
    for word in noise_words:
        normalized = normalized.replace(f' {word} ', ' ')

    # Remove common separators for comparison
    normalized = normalized.replace(' - ', ' ').replace('-', ' ')
    normalized = normalized.replace(' & ', ' ').replace('&', ' ')

    # Remove special characters but keep alphanumeric and spaces
    normalized = ''.join(c if c.isalnum() or c.isspace() else ' ' for c in normalized)

    # Remove extra spaces again
    normalized = ' '.join(normalized.split())
    return normalized


def get_significant_tokens(name: str) -> tuple:
    """
    Extract significant tokens from app name
    Returns: (primary_tokens, all_tokens)
    Primary tokens are from main part (before parentheses), more important for matching
    """
    import re

    # Convert to lowercase and split
    text = name.lower()

    # Separate primary (before parens) and secondary (in parens) content
    text_no_paren = re.sub(r'\([^)]*\)', '', text)
    paren_content = re.findall(r'\(([^)]+)\)', text)

    # Expanded noise words - common descriptive terms that don't identify the app
    noise = {
        'the', 'a', 'an', 'and', 'or', 'of', 'in', 'on', 'at', 'to', 'for', 'with', 'by',
        'file', 'app', 'application', 'system', 'tool', 'software', 'ms', 'project',
        'database', 'db', 'program', 'service'
    }

    # Extract primary tokens (before parentheses) - most important
    primary_tokens = set(re.findall(r'\w+', text_no_paren))
    primary_tokens = {t for t in primary_tokens if t not in noise and len(t) >= 2}

    # Extract all tokens including parentheses content
    all_text = text_no_paren + ' ' + ' '.join(paren_content)
    all_tokens = set(re.findall(r'\w+', all_text))
    all_tokens = {t for t in all_tokens if t not in noise and len(t) >= 2}

    return primary_tokens, all_tokens


def find_matching_application(file_app_name: str, app_dict: dict) -> tuple:
    """
    Find matching application using smart matching algorithm
    Returns: (matched_app, match_type) or (None, None)

    Match types: 'exact', 'normalized', 'substring', 'token', 'fuzzy', 'fuzzy_normalized'
    """
    file_app_lower = file_app_name.strip().lower()

    # Strategy 1: Exact match (case-insensitive)
    if file_app_lower in app_dict:
        return app_dict[file_app_lower], 'exact'

    # Strategy 2: Normalized match (remove dashes and extra spaces)
    file_normalized = normalize_app_name(file_app_name)

    for app_name_lower, app in app_dict.items():
        app_normalized = normalize_app_name(app_name_lower)
        if file_normalized == app_normalized:
            return app, 'normalized'

    # Strategy 3: Substring containment - check if file name contains an app name
    # or vice versa. Prefer longest match to avoid "Bentley" matching when
    # "Bentley - PLS-CADD" exists.
    substring_matches = []
    for app_name_lower, app in app_dict.items():
        # Check both directions
        if app_name_lower in file_app_lower or file_app_lower in app_name_lower:
            # Score by length of overlap (prefer longer/more specific matches)
            overlap_len = min(len(app_name_lower), len(file_app_lower))
            substring_matches.append((app, overlap_len, app_name_lower))

    if substring_matches:
        # Pick the longest (most specific) match
        substring_matches.sort(key=lambda x: x[1], reverse=True)
        return substring_matches[0][0], 'substring'

    # Strategy 4: Token-based matching - prioritize primary tokens
    file_primary, file_all = get_significant_tokens(file_app_name)

    best_match = None
    best_score = 0

    for app_name_lower, app in app_dict.items():
        app_primary, app_all = get_significant_tokens(app_name_lower)

        if not file_primary or not app_primary:
            continue

        # First check primary tokens match (more important)
        primary_intersection = len(file_primary & app_primary)
        primary_union = len(file_primary | app_primary)
        primary_similarity = primary_intersection / primary_union if primary_union > 0 else 0

        # If primary tokens match well (>=80%), it's a strong match
        if primary_similarity >= 0.8:
            if primary_similarity > best_score:
                best_score = primary_similarity
                best_match = app
            continue

        # Otherwise, check all tokens with lower threshold
        all_intersection = len(file_all & app_all)
        all_union = len(file_all | app_all)
        all_similarity = all_intersection / all_union if all_union > 0 else 0

        if all_similarity > best_score and all_similarity >= 0.4:  # Lower threshold for all tokens
            best_score = all_similarity
            best_match = app

    if best_match:
        return best_match, 'token'

    # Strategy 5: Fuzzy match using difflib (similarity > 80%)
    matches = difflib.get_close_matches(
        file_app_lower,
        app_dict.keys(),
        n=1,
        cutoff=0.80
    )

    if matches:
        return app_dict[matches[0]], 'fuzzy'

    # Strategy 6: Try normalized fuzzy match
    app_names_normalized = {normalize_app_name(k): v for k, v in app_dict.items()}

    matches_normalized = difflib.get_close_matches(
        file_normalized,
        app_names_normalized.keys(),
        n=1,
        cutoff=0.80
    )

    if matches_normalized:
        return app_names_normalized[matches_normalized[0]], 'fuzzy_normalized'

    return None, None


def read_transcript_file(uploaded_file) -> str:
    """Read transcript from various file formats"""
    try:
        file_extension = uploaded_file.name.split('.')[-1].lower()

        if file_extension == 'txt':
            return uploaded_file.getvalue().decode('utf-8')

        elif file_extension == 'pdf':
            pdf_reader = PdfReader(io.BytesIO(uploaded_file.getvalue()))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            return text

        elif file_extension in ['docx', 'doc']:
            doc = Document(io.BytesIO(uploaded_file.getvalue()))
            text = "\n".join([para.text for para in doc.paragraphs])
            return text

        else:
            st.warning(f"Unsupported file format: {file_extension}")
            return ""

    except Exception as e:
        st.error(f"Error reading transcript file: {e}")
        return ""


def save_application_to_db(app_data: dict, session):
    """Save application and its answers to database"""
    try:
        # Check if application exists
        existing_app = session.query(Application).filter_by(name=app_data['name']).first()

        if existing_app:
            app = existing_app
        else:
            app = Application(
                id=str(uuid.uuid4()),
                name=app_data['name'],
                safe_name=app_data.get('safe_name', app_data['name'][:31]),
                is_green=app_data.get('is_green', False)
            )
            session.add(app)

        # Save questionnaire answers
        for question, answer_obj in app_data['answers'].items():
            existing_answer = session.query(QuestionnaireAnswer).filter_by(
                application_id=app.id,
                question_text=question
            ).first()

            new_answer_text = answer_obj.get('a', '')

            if not existing_answer:
                # No answer exists - create new
                qa = QuestionnaireAnswer(
                    id=str(uuid.uuid4()),
                    application_id=app.id,
                    question_text=question,
                    answer_text=new_answer_text,
                    score=answer_obj.get('s'),
                    synergy_block=answer_obj.get('block', 'Unknown')
                )
                session.add(qa)
            else:
                # Answer exists - check if incomplete (empty or very short)
                is_incomplete = (
                    not existing_answer.answer_text or
                    len(existing_answer.answer_text.strip()) < 5
                )

                # Only update if existing answer is incomplete AND new answer has content
                if is_incomplete and new_answer_text and len(new_answer_text.strip()) >= 5:
                    existing_answer.answer_text = new_answer_text
                    existing_answer.score = answer_obj.get('s')
                    existing_answer.synergy_block = answer_obj.get('block', 'Unknown')

        session.commit()
        return app

    except Exception as e:
        session.rollback()
        st.error(f"Error saving to database: {e}")
        return None


def get_all_applications_from_db(session):
    """Get all applications with their data"""
    try:
        apps = session.query(Application).all()
        return apps
    except Exception as e:
        st.error(f"Error fetching applications: {e}")
        return []


def get_aggregated_rationale(app_id: str, session) -> Dict:
    """
    Get aggregated rationale from all synergy block scores for an application.
    Returns dict with 'summary' and 'details' keys.
    """
    try:
        scores = session.query(SynergyScore).filter_by(
            application_id=app_id,
            approved=True
        ).all()

        if not scores:
            return {
                'summary': 'No assessment available',
                'details': {}
            }

        # Build summary from all rationales
        rationale_parts = []
        details = {}

        for score in scores:
            if score.rationale and score.rationale.strip():
                # Clean rationale
                rationale_clean = score.rationale.replace('⚠️ NO DATA -', '').strip()
                if rationale_clean and not rationale_clean.startswith('Conservative score'):
                    rationale_parts.append(f"{score.block_name}: {rationale_clean}")
                    details[score.block_name] = rationale_clean

        # Create summary
        if rationale_parts:
            summary = " | ".join(rationale_parts[:3])  # First 3 blocks
            if len(rationale_parts) > 3:
                summary += f" (+{len(rationale_parts) - 3} more)"
        else:
            summary = "Assessment based on questionnaire and transcript data"

        return {
            'summary': summary,
            'details': details
        }

    except Exception as e:
        return {
            'summary': f'Error: {str(e)}',
            'details': {}
        }


# ==================== MODERN CARD HELPER ====================
def render_modern_card(icon, title, items, color="orange"):
    """Render a modern card using HTML components"""
    color_map = {
        "orange": {"bg": "#FEF3E7", "text": "#E87722"},
        "blue": {"bg": "#E8F4FB", "text": "#0066B3"},
        "green": {"bg": "#ECFDF5", "text": "#10B981"}
    }

    c = color_map.get(color, color_map["orange"])
    items_html = "".join([f"<li>{item}</li>" for item in items])

    html = f"""
    <style>
        .modern-card {{
            background: white;
            border-radius: 12px;
            padding: 2rem;
            border: 1px solid #E5E7EB;
            box-shadow: 0 1px 3px rgba(0, 0, 0, 0.05);
            transition: all 0.3s ease;
            height: 100%;
        }}
        .modern-card:hover {{
            transform: translateY(-4px);
            box-shadow: 0 12px 24px rgba(0, 0, 0, 0.1);
            border-color: {c['text']};
        }}
        .card-icon {{
            width: 48px;
            height: 48px;
            border-radius: 10px;
            background: {c['bg']};
            color: {c['text']};
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.5rem;
            margin-bottom: 1rem;
        }}
        .card-title {{
            font-size: 0.875rem;
            font-weight: 600;
            color: #111827;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 1rem;
        }}
        .card-items {{
            list-style: none;
            padding: 0;
            margin: 0;
        }}
        .card-items li {{
            color: #6B7280;
            font-size: 0.9375rem;
            line-height: 1.8;
            padding-left: 1.25rem;
            position: relative;
        }}
        .card-items li:before {{
            content: "•";
            color: {c['text']};
            font-weight: bold;
            position: absolute;
            left: 0;
        }}
    </style>
    <div class="modern-card">
        <div class="card-icon">{icon}</div>
        <div class="card-title">{title}</div>
        <ul class="card-items">{items_html}</ul>
    </div>
    """
    components.html(html, height=280)


# ==================== PAGE: INTRODUCTION ====================
def page_introduction():
    """Executive Introduction - Strategic Portfolio Management Overview"""

    # Hero section - Professional, executive style
    st.markdown("""
    <div style="background: linear-gradient(135deg, #F9FAFB 0%, #FFFFFF 100%);
                padding: 3rem 2.5rem;
                border-radius: 16px;
                border-left: 4px solid #E87722;
                margin-bottom: 2.5rem;
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.05);">
        <h1 style="color: #111827; margin: 0 0 1rem 0; font-size: 2.25rem; font-weight: 700; line-height: 1.2;">
            Application Portfolio Management
        </h1>
        <p style="color: #6B7280; font-size: 1.125rem; margin: 0; line-height: 1.6;">
            Data-driven framework for strategic portfolio optimization across 70+ enterprise applications,
            enabling informed investment decisions through systematic assessment of business value and technical health.
        </p>
    </div>
    """, unsafe_allow_html=True)

    # Executive Summary Cards
    st.markdown("### Executive Summary")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
        <div style="background: #FFFFFF; padding: 1.75rem; border-radius: 12px;
                    border: 1px solid #E5E7EB; height: 100%; margin-bottom: 1rem;">
            <div style="display: flex; align-items: center; margin-bottom: 1rem;">
                <div style="background: rgba(232, 119, 34, 0.1); padding: 0.75rem; border-radius: 10px; margin-right: 1rem;">
                    <span style="font-size: 1.5rem;">🎯</span>
                </div>
                <h4 style="color: #111827; margin: 0; font-weight: 600;">Strategic Objective</h4>
            </div>
            <p style="color: #4B5563; font-size: 0.9375rem; line-height: 1.6; margin: 0;">
                Transform application portfolio into a strategic asset by identifying optimization opportunities,
                eliminating redundancies, and aligning IT investments with business priorities. Enable data-driven
                decision-making through quantitative assessment of business value and technical health.
            </p>
        </div>
        """, unsafe_allow_html=True)

    with col2:
        st.markdown("""
        <div style="background: #FFFFFF; padding: 1.75rem; border-radius: 12px;
                    border: 1px solid #E5E7EB; height: 100%; margin-bottom: 1rem;">
            <div style="display: flex; align-items: center; margin-bottom: 1rem;">
                <div style="background: rgba(232, 119, 34, 0.1); padding: 0.75rem; border-radius: 10px; margin-right: 1rem;">
                    <span style="font-size: 1.5rem;">📊</span>
                </div>
                <h4 style="color: #111827; margin: 0; font-weight: 600;">Portfolio Scope</h4>
            </div>
            <p style="color: #4B5563; font-size: 0.9375rem; line-height: 1.6; margin: 0;">
                Comprehensive assessment of 70+ mission-critical applications spanning Electric & Gas Operations,
                Corporate Functions, Customer Solutions, and Grid Management. Evaluation encompasses both
                commercial products (SAP, Oracle, Maximo) and custom-built internal systems.
            </p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

    # Assessment Framework
    st.markdown("### Two-Dimensional Assessment Framework")

    # Get current weights from session_state or use defaults
    current_weights = get_current_weights()

    col1, col2 = st.columns(2)

    # BVI blocks with descriptions
    bvi_blocks = {
        "Strategic Fit": "Alignment with organizational strategy and long-term goals",
        "Business Efficiency": "Process optimization, automation impact, and cost savings",
        "User Value": "User satisfaction, adoption rates, and productivity gains",
        "Financial Value": "ROI, cost avoidance, and revenue impact"
    }

    # THI blocks with descriptions
    thi_blocks = {
        "Architecture": "Scalability, flexibility, integration capabilities, and tech stack modernity",
        "Operational Risk": "Reliability, security vulnerabilities, and business continuity risk",
        "Maintainability": "Code quality, documentation, ease of updates and extensibility",
        "Support Quality": "Vendor support, internal expertise, and knowledge transfer capability"
    }

    with col1:
        bvi_items = ""
        for block, description in bvi_blocks.items():
            weight = current_weights.get(block, 25)
            bvi_items += f'<div style="display:flex;align-items:baseline;margin-bottom:0.75rem;"><span style="color:#10B981;font-weight:700;min-width:40px;">{weight}%</span><div><strong style="color:#111827;">{block}</strong><p style="color:#6B7280;font-size:0.8125rem;margin:0.25rem 0 0 0;">{description}</p></div></div>'

        st.markdown(f'<div style="background:#FFFFFF;padding:2rem;border-radius:12px;border:1px solid #E5E7EB;border-top:3px solid #10B981;"><h4 style="color:#111827;margin:0 0 1.25rem 0;font-weight:600;">Business Value Index (BVI)</h4><p style="color:#6B7280;font-size:0.875rem;margin-bottom:1.5rem;line-height:1.6;">Quantifies application contribution to business objectives through weighted assessment of:</p><div style="margin-left:0.5rem;">{bvi_items}</div></div>', unsafe_allow_html=True)

    with col2:
        thi_items = ""
        for block, description in thi_blocks.items():
            weight = current_weights.get(block, 25)
            thi_items += f'<div style="display:flex;align-items:baseline;margin-bottom:0.75rem;"><span style="color:#3B82F6;font-weight:700;min-width:40px;">{weight}%</span><div><strong style="color:#111827;">{block}</strong><p style="color:#6B7280;font-size:0.8125rem;margin:0.25rem 0 0 0;">{description}</p></div></div>'

        st.markdown(f'<div style="background:#FFFFFF;padding:2rem;border-radius:12px;border:1px solid #E5E7EB;border-top:3px solid #3B82F6;"><h4 style="color:#111827;margin:0 0 1.25rem 0;font-weight:600;">Technical Health Index (THI)</h4><p style="color:#6B7280;font-size:0.875rem;margin-bottom:1.5rem;line-height:1.6;">Evaluates technical sustainability and operational risk through analysis of:</p><div style="margin-left:0.5rem;">{thi_items}</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

    # Strategic Decision Matrix
    st.markdown("### Strategic Decision Matrix")
    st.markdown("""
    <p style="color: #6B7280; font-size: 0.9375rem; margin-bottom: 1.5rem; line-height: 1.6;">
        Applications are positioned on a 2x2 matrix (BVI vs THI) with threshold at 60 points,
        resulting in four strategic pathways that guide investment priorities and portfolio actions.
    </p>
    """, unsafe_allow_html=True)

    # 2x2 Matrix - Professional, minimal color
    row1_col1, row1_col2 = st.columns(2)

    with row1_col1:
        st.markdown("""
        <div style="background: rgba(251, 191, 36, 0.08); padding: 1.5rem; border-radius: 10px;
                    border: 1px solid rgba(251, 191, 36, 0.2); border-left: 4px solid #F59E0B;">
            <div style="display: flex; align-items: center; margin-bottom: 0.75rem;">
                <span style="font-size: 1.5rem; margin-right: 0.75rem;">⚠️</span>
                <h4 style="color: #92400E; margin: 0; font-weight: 600;">INVEST</h4>
            </div>
            <p style="color: #78350F; font-size: 0.8125rem; font-weight: 600; margin: 0.5rem 0;">
                High Business Value · Low Technical Health
            </p>
            <p style="color: #92400E; font-size: 0.875rem; margin: 0.75rem 0 0 0; line-height: 1.5;">
                Critical business applications with technical debt requiring urgent modernization to prevent business risk.
            </p>
        </div>
        """, unsafe_allow_html=True)

    with row1_col2:
        st.markdown("""
        <div style="background: rgba(16, 185, 129, 0.08); padding: 1.5rem; border-radius: 10px;
                    border: 1px solid rgba(16, 185, 129, 0.2); border-left: 4px solid #10B981;">
            <div style="display: flex; align-items: center; margin-bottom: 0.75rem;">
                <span style="font-size: 1.5rem; margin-right: 0.75rem;">⭐</span>
                <h4 style="color: #065F46; margin: 0; font-weight: 600;">EVOLVE</h4>
            </div>
            <p style="color: #064E3B; font-size: 0.8125rem; font-weight: 600; margin: 0.5rem 0;">
                High Business Value · High Technical Health
            </p>
            <p style="color: #065F46; font-size: 0.875rem; margin: 0.75rem 0 0 0; line-height: 1.5;">
                Strategic assets delivering strong value on solid foundation - continue investment to enhance capabilities.
            </p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='margin: 1.5rem 0;'></div>", unsafe_allow_html=True)

    row2_col1, row2_col2 = st.columns(2)

    with row2_col1:
        st.markdown("""
        <div style="background: rgba(239, 68, 68, 0.08); padding: 1.5rem; border-radius: 10px;
                    border: 1px solid rgba(239, 68, 68, 0.2); border-left: 4px solid #EF4444;">
            <div style="display: flex; align-items: center; margin-bottom: 0.75rem;">
                <span style="font-size: 1.5rem; margin-right: 0.75rem;">🗑️</span>
                <h4 style="color: #991B1B; margin: 0; font-weight: 600;">ELIMINATE</h4>
            </div>
            <p style="color: #7F1D1D; font-size: 0.8125rem; font-weight: 600; margin: 0.5rem 0;">
                Low Business Value · Low Technical Health
            </p>
            <p style="color: #991B1B; font-size: 0.875rem; margin: 0.75rem 0 0 0; line-height: 1.5;">
                Candidates for decommissioning, consolidation, or replacement - minimal value with high maintenance cost.
            </p>
        </div>
        """, unsafe_allow_html=True)

    with row2_col2:
        st.markdown("""
        <div style="background: rgba(59, 130, 246, 0.08); padding: 1.5rem; border-radius: 10px;
                    border: 1px solid rgba(59, 130, 246, 0.2); border-left: 4px solid #3B82F6;">
            <div style="display: flex; align-items: center; margin-bottom: 0.75rem;">
                <span style="font-size: 1.5rem; margin-right: 0.75rem;">🔧</span>
                <h4 style="color: #1E3A8A; margin: 0; font-weight: 600;">MAINTAIN</h4>
            </div>
            <p style="color: #1E40AF; font-size: 0.8125rem; font-weight: 600; margin: 0.5rem 0;">
                Low Business Value · High Technical Health
            </p>
            <p style="color: #1E3A8A; font-size: 0.875rem; margin: 0.75rem 0 0 0; line-height: 1.5;">
                Technically sound but low strategic value - maintain with minimal investment, consider for consolidation.
            </p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

    # Data Sources
    st.markdown("### Assessment Data Sources")
    st.markdown("""
    <div style="background: #F9FAFB; padding: 1.5rem; border-radius: 12px; border: 1px solid #E5E7EB;">
        <p style="color: #4B5563; font-size: 0.9375rem; line-height: 1.6; margin: 0;">
            Portfolio assessments combine <strong>structured questionnaires</strong> (standardized application attributes),
            <strong>stakeholder meeting transcripts</strong> (qualitative insights from business owners and IT teams),
            and <strong>expert notes</strong> from portfolio review sessions. AI-powered analysis extracts answers
            to master questions across all synergy blocks, with confidence scoring to prioritize high-quality data.
        </p>
    </div>
    """, unsafe_allow_html=True)


# ==================== PAGE: DASHBOARD (REMOVED - Replaced by Analyses page) ====================
def page_dashboard():
    # DEPRECATED - This page has been replaced by Introduction + Analyses
    st.warning("⚠️ This page has been deprecated. Please use Introduction or Analyses instead.")
    return
    # Hero Section
    st.markdown("""
    <div style="margin-bottom: 2rem;">
        <h1 style="font-size: 2.5rem; font-weight: 800; color: #111827; margin-bottom: 0.5rem; letter-spacing: -0.025em;">
            Avangrid APM Platform
        </h1>
        <p style="font-size: 1.125rem; color: #6B7280; margin: 0;">
            Creating the "Self-Driving Enterprise": From functional efficiency to autonomous value engines.
        </p>
    </div>
    """, unsafe_allow_html=True)

    session = get_session()

    try:
        # Get all applications
        apps = get_all_applications_from_db(session)

        if not apps:
            st.markdown("""
            <div style="background: #FFFFFF; border-radius: 12px; padding: 3rem; border: 1px solid #E5E7EB; text-align: center; margin: 2rem 0;">
                <div style="font-size: 3rem; margin-bottom: 1rem;">👋</div>
                <h3 style="color: #111827; margin-bottom: 0.5rem;">Welcome to Avangrid APM</h3>
                <p style="color: #6B7280; margin-bottom: 1.5rem;">Start by uploading your application questionnaire to begin portfolio analysis</p>
            </div>
            """, unsafe_allow_html=True)

            if st.button("📤 Go to Uploads", key="welcome_upload"):
                st.session_state.current_page = 'Uploads'
                st.rerun()
            return

        # Calculate metrics
        total_apps = len(apps)
        total_transcripts = session.query(MeetingTranscript).count()
        insights_count = session.query(Insight).count()

        # Get scores
        avg_bvi = 0
        avg_thi = 0
        evolve_count = 0
        invest_count = 0
        maintain_count = 0
        eliminate_count = 0

        for app in apps:
            scores_data = session.query(SynergyScore).filter_by(
                application_id=app.id,
                approved=True
            ).all()

            if scores_data:
                scores = {s.block_name: s.score for s in scores_data}
                bvi, thi = calculate_bvi_thi(scores, {b: {'Weight': w} for b, w in get_current_weights().items()})
                avg_bvi += bvi
                avg_thi += thi

                rec = get_recommendation(bvi, thi)
                if rec == "EVOLVE":
                    evolve_count += 1
                elif rec == "INVEST":
                    invest_count += 1
                elif rec == "MAINTAIN":
                    maintain_count += 1
                elif rec == "ELIMINATE":
                    eliminate_count += 1

        if total_apps > 0:
            avg_bvi = round(avg_bvi / total_apps, 1)
            avg_thi = round(avg_thi / total_apps, 1)

        # Key Metrics
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric("Applications", total_apps)

        with col2:
            st.metric("Average BVI", f"{avg_bvi:.1f}")

        with col3:
            st.metric("Average THI", f"{avg_thi:.1f}")

        with col4:
            st.metric("AI Insights", insights_count)

        st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

        # Feature Cards Section - Modern HTML Components
        col1, col2, col3 = st.columns(3)

        with col1:
            render_modern_card(
                icon="💼",
                title="Portfolio Intelligence",
                items=["AI-Powered Analysis", "Strategic Assessments", "Value Optimization"],
                color="orange"
            )

        with col2:
            render_modern_card(
                icon="📊",
                title="Advanced Analytics",
                items=["BVI/THI Scoring", "2x2 Matrix Mapping", "Strategic Roadmaps"],
                color="blue"
            )

        with col3:
            render_modern_card(
                icon="🤖",
                title="Intelligent Automation",
                items=["Transcript Analysis", "Auto Score Suggestions", "Insight Generation"],
                color="green"
            )

        st.markdown("<div style='margin: 3rem 0;'></div>", unsafe_allow_html=True)

        # Portfolio Overview Section
        st.markdown("""
        <h2 style="font-size: 1.5rem; font-weight: 700; color: #111827; margin-bottom: 1.5rem;">
            Portfolio Overview
        </h2>
        """, unsafe_allow_html=True)

        col1, col2 = st.columns([1, 1])

        with col1:
            # Recommendation Distribution
            rec_data = pd.DataFrame({
                'Recommendation': ['EVOLVE', 'INVEST', 'MAINTAIN', 'ELIMINATE'],
                'Count': [evolve_count, invest_count, maintain_count, eliminate_count],
            })

            fig = px.pie(
                rec_data,
                values='Count',
                names='Recommendation',
                color='Recommendation',
                color_discrete_map={
                    'EVOLVE': '#10B981',
                    'INVEST': '#F59E0B',
                    'MAINTAIN': '#0066B3',
                    'ELIMINATE': '#EF4444'
                },
                hole=0.4
            )
            fig.update_traces(textposition='inside', textinfo='percent+label')
            fig.update_layout(
                showlegend=True,
                height=350,
                margin=dict(t=30, b=30, l=30, r=30)
            )
            st.plotly_chart(fig, width="stretch")

        with col2:
            # Key Highlights
            st.markdown("""
            <div style="background: #FFFFFF; border-radius: 12px; padding: 1.5rem; border: 1px solid #E5E7EB; height: 350px; display: flex; flex-direction: column; justify-content: space-around;">
                <div class="metric-highlight">
                    <div class="metric-highlight-value">""" + str(total_apps) + """</div>
                    <div class="metric-highlight-label">Total Applications</div>
                </div>
                <div class="metric-highlight">
                    <div class="metric-highlight-value">""" + str(total_transcripts) + """</div>
                    <div class="metric-highlight-label">Transcripts Analyzed</div>
                </div>
                <div class="metric-highlight">
                    <div class="metric-highlight-value">""" + str(insights_count) + """</div>
                    <div class="metric-highlight-label">AI-Generated Insights</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

        # Quick Actions
        st.markdown("""
        <h2 style="font-size: 1.5rem; font-weight: 700; color: #111827; margin-bottom: 1rem;">
            Quick Actions
        </h2>
        """, unsafe_allow_html=True)

        col1, col2, col3 = st.columns(3)

        with col1:
            if st.button("📤 Upload Data", width="stretch", key="qa_upload"):
                st.session_state.current_page = 'Uploads'
                st.rerun()

        with col2:
            if st.button("📊 View Analytics", width="stretch", key="qa_analytics"):
                st.session_state.current_page = 'Analyses'
                st.rerun()

        with col3:
            if st.button("💡 Generate Insights", width="stretch", key="qa_insights"):
                st.session_state.current_page = 'Insights'
                st.rerun()

        with col3:
            if st.button("📱 View Applications", width="stretch"):
                st.session_state.current_page = 'Applications'
                st.rerun()

    finally:
        close_session(session)


# ==================== PAGE: UPLOADS ====================
def page_uploads():
    st.title("📤 Uploads")
    st.markdown("### Upload questionnaires and meeting transcripts")

    tab1, tab2 = st.tabs(["📋 Questionnaire", "🎙️ Transcripts"])

    # Tab 1: Questionnaire Upload
    with tab1:
        st.markdown("#### Upload Application Questionnaire Excel")
        st.info("Upload your completed questionnaire Excel file. The system will automatically parse all applications and their answers.")

        uploaded_file = st.file_uploader(
            "Choose Excel file",
            type=['xlsx', 'xls'],
            key='questionnaire_upload'
        )

        if uploaded_file:
            with st.spinner("Parsing questionnaire..."):
                apps_data = parse_questionnaire_excel(uploaded_file)

            if apps_data:
                st.success(f"✅ Found {len(apps_data)} applications!")

                # Preview
                st.markdown("##### Preview:")
                for app in apps_data[:5]:  # Show first 5
                    st.markdown(f"- **{app['name']}** ({len(app['answers'])} answers)")

                if len(apps_data) > 5:
                    st.markdown(f"... and {len(apps_data) - 5} more")

                # Save to database
                if st.button("💾 Save to Database", type="primary"):
                    session = get_session()
                    try:
                        # Step 1: Save applications
                        progress = st.progress(0)
                        status_text = st.empty()

                        status_text.text("📥 Saving applications to database...")
                        for idx, app_data in enumerate(apps_data):
                            save_application_to_db(app_data, session)
                            progress.progress((idx + 1) / len(apps_data) * 0.5)  # First 50%

                        st.success(f"✅ Successfully saved {len(apps_data)} applications!")

                        # Step 2: Auto-calculate scores
                        status_text.text("🤖 Auto-calculating scores with AI...")
                        calculated_count = 0

                        for idx, app_data in enumerate(apps_data):
                            app_name = app_data['name']

                            # Find the application in database
                            app = session.query(Application).filter_by(name=app_name).first()

                            if app:
                                # Check if manually edited scores exist (don't recalculate those)
                                manual_scores = session.query(SynergyScore).filter(
                                    SynergyScore.application_id == app.id,
                                    SynergyScore.approved == True,
                                    SynergyScore.approved_by != 'auto_ai_generated'
                                ).first()

                                if manual_scores:
                                    # User has manually edited scores - don't overwrite
                                    pass
                                else:
                                    # Check if questionnaire has enough complete answers
                                    qa_answers = session.query(QuestionnaireAnswer).filter_by(
                                        application_id=app.id
                                    ).all()

                                    # Count complete answers (not empty)
                                    complete_answers = [qa for qa in qa_answers if qa.answer_text and len(qa.answer_text.strip()) >= 5]
                                    total_answers = len(qa_answers)

                                    # Only calculate if we have at least 50% complete answers
                                    if total_answers > 0 and (len(complete_answers) / total_answers) >= 0.5:
                                        # Delete old auto-generated scores to recalculate
                                        session.query(SynergyScore).filter_by(
                                            application_id=app.id,
                                            approved_by='auto_ai_generated'
                                        ).delete()

                                        session.commit()  # Commit deletion before calculating new scores

                                        # Gather questionnaire data (only complete answers)
                                        questionnaire_dict = {}
                                        for qa in complete_answers:
                                            questionnaire_dict[qa.question_text] = {'a': qa.answer_text, 's': qa.score}

                                        # Generate scores with AI
                                        if questionnaire_dict:
                                            try:
                                                result = suggest_scores(questionnaire_dict, [])

                                                if result.get('scores'):
                                                    for block_name, score_data in result['scores'].items():
                                                        score = SynergyScore(
                                                            id=str(uuid.uuid4()),
                                                            application_id=app.id,
                                                            block_name=block_name,
                                                            score=score_data['score'],
                                                            suggested_by='ai_questionnaire',
                                                            confidence=score_data.get('confidence', 0.8),
                                                            rationale=score_data.get('rationale', ''),
                                                            approved=True,
                                                            approved_by='auto_ai_generated',
                                                            approved_at=datetime.now(timezone.utc)
                                                        )
                                                        session.add(score)

                                                    calculated_count += 1
                                            except Exception as e:
                                                st.warning(f"⚠️ Could not calculate scores for {app_name}: {str(e)}")
                                    else:
                                        # Not enough complete answers - skip score calculation
                                        pass

                            progress.progress(0.5 + ((idx + 1) / len(apps_data) * 0.5))  # Second 50%

                        session.commit()

                        if calculated_count > 0:
                            st.success(f"🎉 Auto-calculated scores for {calculated_count} applications!")

                        status_text.empty()
                        st.balloons()

                    finally:
                        close_session(session)
            else:
                st.warning("No applications found in the file.")

    # Tab 2: Transcripts Upload
    with tab2:
        st.markdown("#### Upload Meeting Transcripts")
        st.info("📁 Upload all transcripts at once. The application name will be automatically detected from the filename.\n\n**Format:** `ApplicationName - ...description.docx`")

        session = get_session()

        try:
            # Get all applications
            apps = get_all_applications_from_db(session)

            if not apps:
                st.warning("⚠️ Please upload a questionnaire first to create applications.")
                return

            # Create a mapping of app names for quick lookup (case-insensitive)
            app_dict = {app.name.strip().lower(): app for app in apps}

            # Upload transcript files
            uploaded_transcripts = st.file_uploader(
                "Choose all transcript files",
                type=['txt', 'pdf', 'docx'],
                accept_multiple_files=True,
                key='transcript_upload',
                help="Upload all transcript files. App name will be detected from filename (before ' - ')"
            )

            if uploaded_transcripts:
                st.markdown("---")
                col1, col2 = st.columns(2)

                with col1:
                    quick_upload = st.button("📁 Quick Upload (Save Only)", type="secondary", width="stretch",
                                            help="Save files to database without processing. Process later in Batch Operations.")
                with col2:
                    process_now = st.button("🤖 Upload & Process Now", type="primary", width="stretch",
                                           help="Save and process all files immediately with AI.")

            # Quick Upload (Save only, no processing)
            if uploaded_transcripts and quick_upload:
                st.markdown(f"### 📁 Uploading {len(uploaded_transcripts)} file(s)...")

                progress_bar = st.progress(0)
                saved_count = 0
                skipped_count = 0
                error_count = 0

                for idx, uploaded_file in enumerate(uploaded_transcripts):
                    try:
                        filename = uploaded_file.name

                        # Extract app name
                        if " - " in filename:
                            app_name_from_file = filename.split(" - ")[0].strip()
                        else:
                            # No separator - use filename without extension as app name
                            app_name_from_file = filename.rsplit('.', 1)[0].strip()

                        # Find app using smart matching
                        matched_app, match_type = find_matching_application(app_name_from_file, app_dict)
                        if not matched_app:
                            st.warning(f"⚠️ {filename}: App '{app_name_from_file}' not found")
                            error_count += 1
                            progress_bar.progress((idx + 1) / len(uploaded_transcripts))
                            continue

                        if match_type != 'exact':
                            st.info(f"ℹ️ {filename}: Matched to '{matched_app.name}' ({match_type})")

                        # Check if exists
                        existing = session.query(MeetingTranscript).filter_by(
                            application_id=matched_app.id,
                            file_name=uploaded_file.name
                        ).first()

                        if existing:
                            st.info(f"⏭️ {filename}: Already exists")
                            skipped_count += 1
                        else:
                            # Read and save
                            transcript_text = read_transcript_file(uploaded_file)
                            if transcript_text:
                                transcript = MeetingTranscript(
                                    id=str(uuid.uuid4()),
                                    application_id=matched_app.id,
                                    file_name=uploaded_file.name,
                                    transcript_text=transcript_text,
                                    processed=False
                                )
                                session.add(transcript)
                                session.commit()
                                saved_count += 1
                            else:
                                st.error(f"❌ {filename}: Could not read file")
                                error_count += 1

                        progress_bar.progress((idx + 1) / len(uploaded_transcripts))

                    except Exception as e:
                        st.error(f"❌ {filename}: {str(e)}")
                        error_count += 1

                st.markdown("---")
                st.markdown("### 📊 Upload Summary")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("✅ Saved", saved_count)
                with col2:
                    st.metric("⏭️ Skipped", skipped_count)
                with col3:
                    st.metric("❌ Errors", error_count)

                if saved_count > 0:
                    st.success(f"🎉 {saved_count} files saved! Go to **Batch Operations** to process them.")

            # Process Immediately (current behavior)
            elif uploaded_transcripts and process_now:
                import time
                start_time = time.time()

                processed_count = 0
                skipped_count = 0
                error_count = 0
                unmatched_count = 0

                # Track which apps were processed
                apps_processed = {}

                with st.status(f"🤖 Processing {len(uploaded_transcripts)} transcript(s)...", expanded=True) as status:
                    progress_bar = st.progress(0)

                    for idx, uploaded_file in enumerate(uploaded_transcripts):
                        # Update progress percentage
                        progress = (idx) / len(uploaded_transcripts)
                        progress_bar.progress(progress)

                        # Estimate time remaining
                        if idx > 0:
                            elapsed = time.time() - start_time
                            avg_time = elapsed / idx
                            remaining = avg_time * (len(uploaded_transcripts) - idx)
                            status.update(label=f"🤖 Processing {idx+1}/{len(uploaded_transcripts)}: {uploaded_file.name} (~{int(remaining)}s remaining)")
                        else:
                            status.update(label=f"🤖 Processing 1/{len(uploaded_transcripts)}: {uploaded_file.name}")

                        try:
                            # Extract application name from filename
                            # Handles multi-part names like "Bentley - PLS-CADD - Application Assessment..."
                            filename = uploaded_file.name
                            name_no_ext = filename.rsplit('.', 1)[0].strip()

                            # Smart suffix removal: find description suffix boundary
                            # This preserves multi-part app names (e.g., "Bentley - PLS-CADD")
                            suffix_idx = name_no_ext.lower().find(' - application assessment')
                            if suffix_idx > 0:
                                app_name_from_file = name_no_ext[:suffix_idx].strip()
                            elif " - " in name_no_ext:
                                # Fallback: take first segment
                                app_name_from_file = name_no_ext.split(" - ")[0].strip()
                            else:
                                app_name_from_file = name_no_ext

                            # Find matching application using smart matching
                            matched_app, match_type = find_matching_application(app_name_from_file, app_dict)

                            if not matched_app:
                                st.warning(f"⚠️ {filename}: Application '{app_name_from_file}' not found in database")
                                unmatched_count += 1
                                continue

                            if match_type != 'exact':
                                st.info(f"ℹ️ {filename}: Matched to '{matched_app.name}' ({match_type})")

                            # Track apps being processed
                            if matched_app.name not in apps_processed:
                                apps_processed[matched_app.name] = []
                            apps_processed[matched_app.name].append(filename)

                            # Check if transcript already exists and was processed
                            existing_transcript = session.query(MeetingTranscript).filter_by(
                                application_id=matched_app.id,
                                file_name=uploaded_file.name
                            ).first()

                            if existing_transcript and existing_transcript.processed:
                                st.info(f"⏭️ Skipping {uploaded_file.name} (already processed)")
                                skipped_count += 1
                                continue

                            # Read transcript
                            transcript_text = read_transcript_file(uploaded_file)

                            if not transcript_text:
                                st.error(f"❌ Could not read {uploaded_file.name}")
                                error_count += 1
                                continue

                            # Create or use existing transcript
                            if existing_transcript:
                                transcript = existing_transcript
                                transcript.transcript_text = transcript_text
                            else:
                                transcript = MeetingTranscript(
                                    id=str(uuid.uuid4()),
                                    application_id=matched_app.id,
                                    file_name=uploaded_file.name,
                                    transcript_text=transcript_text,
                                    processed=False
                                )
                                session.add(transcript)

                            session.commit()

                            # Extract answers using AI
                            st.write(f"🤖 AI analyzing for **{matched_app.name}**...")
                            result = extract_answers_from_transcript(transcript_text, matched_app.name)

                            if result.get('answers'):
                                answer_count = 0
                                # Save extracted answers (avoid duplicates)
                                for answer_data in result['answers']:
                                    if answer_data.get('answer') and answer_data.get('confidence', 0) > 0.3:
                                        # Check if this answer already exists for this transcript
                                        existing_answer = session.query(TranscriptAnswer).filter_by(
                                            transcript_id=transcript.id,
                                            question_text=answer_data['question']
                                        ).first()

                                        if not existing_answer:
                                            ta = TranscriptAnswer(
                                                id=str(uuid.uuid4()),
                                                application_id=matched_app.id,
                                                transcript_id=transcript.id,
                                                question_text=answer_data['question'],
                                                answer_text=answer_data['answer'],
                                                confidence_score=answer_data['confidence'],
                                                synergy_block=answer_data.get('synergy_block', 'Unknown')
                                            )
                                            session.add(ta)
                                            answer_count += 1

                                transcript.processed = True
                                session.commit()

                                st.success(f"✅ **{matched_app.name}** - {uploaded_file.name}: Extracted {answer_count} new answers")
                                processed_count += 1

                            else:
                                st.warning(f"⚠️ {uploaded_file.name}: No answers extracted")
                                processed_count += 1

                        except Exception as e:
                            st.error(f"❌ Error processing {uploaded_file.name}: {str(e)}")
                            error_count += 1

                    # Update final progress
                    progress_bar.progress(1.0)
                    total_time = time.time() - start_time
                    status.update(label=f"✅ Complete! Processed {len(uploaded_transcripts)} files in {int(total_time)}s", state="complete", expanded=False)

                # Final summary - Make it very prominent
                st.markdown("---")
                st.success("🎉 **PROCESSING COMPLETE!**")
                st.markdown("### 📊 Summary")

                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("✅ Processed", processed_count)
                with col2:
                    st.metric("⏭️ Skipped", skipped_count)
                with col3:
                    st.metric("❌ Errors", error_count)
                with col4:
                    st.metric("⚠️ Unmatched", unmatched_count)

                # Show which applications were processed
                if apps_processed:
                    st.markdown("#### Applications Updated:")
                    for app_name, files in apps_processed.items():
                        with st.expander(f"📱 {app_name} ({len(files)} files)"):
                            for file in files:
                                st.markdown(f"- {file}")

                if processed_count > 0:
                    st.success("🎉 All transcripts processed!")

                    # Auto-recalculate scores for affected applications
                    if apps_processed:
                        st.markdown("---")
                        st.markdown("### 🤖 Auto-Recalculating Scores")
                        recalc_progress = st.progress(0)
                        recalc_status = st.empty()

                        recalculated_count = 0
                        app_names = list(apps_processed.keys())

                        for idx, app_name in enumerate(app_names):
                            recalc_status.text(f"Recalculating {idx+1}/{len(app_names)}: {app_name}")

                            app = session.query(Application).filter_by(name=app_name).first()

                            if app:
                                # Gather questionnaire data
                                questionnaire_dict = {}
                                qa_answers = session.query(QuestionnaireAnswer).filter_by(
                                    application_id=app.id
                                ).all()
                                for qa in qa_answers:
                                    questionnaire_dict[qa.question_text] = {'a': qa.answer_text, 's': qa.score}

                                # Gather transcript data
                                transcript_list = []
                                ta_answers = session.query(TranscriptAnswer).filter_by(
                                    application_id=app.id
                                ).all()
                                for ta in ta_answers:
                                    transcript_list.append({
                                        'question': ta.question_text,
                                        'answer': ta.answer_text,
                                        'confidence': ta.confidence_score
                                    })

                                # Recalculate scores
                                if questionnaire_dict or transcript_list:
                                    try:
                                        result = suggest_scores(questionnaire_dict, transcript_list)

                                        if result.get('scores'):
                                            # Delete old auto-generated scores
                                            session.query(SynergyScore).filter_by(
                                                application_id=app.id,
                                                approved_by='auto_ai_generated'
                                            ).delete()

                                            # Insert new scores
                                            for block_name, score_data in result['scores'].items():
                                                score = SynergyScore(
                                                    id=str(uuid.uuid4()),
                                                    application_id=app.id,
                                                    block_name=block_name,
                                                    score=score_data['score'],
                                                    suggested_by='ai_combined',
                                                    confidence=score_data.get('confidence', 0.8),
                                                    rationale=score_data.get('rationale', ''),
                                                    approved=True,
                                                    approved_by='auto_ai_generated',
                                                    approved_at=datetime.now(timezone.utc)
                                                )
                                                session.add(score)

                                            recalculated_count += 1
                                    except Exception as e:
                                        st.warning(f"⚠️ Could not recalculate scores for {app_name}: {str(e)}")

                            recalc_progress.progress((idx + 1) / len(app_names))

                        session.commit()

                        if recalculated_count > 0:
                            st.success(f"🎉 Recalculated scores for {recalculated_count} applications!")

                        recalc_status.empty()

                    st.balloons()

        finally:
            close_session(session)


# ==================== PAGE: APPLICATIONS ====================
def page_applications():
    """Simplified Applications page with table view grouped by synergy blocks"""

    st.title("📱 Applications Portfolio")
    st.markdown("Review and edit application scores by synergy block")

    session = get_session()

    try:
        apps = get_all_applications_from_db(session)
        apps = [app for app in apps if app.name != "Questions Template"]

        if not apps:
            st.info("📋 No applications found. Upload a questionnaire to get started.")
            return

        # Application selector
        app_names = [app.name for app in apps]
        selected_app_name = st.selectbox("Select Application", app_names, key="selected_app")
        selected_app = next(app for app in apps if app.name == selected_app_name)

        st.markdown("---")

        # Get scores
        scores_data = session.query(SynergyScore).filter_by(
            application_id=selected_app.id,
            approved=True
        ).all()

        scores = {s.block_name: s.score for s in scores_data} if scores_data else {}

        # Calculate BVI/THI
        if scores:
            bvi, thi = calculate_bvi_thi(scores, {b: {'Weight': w} for b, w in get_current_weights().items()})
            rec = get_recommendation(bvi, thi)

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("BVI", f"{bvi:.1f}")
            with col2:
                st.metric("THI", f"{thi:.1f}")
            with col3:
                st.metric("Decision", rec)

        st.markdown("---")

        # David's Key Insights (show ONCE per application, not per block)
        from database import DavidNote
        david_insights = session.query(DavidNote).filter_by(
            application_id=selected_app.id,
            note_type='insight'
        ).first()

        if david_insights:
            st.markdown("### 💡 David's Key Insights")
            st.markdown("""
            <div style="background: #FEF3C7; padding: 1.5rem; border-radius: 8px; border-left: 4px solid #F59E0B; margin-bottom: 1.5rem;">
                <p style="margin: 0; color: #92400E; white-space: pre-wrap;">
            """ + david_insights.answer_text + """
                </p>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")

        # Display by synergy block
        current_weights = get_current_weights()
        for block_name, block_info in SYNERGY_BLOCKS.items():
            w = current_weights.get(block_name, block_info['Weight'])
            with st.expander(f"📋 {block_name} (Weight: {w}%)", expanded=False):
                # Get questions for this block
                block_questions = MASTER_QUESTIONS.get(block_name, [])

                # Get questionnaire answers
                qa_answers = session.query(QuestionnaireAnswer).filter_by(
                    application_id=selected_app.id,
                    synergy_block=block_name
                ).all()
                qa_dict = {qa.question_text: qa.answer_text for qa in qa_answers}

                # Get transcript answers
                ta_answers = session.query(TranscriptAnswer).filter_by(
                    application_id=selected_app.id,
                    synergy_block=block_name
                ).all()
                ta_dict = {ta.question_text: (ta.answer_text, ta.confidence_score) for ta in ta_answers}

                # Get David's notes
                david_answers = session.query(DavidNote).filter_by(
                    application_id=selected_app.id,
                    synergy_block=block_name,
                    note_type='answer'
                ).all()
                david_dict = {dn.question_text: dn.answer_text for dn in david_answers}

                # Build table data - Show ALL questions (with or without answers)
                table_data = []
                for question in block_questions:
                    qa_answer = qa_dict.get(question, "-")
                    ta_info = ta_dict.get(question, ("-", 0))
                    ta_answer = ta_info[0] if isinstance(ta_info, tuple) else "-"
                    david_answer = david_dict.get(question, "-")
                    confidence = f"{ta_info[1]:.0%}" if isinstance(ta_info, tuple) and ta_info[1] > 0 else "-"

                    # Merge transcript and David notes intelligently
                    combined_answer = ""
                    if ta_answer != "-" and david_answer != "-":
                        # Both exist - check if they're similar
                        if ta_answer.strip().lower() == david_answer.strip().lower():
                            # Same answer, show once
                            combined_answer = ta_answer
                        else:
                            # Different answers, show both with labels
                            combined_answer = f"Transcript: {ta_answer}\n\nDavid: {david_answer}"
                    elif ta_answer != "-":
                        combined_answer = ta_answer
                    elif david_answer != "-":
                        combined_answer = f"David: {david_answer}"
                    else:
                        combined_answer = "-"

                    table_data.append({
                        "Question": question,  # No truncation
                        "Questionnaire": qa_answer,  # No truncation
                        "Transcript and David notes": combined_answer,  # Combined, no truncation
                        "Confidence": confidence
                    })

                if table_data:
                    df = pd.DataFrame(table_data)

                    # Calculate dynamic height: ~35px per row (header + data rows), min 150px, max 800px
                    row_height = 35
                    calculated_height = min(max((len(table_data) + 1) * row_height, 150), 800)

                    # Configure column widths and text wrapping
                    st.dataframe(
                        df,
                        width="stretch",
                        height=calculated_height,  # Dynamic height based on number of rows
                        column_config={
                            "Question": st.column_config.TextColumn(
                                "Question",
                                width="medium",
                                help="Master question from questionnaire"
                            ),
                            "Questionnaire": st.column_config.TextColumn(
                                "Questionnaire",
                                width="large",
                                help="Answer from uploaded questionnaire"
                            ),
                            "Transcript and David notes": st.column_config.TextColumn(
                                "Transcript and David notes",
                                width="large",
                                help="Combined answers from transcript and David's notes"
                            ),
                            "Confidence": st.column_config.TextColumn(
                                "Confidence",
                                width="small",
                                help="AI confidence score for transcript answer"
                            )
                        }
                    )
                else:
                    st.info("No responses for this block")

                # Score editor
                current_score = scores.get(block_name, 3)
                new_score = st.slider(
                    f"Score for {block_name}",
                    min_value=1,
                    max_value=5,
                    value=current_score,
                    key=f"score_{block_name}_{selected_app.id}"
                )

                if new_score != current_score:
                    if st.button(f"Save Score", key=f"save_{block_name}_{selected_app.id}"):
                        # Update score
                        existing_score = session.query(SynergyScore).filter_by(
                            application_id=selected_app.id,
                            block_name=block_name,
                            approved=True
                        ).first()

                        if existing_score:
                            existing_score.score = new_score
                            existing_score.approved_by = 'manual_edit'
                        else:
                            new_score_obj = SynergyScore(
                                id=str(uuid.uuid4()),
                                application_id=selected_app.id,
                                block_name=block_name,
                                score=new_score,
                                suggested_by='manual',
                                confidence=1.0,
                                rationale='Manually set score',
                                approved=True,
                                approved_by='manual_edit',
                                approved_at=datetime.now(timezone.utc)
                            )
                            session.add(new_score_obj)

                        session.commit()
                        st.success(f"✅ Score updated for {block_name}!")
                        st.rerun()

        # Strategic Insights Section
        st.markdown("---")
        st.markdown("### 💡 Strategic Insights")

        # Gather insights from multiple sources
        from database import AppInsight, DavidNote

        # Priority 1: David's insights (most valuable)
        david_insights = session.query(DavidNote).filter_by(
            application_id=selected_app.id,
            note_type='insight'
        ).first()

        # Priority 2: David's answers (rich context)
        david_answers = session.query(DavidNote).filter_by(
            application_id=selected_app.id,
            note_type='answer'
        ).all()

        # Priority 3: AI-generated insights (from generate_insights.py)
        app_insights = session.query(AppInsight).filter_by(
            application_id=selected_app.id
        ).all()

        # Priority 4: Transcript answers (user feedback)
        transcript_answers = session.query(TranscriptAnswer).filter_by(
            application_id=selected_app.id
        ).all()

        has_any_insights = david_insights or david_answers or app_insights or transcript_answers

        if has_any_insights:
            st.markdown("""
            <div style="background: linear-gradient(135deg, #E87722 10%, #FF8C42 100%);
                        padding: 1rem; border-radius: 8px; margin-bottom: 1rem;">
                <p style="color: white; margin: 0; font-size: 0.9rem;">
                    ✨ Strategic insights from multiple sources: David's meeting notes, AI analysis, transcripts, and questionnaires
                </p>
            </div>
            """, unsafe_allow_html=True)

            # ===========================================
            # PRIORITY 1: David's General Insights
            # ===========================================
            if david_insights:
                with st.expander("🎯 **David's Key Insights** (Source: Stakeholder Meetings)", expanded=True):
                    st.markdown("""
                    <div style="background: #FEF3C7; padding: 1rem; border-radius: 8px; border-left: 4px solid #F59E0B;">
                        <p style="margin: 0; color: #92400E;">
                            <strong>📍 Source:</strong> Detailed notes from David's meetings with stakeholders and SMEs
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

                    st.markdown("#### Executive Summary")
                    st.info(david_insights.answer_text)

            # ===========================================
            # PRIORITY 2: David's Detailed Answers
            # ===========================================
            if david_answers:
                with st.expander(f"📝 **David's Detailed Analysis** ({len(david_answers)} insights)", expanded=False):
                    st.markdown("""
                    <div style="background: #FEF3C7; padding: 1rem; border-radius: 8px; border-left: 4px solid #F59E0B;">
                        <p style="margin: 0; color: #92400E;">
                            <strong>📍 Source:</strong> Specific responses collected by David during interviews with users and managers
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

                    # Group by synergy block
                    david_by_block = {}
                    for da in david_answers:
                        block = da.synergy_block or "General"
                        if block not in david_by_block:
                            david_by_block[block] = []
                        david_by_block[block].append(da)

                    for block_name, answers in david_by_block.items():
                        st.markdown(f"##### 📊 {block_name}")
                        for answer in answers:
                            with st.container():
                                st.markdown(f"**❓ {answer.question_text}**")
                                st.markdown(f"""
                                <div style="background: white; padding: 0.75rem; border-radius: 6px; border-left: 3px solid #E87722; margin-bottom: 1rem;">
                                    {answer.answer_text}
                                </div>
                                """, unsafe_allow_html=True)

            # ===========================================
            # PRIORITY 3: High-Confidence Transcript Insights
            # ===========================================
            high_conf_transcripts = [ta for ta in transcript_answers if ta.confidence_score >= 0.7]
            if high_conf_transcripts:
                with st.expander(f"🎙️ **Transcript Insights** ({len(high_conf_transcripts)} with high confidence)", expanded=False):
                    st.markdown("""
                    <div style="background: #DBEAFE; padding: 1rem; border-radius: 8px; border-left: 4px solid #3B82F6;">
                        <p style="margin: 0; color: #1E3A8A;">
                            <strong>📍 Source:</strong> Extracted from meeting transcripts using AI (confidence >= 70%)
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

                    # Group by synergy block
                    trans_by_block = {}
                    for ta in high_conf_transcripts:
                        block = ta.synergy_block or "General"
                        if block not in trans_by_block:
                            trans_by_block[block] = []
                        trans_by_block[block].append(ta)

                    for block_name, answers in trans_by_block.items():
                        st.markdown(f"##### 📊 {block_name}")
                        for answer in answers[:3]:  # Show top 3 per block
                            conf_color = "🟢" if answer.confidence_score >= 0.8 else "🟡"
                            st.markdown(f"**❓ {answer.question_text}** {conf_color} {answer.confidence_score:.0%}")
                            st.markdown(f"> {answer.answer_text}")
                            st.markdown("---")

            # ===========================================
            # PRIORITY 4: AI-Generated Strategic Insights
            # ===========================================
            if app_insights:
                with st.expander(f"🤖 **AI-Generated Insights** ({len(app_insights)} analyses)", expanded=False):
                    st.markdown("""
                    <div style="background: #F3E8FF; padding: 1rem; border-radius: 8px; border-left: 4px solid #9333EA;">
                        <p style="margin: 0; color: #581C87;">
                            <strong>📍 Source:</strong> Strategic analysis generated by GPT-4o based on questionnaires, transcripts, and market research
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

                    # Group insights by type
                    insights_by_type = {}
                    for insight in app_insights:
                        if insight.insight_type not in insights_by_type:
                            insights_by_type[insight.insight_type] = []
                        insights_by_type[insight.insight_type].append(insight)

                    # Display each insight type
                    for insight_type, insights_list in insights_by_type.items():
                        insight_type_display = insight_type.replace('_', ' ').title()

                        st.markdown(f"##### 📊 {insight_type_display}")
                        for insight in insights_list:
                            try:
                                content = json.loads(insight.content)

                                if insight_type == 'capabilities':
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.markdown("**Strengths:**")
                                        for strength in content.get('strengths', []):
                                            st.markdown(f"✅ {strength}")
                                    with col2:
                                        st.markdown("**Limitations:**")
                                        for limitation in content.get('limitations', []):
                                            st.markdown(f"⚠️ {limitation}")

                                    if content.get('unique_value'):
                                        st.info(f"💎 Unique Value: {content['unique_value']}")

                                elif insight_type == 'user_satisfaction':
                                    st.markdown(f"**Sentiment:** {content.get('sentiment', 'N/A').upper()}")

                                    if content.get('pain_points'):
                                        st.markdown("**Pain Points:**")
                                        for pain in content['pain_points']:
                                            st.markdown(f"- {pain}")

                                    if content.get('key_quotes'):
                                        with st.expander("📢 Key User Quotes"):
                                            for quote in content['key_quotes']:
                                                st.markdown(f"> {quote}")

                                elif insight_type == 'technical_debt':
                                    severity = content.get('severity', 'unknown')
                                    severity_emoji = {'high': '🔴', 'medium': '🟡', 'low': '🟢'}.get(severity, '⚪')
                                    st.markdown(f"**Severity:** {severity_emoji} {severity.upper()}")

                                    if content.get('issues'):
                                        st.markdown("**Issues:**")
                                        for issue in content['issues']:
                                            st.markdown(f"- {issue}")

                                elif insight_type == 'integration_opportunities':
                                    if content.get('should_integrate_into'):
                                        st.success(f"🔗 Recommended: Integrate into **{content['should_integrate_into']}**")

                                    if content.get('can_consolidate_with'):
                                        st.markdown("**Could consolidate with:**")
                                        for app_name in content['can_consolidate_with']:
                                            st.markdown(f"- {app_name}")

                                elif insight_type == 'market_alternatives':
                                    if content.get('alternatives'):
                                        st.markdown("**Market Alternatives:**")
                                        for alt in content['alternatives']:
                                            st.markdown(f"- {alt}")

                                    if content.get('migration_path'):
                                        st.info(f"📋 Migration Path: {content['migration_path']}")

                                elif insight_type == 'strategic_recommendation':
                                    action = content.get('action', 'N/A')
                                    target = content.get('target', '')
                                    priority = content.get('priority', 'P3')
                                    impact = content.get('estimated_impact', 'unknown')
                                    complexity = content.get('complexity', 'unknown')

                                    # Color coding
                                    action_colors = {
                                        'EVOLVE': '🟢',
                                        'INVEST': '🔵',
                                        'MAINTAIN': '🟡',
                                        'ELIMINATE': '🔴',
                                        'INTEGRATE': '🟣',
                                        'MIGRATE': '🟠',
                                        'CONSOLIDATE': '🟤'
                                    }

                                    st.markdown(f"### {action_colors.get(action, '⚪')} {action}")
                                    if target:
                                        st.markdown(f"**Target:** {target}")

                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.metric("Priority", priority)
                                    with col2:
                                        st.metric("Impact", impact.upper())
                                    with col3:
                                        st.metric("Complexity", complexity.upper())

                                    st.markdown(f"**Rationale:** {content.get('rationale', 'N/A')}")

                                    # Action items
                                    if insight.action_items:
                                        st.markdown("**📋 Action Items:**")
                                        for item in insight.action_items:
                                            st.markdown(f"- **{item.get('action')}** ({item.get('owner')}) - {item.get('timeline')}")

                                    # Evidence
                                    if insight.evidence:
                                        with st.expander("📚 Supporting Evidence"):
                                            for evidence in insight.evidence:
                                                st.markdown(f"> {evidence}")

                            except json.JSONDecodeError:
                                st.error(f"Error parsing insight content for {insight_type}")
                            except Exception as e:
                                st.error(f"Error displaying insight: {e}")
        else:
            st.info("""
            💡 No strategic insights available yet.

            Run the insight generator to get AI-powered strategic recommendations:
            ```
            cd webapp
            python generate_insights.py
            ```

            This will analyze all transcripts, questionnaires, and market data to provide:
            - Capability assessments
            - User satisfaction analysis
            - Technical debt evaluation
            - Integration opportunities
            - Market alternatives
            - Strategic recommendations with action items
            """)

    finally:
        close_session(session)
def page_analyses():
    st.title("📈 Strategic Analyses")

    session = get_session()

    try:
        apps = get_all_applications_from_db(session)

        if not apps:
            st.info("No applications found.")
            return

        # Get all apps with scores and extended roadmap data
        apps_with_scores = []
        for app in apps:
            # Check if questionnaire is mostly empty (95%+ empty responses)
            qa_answers = session.query(QuestionnaireAnswer).filter_by(
                application_id=app.id
            ).all()

            if qa_answers:
                # Count empty or very short answers
                empty_count = sum(1 for qa in qa_answers if not qa.answer_text or len(qa.answer_text.strip()) < 5)
                empty_percentage = (empty_count / len(qa_answers)) * 100

                # Skip apps with 95%+ empty questionnaires
                if empty_percentage >= 95:
                    continue

            scores_data = session.query(SynergyScore).filter_by(
                application_id=app.id,
                approved=True
            ).all()

            if scores_data:
                scores = {s.block_name: s.score for s in scores_data}
                bvi, thi = calculate_bvi_thi(scores, {b: {'Weight': w} for b, w in get_current_weights().items()})
                rec = get_recommendation(bvi, thi)

                # Get individual block scores
                arch_score = next((s.score for s in scores_data if s.block_name == 'Architecture'), 3)
                maint_score = next((s.score for s in scores_data if s.block_name == 'Maintainability'), 3)

                # Calculate dependencies
                deps_info = extract_dependencies_info(app.id, session)

                apps_with_scores.append({
                    'id': app.id,
                    'name': app.name,
                    'bvi': bvi,
                    'thi': thi,
                    'recommendation': rec,
                    'arch_score': arch_score,
                    'maint_score': maint_score,
                    'integration_count': deps_info['count'],
                    'integration_type': deps_info['type'],
                    'integration_systems': deps_info['systems'],
                    'dependencies_display': deps_info['display']
                })

        if not apps_with_scores:
            st.warning("No applications with approved scores found. Please review and approve scores in the Applications section.")
            return

        # Tabs for different analyses
        tab1, tab2 = st.tabs(["📊 2x2 Matrix", "💡 Insights"])

        with tab1:
            # 2x2 Matrix - Strategic Decision Matrix
            st.markdown("#### Portfolio Strategic Matrix")
            st.caption("Technical Health Index (THI) vs. Business Value Index (BVI)")

            # Create scatter plot
            df = pd.DataFrame(apps_with_scores)

            fig = go.Figure()

            # Add quadrant backgrounds (X=THI, Y=BVI - Standard APM Matrix)
            fig.add_shape(type="rect", x0=0, y0=0, x1=60, y1=60,
                         fillcolor="rgba(239, 68, 68, 0.1)", line_width=0)  # ELIMINATE (low THI, low BVI)
            fig.add_shape(type="rect", x0=60, y0=0, x1=100, y1=60,
                         fillcolor="rgba(59, 130, 246, 0.1)", line_width=0)  # MAINTAIN (high THI, low BVI)
            fig.add_shape(type="rect", x0=0, y0=60, x1=60, y1=100,
                         fillcolor="rgba(245, 158, 11, 0.1)", line_width=0)  # INVEST (low THI, high BVI)
            fig.add_shape(type="rect", x0=60, y0=60, x1=100, y1=100,
                         fillcolor="rgba(16, 185, 129, 0.1)", line_width=0)  # EVOLVE (high THI, high BVI)

            # Add quadrant lines
            fig.add_hline(y=60, line_dash="dash", line_color="gray")
            fig.add_vline(x=60, line_dash="dash", line_color="gray")

            # Add points (X=THI, Y=BVI)
            colors_map = {
                'EVOLVE': '#10B981',
                'INVEST': '#F59E0B',
                'MAINTAIN': '#3B82F6',
                'ELIMINATE': '#EF4444'
            }

            import math

            # All 9 plotly textposition options with their offset vectors (dx, dy)
            TEXT_POSITIONS = [
                ('top right',      ( 1,  1)),
                ('top left',       (-1,  1)),
                ('bottom right',   ( 1, -1)),
                ('bottom left',    (-1, -1)),
                ('middle right',   ( 1,  0)),
                ('middle left',    (-1,  0)),
                ('top center',     ( 0,  1)),
                ('bottom center',  ( 0, -1)),
            ]

            def assign_text_positions(apps_list):
                """Assign plotly textposition for each app to minimize neighbor conflicts.
                Returns list of textposition strings in the same order as input."""
                coords = [(a['thi'], a['bvi']) for a in apps_list]
                n = len(coords)
                assigned = [None] * n

                # Process most crowded points first
                neighbor_counts = []
                for i, (x1, y1) in enumerate(coords):
                    cnt = sum(1 for j, (x2, y2) in enumerate(coords)
                              if i != j and math.hypot(x2 - x1, y2 - y1) < 15)
                    neighbor_counts.append(cnt)
                order = sorted(range(n), key=lambda i: -neighbor_counts[i])

                # Track which positions are used by nearby points
                for idx in order:
                    x, y = coords[idx]
                    best_pos = 'top right'
                    best_conflicts = float('inf')

                    for pos_name, (dx, dy) in TEXT_POSITIONS:
                        conflicts = 0
                        # Check how many nearby points use the same position
                        for j in range(n):
                            if j == idx or assigned[j] is None:
                                continue
                            dist = math.hypot(coords[j][0] - x, coords[j][1] - y)
                            if dist < 15 and assigned[j] == pos_name:
                                conflicts += 3 if dist < 8 else 1
                        # Prefer positions that push text away from chart center/edges
                        if x > 70 and dx > 0: conflicts += 0.5   # near right edge, don't go right
                        if x < 30 and dx < 0: conflicts += 0.5   # near left edge, don't go left
                        if y > 80 and dy > 0: conflicts += 0.5   # near top, don't go up
                        if y < 20 and dy < 0: conflicts += 0.5   # near bottom, don't go down

                        if conflicts < best_conflicts:
                            best_conflicts = conflicts
                            best_pos = pos_name
                            if conflicts == 0:
                                break

                    assigned[idx] = best_pos
                return assigned

            # Compute best text positions
            text_positions = assign_text_positions(apps_with_scores)

            # Build a lookup: app_name -> textposition
            pos_lookup = {app['name']: text_positions[i] for i, app in enumerate(apps_with_scores)}

            for rec in ['EVOLVE', 'INVEST', 'MAINTAIN', 'ELIMINATE']:
                df_rec = df[df['recommendation'] == rec].copy()
                if not df_rec.empty:
                    for idx, (_, row) in enumerate(df_rec.iterrows()):
                        text_pos = pos_lookup.get(row['name'], 'top right')
                        display_name = row['name'] if len(row['name']) <= 22 else row['name'][:19] + '...'

                        fig.add_trace(go.Scatter(
                            x=[row['thi']],
                            y=[row['bvi']],
                            mode='markers+text',
                            name=rec,
                            text=[display_name],
                            textposition=text_pos,
                            textfont=dict(
                                size=8,
                                color='#444444',
                                family='Arial, sans-serif'
                            ),
                            marker=dict(
                                size=10,
                                color=colors_map[rec],
                                line=dict(width=1.5, color='white'),
                                opacity=0.85
                            ),
                            showlegend=(idx == 0),
                            legendgroup=rec,
                            hovertemplate=f'<b>{row["name"]}</b><br>THI: {row["thi"]:.1f}<br>BVI: {row["bvi"]:.1f}<extra></extra>',
                            hoverlabel=dict(
                                bgcolor='white',
                                bordercolor=colors_map[rec],
                                font=dict(size=12, color='black')
                            )
                        ))

            # Add quadrant labels (X=THI, Y=BVI) - Increased font size
            fig.add_annotation(x=80, y=80, text="EVOLVE", showarrow=False, font=dict(size=20, color="green", family="Arial Black"))
            fig.add_annotation(x=30, y=80, text="INVEST", showarrow=False, font=dict(size=20, color="orange", family="Arial Black"))
            fig.add_annotation(x=80, y=30, text="MAINTAIN", showarrow=False, font=dict(size=20, color="blue", family="Arial Black"))
            fig.add_annotation(x=30, y=30, text="ELIMINATE", showarrow=False, font=dict(size=20, color="red", family="Arial Black"))

            fig.update_layout(
                xaxis_title="Technical Health Index (THI)",
                yaxis_title="Business Value Index (BVI)",
                xaxis=dict(
                    range=[0, 100],  # Appropriate zoom level for actual data range
                    gridcolor='lightgray',
                    title_font=dict(size=16),
                    showgrid=True,
                    zeroline=False
                ),
                yaxis=dict(
                    range=[0, 100],  # Appropriate zoom level for actual data range
                    gridcolor='lightgray',
                    title_font=dict(size=16),
                    showgrid=True,
                    zeroline=False
                ),
                height=800,  # Standard height for better visualization
                showlegend=True,
                legend=dict(font=dict(size=14), x=1.02, y=1),
                plot_bgcolor='white',
                hoverlabel=dict(font_size=12),
                margin=dict(l=80, r=100, t=60, b=80),
                font=dict(size=12)
            )

            st.plotly_chart(fig, width="stretch", key="strategic_matrix_thi_bvi_v2")

        with tab2:
            # Enhanced Insights Tab with Portfolio-Level Analysis
            st.markdown("#### 💡 Strategic Portfolio Insights")
            st.markdown("AI-powered analysis combining transcripts, questionnaires, and market research")

            # Check if new insights exist
            from database import PortfolioInsight, AppInsight
            portfolio_insights = session.query(PortfolioInsight).all()
            app_insights_count = session.query(AppInsight).count()

            if not portfolio_insights and app_insights_count == 0:
                st.info("""
                **No insights generated yet.**

                Run the enhanced insight generator to get:
                - Deep per-application analysis (capabilities, user satisfaction, technical debt)
                - Portfolio-level patterns (consolidation opportunities, redundancies, gaps)
                - Market research integration for commercial products
                - Actionable recommendations with evidence

                **To generate insights:**
                ```bash
                cd webapp
                python generate_insights.py
                ```

                Estimated cost: ~$5-10 (one-time) | Uses GPT-4o for highest quality analysis
                """)
            else:
                # Statistics
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Portfolio Insights", len(portfolio_insights))
                with col2:
                    consolidation_count = len([i for i in portfolio_insights if i.insight_type == 'consolidation'])
                    st.metric("Consolidation Opps", consolidation_count)
                with col3:
                    quick_wins = len([i for i in portfolio_insights if i.insight_type == 'quick_win'])
                    st.metric("Quick Wins", quick_wins)
                with col4:
                    risks = len([i for i in portfolio_insights if i.insight_type == 'risk'])
                    st.metric("Risk Areas", risks)

                st.markdown("---")

                # Display Portfolio Insights
                if portfolio_insights:
                    st.markdown("### 🌐 Portfolio-Level Strategic Insights")

                    # Group by type
                    insights_by_type = {}
                    for insight in portfolio_insights:
                        t = insight.insight_type
                        if t not in insights_by_type:
                            insights_by_type[t] = []
                        insights_by_type[t].append(insight)

                    # Type icons and colors
                    type_config = {
                        'consolidation': {'icon': '🔄', 'color': '#E87722', 'label': 'Consolidation Opportunities'},
                        'quick_win': {'icon': '⚡', 'color': '#10B981', 'label': 'Quick Wins'},
                        'risk': {'icon': '⚠️', 'color': '#EF4444', 'label': 'Risk Areas'},
                        'integration': {'icon': '🔗', 'color': '#3B82F6', 'label': 'Integration Points'},
                        'gap': {'icon': '🎯', 'color': '#8B5CF6', 'label': 'Portfolio Gaps'},
                        'priority': {'icon': '📌', 'color': '#F59E0B', 'label': 'Strategic Priorities'}
                    }

                    # Display each type
                    for insight_type, insights_list in sorted(insights_by_type.items()):
                        config = type_config.get(insight_type, {'icon': '💡', 'color': '#6B7280', 'label': insight_type.replace('_', ' ').title()})

                        st.markdown(f"### {config['icon']} {config['label']} ({len(insights_list)})")

                        for insight in insights_list:
                            # Priority color
                            priority_colors = {
                                'P1': '🔴',
                                'P2': '🟡',
                                'P3': '🟢'
                            }
                            priority_icon = priority_colors.get(insight.priority, '⚪')

                            with st.expander(f"{priority_icon} {insight.title}"):
                                st.markdown(insight.description)

                                # Details
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.metric("Priority", insight.priority)
                                with col2:
                                    if insight.estimated_impact:
                                        impact_emoji = {'high': '🔴', 'medium': '🟡', 'low': '🟢'}.get(insight.estimated_impact, '⚪')
                                        st.metric("Impact", f"{impact_emoji} {insight.estimated_impact.upper()}")
                                with col3:
                                    if insight.complexity:
                                        complexity_emoji = {'high': '🔴', 'medium': '🟡', 'low': '🟢'}.get(insight.complexity, '⚪')
                                        st.metric("Complexity", f"{complexity_emoji} {insight.complexity.upper()}")

                                # Affected apps
                                if insight.affected_apps:
                                    st.markdown("**📱 Affected Applications:**")
                                    # Show as pills
                                    apps_html = " ".join([
                                        f'<span style="background: #F3F4F6; padding: 0.25rem 0.75rem; border-radius: 12px; margin: 0.25rem; display: inline-block;">{app}</span>'
                                        for app in insight.affected_apps
                                    ])
                                    st.markdown(apps_html, unsafe_allow_html=True)

                                # Recommended action
                                if insight.recommended_action:
                                    st.markdown("**🎯 Recommended Action:**")
                                    st.info(insight.recommended_action)

                                st.caption(f"Generated: {insight.generated_at.strftime('%Y-%m-%d %H:%M')} | Model: {insight.model_version}")

                        st.markdown("---")

                # Application-Level Insights Summary
                if app_insights_count > 0:
                    st.markdown("### 📊 Application-Level Insights")
                    st.info(f"""
                    **{app_insights_count} detailed insights** generated across all applications.

                    View per-application insights in the **Applications** page → Select app → Strategic Insights section

                    Each application has insights on:
                    - Capabilities & Limitations
                    - User Satisfaction
                    - Technical Debt
                    - Integration Opportunities
                    - Market Alternatives
                    - Strategic Recommendations
                    """)

    finally:
        close_session(session)


# ==================== PAGE: INSIGHTS (REMOVED - Merged into Analyses tab 4) ====================
def page_insights():
    # DEPRECATED - This page has been merged into Analyses as tab 4
    st.warning("⚠️ This page has been deprecated. Please use Analyses → Insights tab instead.")
    return

    st.title("💡 Strategic Insights")
    st.markdown("### AI-Generated Portfolio Insights")

    session = get_session()

    try:
        # Check if insights exist
        existing_insights = session.query(Insight).all()

        if not existing_insights:
            st.info("No insights generated yet. Click the button below to generate insights.")

        # Generate insights button
        if st.button("🤖 Generate Insights with AI", type="primary"):
            with st.spinner("Analyzing portfolio and generating insights..."):
                # Get all applications with scores
                apps = get_all_applications_from_db(session)
                apps_data = []

                for app in apps:
                    scores_data = session.query(SynergyScore).filter_by(
                        application_id=app.id,
                        approved=True
                    ).all()

                    if scores_data:
                        scores = {s.block_name: s.score for s in scores_data}
                        bvi, thi = calculate_bvi_thi(scores, {b: {'Weight': w} for b, w in get_current_weights().items()})
                        rec = get_recommendation(bvi, thi)

                        # Get key facts
                        qa_data = session.query(QuestionnaireAnswer).filter_by(
                            application_id=app.id
                        ).all()

                        business_critical = ""
                        technology = ""
                        cost = ""
                        integrations = []

                        for qa in qa_data:
                            if "business-critical" in qa.question_text.lower():
                                business_critical = qa.answer_text
                            elif "technolog" in qa.question_text.lower():
                                technology = qa.answer_text
                            elif "cost" in qa.question_text.lower():
                                cost = qa.answer_text
                            elif "integrate" in qa.question_text.lower():
                                integrations.append(qa.answer_text)

                        apps_data.append({
                            'name': app.name,
                            'bvi': bvi,
                            'thi': thi,
                            'recommendation': rec,
                            'business_critical': business_critical,
                            'technology': technology,
                            'cost': cost,
                            'integrations': integrations
                        })

                # Generate insights
                insights = generate_insights(apps_data)

                # Save to database
                for insight_data in insights:
                    # Find affected apps
                    affected_apps = insight_data.get('affected_apps', [])
                    app_id = None
                    if affected_apps:
                        for app in apps:
                            if app.name in affected_apps:
                                app_id = app.id
                                break

                    insight = Insight(
                        id=str(uuid.uuid4()),
                        application_id=app_id,
                        insight_type=insight_data.get('type', 'general'),
                        title=insight_data.get('title', ''),
                        description=insight_data.get('description', ''),
                        priority=insight_data.get('priority', 'P3'),
                        recommendation=insight_data.get('recommendation'),
                        supporting_data={'affected_apps': affected_apps}
                    )
                    session.add(insight)

                session.commit()
                st.success(f"✅ Generated {len(insights)} insights!")
                st.balloons()
                st.rerun()

        # Display insights
        insights = session.query(Insight).all()

        if insights:
            # Group by priority
            p1_insights = [i for i in insights if i.priority == 'P1']
            p2_insights = [i for i in insights if i.priority == 'P2']
            p3_insights = [i for i in insights if i.priority == 'P3']

            # Display by priority
            for priority, insights_list, color in [
                ('P1 - High Priority', p1_insights, '#EF4444'),
                ('P2 - Medium Priority', p2_insights, '#F59E0B'),
                ('P3 - Low Priority', p3_insights, '#3B82F6')
            ]:
                if insights_list:
                    st.markdown(f"### {priority}")

                    for insight in insights_list:
                        # Icon based on type
                        type_icons = {
                            'integration': '🔗',
                            'absorption': '📥',
                            'technology_update': '🚀',
                            'risk': '⚠️',
                            'financial': '💰',
                            'quick_win': '⚡'
                        }
                        icon = type_icons.get(insight.insight_type, '💡')

                        # Card
                        with st.expander(f"{icon} {insight.title}", expanded=priority == 'P1 - High Priority'):
                            st.markdown(f"**Type:** {insight.insight_type.replace('_', ' ').title()}")

                            if insight.recommendation:
                                st.markdown(f"**Recommendation:** :{color}[{insight.recommendation}]")

                            st.markdown(f"**Description:**")
                            st.write(insight.description)

                            if insight.supporting_data and insight.supporting_data.get('affected_apps'):
                                st.markdown("**Affected Applications:**")
                                for app_name in insight.supporting_data['affected_apps']:
                                    st.markdown(f"- {app_name}")

                            st.caption(f"Generated: {insight.created_at.strftime('%Y-%m-%d %H:%M')}")

    finally:
        close_session(session)


# ==================== PAGE: Q&A ASSISTANT ====================
def page_qa_assistant():
    st.title("🤖 AI Assistant")
    st.markdown("### Ask questions about your application portfolio")

    st.info("💡 Ask me anything about the applications, their scores, integrations, costs, risks, and more!")

    # Suggested questions
    with st.expander("💭 Suggested Questions"):
        st.markdown("""
        - Which applications are candidates for consolidation?
        - What is the total cost of legacy applications?
        - Which applications have security concerns?
        - How does application X integrate with application Y?
        - What would be the impact of eliminating application Z?
        - Which applications should we prioritize for modernization?
        - What are the main technical risks in our portfolio?
        - Which applications are business-critical but have low technical health?
        """)

    session = get_session()

    try:
        # Chat input
        user_question = st.text_area("Your question:", height=100, placeholder="e.g., Which applications should we prioritize for cloud migration?")

        if st.button("🔍 Ask", type="primary") and user_question:
            with st.spinner("Thinking..."):
                # Gather context
                apps = get_all_applications_from_db(session)
                context_data = {'applications': []}

                for app in apps:
                    # Get scores
                    scores_data = session.query(SynergyScore).filter_by(
                        application_id=app.id,
                        approved=True
                    ).all()

                    if scores_data:
                        scores = {s.block_name: s.score for s in scores_data}
                        bvi, thi = calculate_bvi_thi(scores, {b: {'Weight': w} for b, w in get_current_weights().items()})
                        rec = get_recommendation(bvi, thi)

                        # Build RICH CONTEXT - Prioritize David's notes
                        from database import DavidNote

                        # 1. Get David's insights (highest priority)
                        david_insights = session.query(DavidNote).filter_by(
                            application_id=app.id,
                            note_type='insight'
                        ).first()
                        david_insight_text = david_insights.answer_text if david_insights else ""

                        # 2. Get David's detailed notes (highest priority for Q&A)
                        david_notes = session.query(DavidNote).filter_by(
                            application_id=app.id,
                            note_type='answer'
                        ).all()
                        david_answers = {dn.question_text: dn.answer_text for dn in david_notes}  # FULL TEXT, no truncation

                        # 3. Get questionnaire answers (complete)
                        qa_data = session.query(QuestionnaireAnswer).filter_by(
                            application_id=app.id
                        ).all()
                        qa_answers = {qa.question_text: qa.answer_text for qa in qa_data}  # FULL TEXT

                        # 4. Get transcript answers (complete)
                        ta_data = session.query(TranscriptAnswer).filter_by(
                            application_id=app.id
                        ).all()
                        transcript_answers = {ta.question_text: f"{ta.answer_text} (confidence: {ta.confidence_score:.0%})" for ta in ta_data}

                        # 5. Merge all answers - David's notes take priority
                        all_answers = {
                            **qa_answers,           # Base: questionnaire
                            **transcript_answers,   # Override: transcripts
                            **david_answers         # Highest priority: David's notes
                        }

                        # 6. Get synergy block scores with rationales
                        scores_with_rationale = {s.block_name: {'score': s.score, 'rationale': s.rationale} for s in scores_data}

                        context_data['applications'].append({
                            'name': app.name,
                            'bvi': bvi,
                            'thi': thi,
                            'recommendation': rec,
                            'scores': scores_with_rationale,  # Include rationales
                            'all_answers': all_answers,  # Complete answers from all sources
                            'david_key_insights': david_insight_text,  # Executive summary from David
                            'data_sources': {
                                'questionnaire_count': len(qa_answers),
                                'transcript_count': len(transcript_answers),
                                'david_notes_count': len(david_answers),
                                'has_david_insights': bool(david_insight_text)
                            }
                        })

                # Call AI
                answer, sources, response_time = answer_question(user_question, context_data)

                # Save to history
                qa_history = QAHistory(
                    id=str(uuid.uuid4()),
                    user_question=user_question,
                    ai_response=answer,
                    context_applications={'count': len(apps)},
                    sources=sources,
                    response_time_ms=response_time
                )
                session.add(qa_history)
                session.commit()

                # Display answer
                st.markdown("### 💬 Answer:")
                st.markdown(answer)

                if sources:
                    st.markdown("**Sources:**")
                    for source in sources:
                        st.markdown(f"- {source}")

                st.caption(f"Response time: {response_time}ms")

                # Feedback
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("👍 Helpful"):
                        qa_history.user_feedback = 'helpful'
                        session.commit()
                        st.success("Thanks for your feedback!")
                with col2:
                    if st.button("👎 Not Helpful"):
                        qa_history.user_feedback = 'not_helpful'
                        session.commit()
                        st.info("Thanks for your feedback!")

        # Chat history
        st.markdown("---")
        st.markdown("### 📜 Recent Conversations")

        history = session.query(QAHistory).order_by(QAHistory.created_at.desc()).limit(5).all()

        if history:
            for qa in history:
                with st.expander(f"Q: {qa.user_question[:80]}..."):
                    st.markdown(f"**Question:** {qa.user_question}")
                    st.markdown(f"**Answer:** {qa.ai_response}")
                    st.caption(f"Asked: {qa.created_at.strftime('%Y-%m-%d %H:%M')}")
        else:
            st.info("No previous conversations")

    finally:
        close_session(session)


# ==================== PAGE: METHODOLOGY ====================
def page_calculator():
    """Calculator page - Overview of all applications with scores"""

    # Hero header
    st.markdown("""
    <div style="background: linear-gradient(135deg, #F3F4F6 0%, #FFFFFF 100%);
                padding: 2rem; border-radius: 16px; margin-bottom: 2rem; border: 1px solid #E5E7EB;">
        <h1 style="color: #111827; margin: 0; font-size: 2rem; font-weight: 800;">📋 Portfolio Calculator</h1>
        <p style="color: #6B7280; margin: 0.5rem 0 0 0;">Complete overview of all applications with scores, indices, and strategic recommendations</p>
    </div>
    """, unsafe_allow_html=True)

    session = get_session()

    try:
        apps = get_all_applications_from_db(session)

        if not apps:
            st.info("No applications found.")
            return

        # Weight Editor
        with st.expander("⚙️ Configure Block Weights", expanded=False):
            st.markdown("""
            Adjust the weights of each synergy block. Weights determine the relative importance of each block
            in calculating BVI and THI indices. Changes apply immediately to all calculations.
            """)

            # Load weights from database
            current_weights = get_current_weights()

            col1, col2 = st.columns(2)

            weights_changed = False

            with col1:
                st.markdown("**📈 Business Value Blocks**")
                for block in ["Strategic Fit", "Business Efficiency", "User Value", "Financial Value"]:
                    new_val = st.number_input(
                        f"{block}",
                        min_value=0,
                        max_value=100,
                        value=current_weights.get(block, 25),
                        step=5,
                        key=f"weight_{block}"
                    )
                    if new_val != current_weights.get(block, 25):
                        weights_changed = True
                    st.session_state.custom_weights[block] = new_val

            with col2:
                st.markdown("**🔧 Technical Health Blocks**")
                for block in ["Architecture", "Operational Risk", "Maintainability", "Support Quality"]:
                    new_val = st.number_input(
                        f"{block}",
                        min_value=0,
                        max_value=100,
                        value=current_weights.get(block, 25),
                        step=5,
                        key=f"weight_{block}"
                    )
                    if new_val != current_weights.get(block, 25):
                        weights_changed = True
                    st.session_state.custom_weights[block] = new_val

            # Auto-save to database whenever weights change
            save_weights_to_db()

            if weights_changed:
                st.toast("Weights saved automatically!", icon="✅")

            # Show totals
            business_total = sum(st.session_state.custom_weights[b] for b in ["Strategic Fit", "Business Efficiency", "User Value", "Financial Value"])
            tech_total = sum(st.session_state.custom_weights[b] for b in ["Architecture", "Operational Risk", "Maintainability", "Support Quality"])

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Business Total", f"{business_total}%")
            with col2:
                st.metric("Tech Total", f"{tech_total}%")
            with col3:
                if st.button("🔄 Reset to Defaults"):
                    st.session_state.custom_weights = {
                        block: info['Weight'] for block, info in SYNERGY_BLOCKS.items()
                    }
                    save_weights_to_db()
                    st.rerun()

        st.markdown("---")

        # Build calculator table data
        calculator_data = []
        app_id_map = {}

        # Get custom weights in the right format
        weights_for_calc = {
            block: {'Weight': st.session_state.custom_weights[block]}
            for block in SYNERGY_BLOCKS.keys()
        }

        # Define subcategory options per decision
        SUBCATEGORY_OPTIONS = {
            'ELIMINATE': ['Replace', 'Retire', 'Absorbed'],
            'INVEST': ['Absorb', 'Modernize'],
            'EVOLVE': ['Migrate', 'Enhance', 'Refactor', 'Upgrade'],
            'MAINTAIN': ['Internalize', 'Maintain']
        }

        # Priority mapping based on subcategory
        PRIORITY_MAP = {
            'Replace': 'P1 - Critical',
            'Retire': 'P1 - Critical',
            'Absorb': 'P1 - Critical',
            'Modernize': 'P1 - Critical',
            'Migrate': 'P1 - Critical',
            'Absorbed': 'P2 - Tactical',
            'Enhance': 'P2 - Strategic',
            'Refactor': 'P2 - Strategic',
            'Upgrade': 'P2 - Strategic',
            'Internalize': 'P2 - Compliance',
            'Maintain': 'P3 - Routine'
        }

        for app in apps:
            scores_data = session.query(SynergyScore).filter_by(
                application_id=app.id,
                approved=True
            ).all()

            if scores_data:
                scores = {s.block_name: s.score for s in scores_data}

                # Calculate BVI/THI with custom weights
                bvi, thi = calculate_bvi_thi(scores, weights_for_calc)
                rec = get_recommendation(bvi, thi)

                # Get subcategory from database (user-filled) - DO NOT auto-calculate
                subcategory = app.subcategory if app.subcategory else ''

                # Calculate priority ONLY if subcategory is filled
                if subcategory:
                    base_priority = PRIORITY_MAP.get(subcategory, 'P3 - Routine')
                    if app.quick_win and base_priority.startswith('P2'):
                        priority = 'P1 - Quick Win'
                    elif app.quick_win and base_priority.startswith('P3'):
                        priority = 'P2 - Quick Win'
                    else:
                        priority = base_priority
                else:
                    priority = ''  # Empty until user selects subcategory

                row = {
                    'app_id': app.id,
                    'Application': app.name,
                    'Strategic Fit': scores.get('Strategic Fit', ''),
                    'Business Efficiency': scores.get('Business Efficiency', ''),
                    'User Value': scores.get('User Value', ''),
                    'Financial Value': scores.get('Financial Value', ''),
                    'BVI': round(bvi, 1),
                    'Architecture': scores.get('Architecture', ''),
                    'Operational Risk': scores.get('Operational Risk', ''),
                    'Maintainability': scores.get('Maintainability', ''),
                    'Support Quality': scores.get('Support Quality', ''),
                    'THI': round(thi, 1),
                    'Decision': rec,
                    'Subcategory': subcategory,
                    'Quick Win': app.quick_win if app.quick_win else False,
                    'Priority': priority
                }
                calculator_data.append(row)
                app_id_map[app.name] = app.id

        if calculator_data:
            calc_df = pd.DataFrame(calculator_data)

            # Reset subcategories button
            with st.expander("🔄 Reset Subcategories"):
                st.warning("This will clear all Subcategory and Priority values for every application.")
                if st.button("Clear All Subcategories", type="primary", key="reset_subcats"):
                    for app in apps:
                        app.subcategory = None
                        app.quick_win = False
                    session.commit()
                    st.success("All subcategories have been cleared!")
                    st.rerun()

            # Reference: Subcategory options per Decision
            st.markdown('<div style="background:#F9FAFB;padding:1rem 1.25rem;border-radius:8px;border:1px solid #E5E7EB;margin-bottom:1rem;"><p style="color:#374151;font-weight:700;font-size:0.85rem;margin:0 0 0.5rem 0;">Subcategory Options by Decision:</p><div style="display:flex;gap:1.5rem;flex-wrap:wrap;"><div><span style="color:#10B981;font-weight:600;font-size:0.8rem;">EVOLVE:</span><span style="color:#6B7280;font-size:0.8rem;"> Migrate, Enhance, Refactor, Upgrade</span></div><div><span style="color:#F59E0B;font-weight:600;font-size:0.8rem;">INVEST:</span><span style="color:#6B7280;font-size:0.8rem;"> Absorb, Modernize</span></div><div><span style="color:#3B82F6;font-weight:600;font-size:0.8rem;">MAINTAIN:</span><span style="color:#6B7280;font-size:0.8rem;"> Internalize, Maintain</span></div><div><span style="color:#EF4444;font-weight:600;font-size:0.8rem;">ELIMINATE:</span><span style="color:#6B7280;font-size:0.8rem;"> Replace, Retire, Absorbed</span></div></div></div>', unsafe_allow_html=True)

            st.markdown("""
            **💡 How to use:** Select a **Subcategory** matching the application's Decision (see reference above). **Quick Win** elevates P2/P3 priorities. **Priority** is calculated automatically. *Changes are saved automatically.*
            """)

            # All subcategory options in a single list (with empty for clearing)
            all_sub_options = ['', 'Absorb', 'Absorbed', 'Enhance', 'Internalize', 'Maintain', 'Migrate', 'Modernize', 'Refactor', 'Replace', 'Retire', 'Upgrade']

            column_config = {
                'app_id': None,
                'Application': st.column_config.TextColumn('Application', disabled=True, width='medium'),
                'Strategic Fit': st.column_config.NumberColumn('Strategic Fit', disabled=True, width='small'),
                'Business Efficiency': st.column_config.NumberColumn('Business Efficiency', disabled=True, width='small'),
                'User Value': st.column_config.NumberColumn('User Value', disabled=True, width='small'),
                'Financial Value': st.column_config.NumberColumn('Financial Value', disabled=True, width='small'),
                'BVI': st.column_config.NumberColumn('BVI', disabled=True, format='%.1f', width='small'),
                'Architecture': st.column_config.NumberColumn('Architecture', disabled=True, width='small'),
                'Operational Risk': st.column_config.NumberColumn('Operational Risk', disabled=True, width='small'),
                'Maintainability': st.column_config.NumberColumn('Maintainability', disabled=True, width='small'),
                'Support Quality': st.column_config.NumberColumn('Support Quality', disabled=True, width='small'),
                'THI': st.column_config.NumberColumn('THI', disabled=True, format='%.1f', width='small'),
                'Decision': st.column_config.TextColumn('Decision', disabled=True, width='small'),
                'Subcategory': st.column_config.SelectboxColumn(
                    'Subcategory',
                    options=all_sub_options,
                    required=False,
                    width='medium'
                ),
                'Quick Win': st.column_config.CheckboxColumn('Quick Win', width='small'),
                'Priority': st.column_config.TextColumn('Priority', disabled=True, width='medium')
            }

            edited_df = st.data_editor(
                calc_df,
                column_config=column_config,
                width="stretch",
                height=600,
                hide_index=True,
                key='calculator_editor'
            )

            # Detect changes and save to database
            if not edited_df.equals(calc_df):
                for idx, row in edited_df.iterrows():
                    original_row = calc_df.loc[idx]
                    app_id = row['app_id']

                    if row['Subcategory'] != original_row['Subcategory'] or row['Quick Win'] != original_row['Quick Win']:
                        app = session.query(Application).filter_by(id=app_id).first()
                        if app:
                            subcategory_val = row['Subcategory'] if row['Subcategory'] else None
                            decision = row['Decision']

                            # Validate subcategory matches decision
                            if subcategory_val and decision in SUBCATEGORY_OPTIONS:
                                valid_options = SUBCATEGORY_OPTIONS[decision]
                                if subcategory_val not in valid_options:
                                    st.error(f"**{row['Application']}**: '{subcategory_val}' is not valid for {decision}. Valid options: {', '.join(valid_options)}")
                                    continue

                            app.subcategory = subcategory_val
                            app.quick_win = bool(row['Quick Win'])

                            if subcategory_val:
                                base_priority = PRIORITY_MAP.get(subcategory_val, 'P3 - Routine')
                                if row['Quick Win'] and base_priority.startswith('P2'):
                                    new_priority = 'P1 - Quick Win'
                                elif row['Quick Win'] and base_priority.startswith('P3'):
                                    new_priority = 'P2 - Quick Win'
                                else:
                                    new_priority = base_priority
                            else:
                                new_priority = ''

                            edited_df.at[idx, 'Priority'] = new_priority

                session.commit()
                st.success("Changes saved successfully!")
                st.rerun()

            # Export buttons
            st.markdown("---")
            st.markdown("### Export Portfolio")
            col_xl, col_ppt = st.columns(2)

            with col_xl:
                if st.button("📥 Generate Portfolio Excel", width="stretch", type="primary", key="gen_excel"):
                    with st.spinner("Generating Excel with all sheets (Calculator, Dashboard, Roadmap, App Groups, Value Chain, App Sheets)..."):
                        from excel_generator import generate_portfolio_excel
                        xlsx_bytes = generate_portfolio_excel(custom_weights=st.session_state.get('custom_weights'))
                        if xlsx_bytes:
                            st.session_state['xlsx_data'] = xlsx_bytes
                            st.session_state['xlsx_ready'] = True
                            st.rerun()
                        else:
                            st.error("No application data available.")

                if st.session_state.get('xlsx_ready'):
                    st.download_button(
                        label="⬇️ Download Portfolio Excel",
                        data=st.session_state['xlsx_data'],
                        file_name=f"Avangrid_Application_Portfolio_Management_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch"
                    )

            with col_ppt:
                if st.button("📊 Generate Portfolio PowerPoint", width="stretch", type="primary", key="gen_ppt"):
                    with st.spinner("Generating PowerPoint with all application cards..."):
                        try:
                            from ppt_generator import generate_portfolio_pptx
                            pptx_bytes = generate_portfolio_pptx()
                            if pptx_bytes:
                                st.session_state['pptx_data'] = pptx_bytes
                                st.session_state['pptx_ready'] = True
                                st.rerun()
                            else:
                                st.error("No application data available.")
                        except Exception as e:
                            st.error(f"Error generating PowerPoint: {str(e)}")

                if st.session_state.get('pptx_ready'):
                    st.download_button(
                        label="⬇️ Download Portfolio PPTX",
                        data=st.session_state['pptx_data'],
                        file_name=f"Avangrid_Portfolio_{datetime.now().strftime('%Y%m%d')}.pptx",
                        mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                        width="stretch"
                    )
        else:
            st.warning("No applications with approved scores found.")

    finally:
        close_session(session)



def page_methodology():
    """Assessment Methodology - Detailed Framework Documentation"""

    # Hero header - Clean, professional
    st.markdown("""
    <div style="background: #FFFFFF; padding: 3rem 2.5rem; border-radius: 16px;
                border: 1px solid #E5E7EB; margin-bottom: 2.5rem;">
        <div style="border-left: 4px solid #E87722; padding-left: 1.5rem;">
            <h1 style="color: #111827; margin: 0 0 0.75rem 0; font-size: 2.25rem; font-weight: 700; line-height: 1.2;">
                Assessment Methodology
            </h1>
            <p style="color: #6B7280; font-size: 1.0625rem; margin: 0; line-height: 1.6;">
                Systematic framework for quantifying business value and technical health across enterprise applications
            </p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Assessment Process
    st.markdown("### Assessment Process")

    process_steps = [
        {
            "number": "01",
            "title": "Data Collection",
            "description": "Structured questionnaires capture application attributes across 8 dimensions. Stakeholder interviews and meeting transcripts provide qualitative context. Expert review sessions document strategic insights and operational knowledge."
        },
        {
            "number": "02",
            "title": "AI-Powered Analysis",
            "description": "GPT-4o processes questionnaire responses and transcript data to extract answers to master questions. Natural language understanding enables automatic scoring with confidence levels. Human experts validate and refine AI-generated assessments."
        },
        {
            "number": "03",
            "title": "Quantitative Scoring",
            "description": "Each synergy block receives a score from 1-5 based on weighted criteria. BVI aggregates 4 business-focused blocks; THI combines 4 technical blocks. Configurable weights allow customization for organizational priorities."
        },
        {
            "number": "04",
            "title": "Strategic Classification",
            "description": "Applications position on 2x2 matrix (BVI vs THI) with 60-point threshold. Four quadrants yield strategic recommendations: EVOLVE, INVEST, MAINTAIN, or ELIMINATE. Portfolio-level analysis identifies consolidation opportunities and quick wins."
        }
    ]

    for idx, step in enumerate(process_steps):
        col1, col2 = st.columns([1, 11])
        with col1:
            st.markdown(f'<div style="background:linear-gradient(135deg,rgba(232,119,34,0.1),rgba(255,140,66,0.05));width:48px;height:48px;border-radius:10px;display:flex;align-items:center;justify-content:center;border:1px solid rgba(232,119,34,0.2);"><span style="color:#E87722;font-weight:700;font-size:1rem;">{step["number"]}</span></div>', unsafe_allow_html=True)

        with col2:
            st.markdown(f'<div style="background:#FFFFFF;padding:1.5rem;border-radius:12px;border:1px solid #E5E7EB;margin-bottom:1.25rem;"><h4 style="color:#111827;margin:0 0 0.5rem 0;font-weight:600;">{step["title"]}</h4><p style="color:#6B7280;font-size:0.9375rem;line-height:1.6;margin:0;">{step["description"]}</p></div>', unsafe_allow_html=True)

    st.markdown("<div style='margin: 2.5rem 0;'></div>", unsafe_allow_html=True)

    # Scoring Framework Detail
    st.markdown("### Detailed Scoring Framework")

    current_weights = get_current_weights()

    # BVI Blocks
    bvi_blocks_method = {
        "Strategic Fit": "Alignment with organizational strategy, support for business priorities, contribution to long-term goals, strategic criticality.",
        "Business Efficiency": "Process automation impact, operational cost reduction, productivity improvements, business agility enablement.",
        "User Value": "User satisfaction scores, adoption rates, productivity impact, user experience quality, training requirements.",
        "Financial Value": "Return on investment, revenue generation, cost avoidance, total cost of ownership, financial risk mitigation."
    }

    bvi_cards = ""
    for block, description in bvi_blocks_method.items():
        weight = current_weights.get(block, SYNERGY_BLOCKS[block]['Weight'])
        bvi_cards += f'<div style="background:#FFFFFF;padding:1.25rem;border-radius:10px;border:1px solid #E5E7EB;margin-bottom:0.75rem;"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.5rem;"><h5 style="color:#111827;margin:0;font-weight:600;">{block}</h5><span style="background:#10B981;color:#FFFFFF;padding:0.25rem 0.75rem;border-radius:20px;font-size:0.75rem;font-weight:600;">{weight}%</span></div><p style="color:#6B7280;font-size:0.875rem;line-height:1.5;margin:0;">{description}</p></div>'

    st.markdown(f'<div style="background:rgba(16,185,129,0.05);padding:1.75rem;border-radius:12px;border:1px solid rgba(16,185,129,0.15);margin-bottom:2rem;"><h4 style="color:#065F46;margin:0 0 1rem 0;font-weight:600;font-size:1.125rem;">Business Value Index (BVI) Components</h4>{bvi_cards}</div>', unsafe_allow_html=True)

    # THI Blocks
    thi_blocks_method = {
        "Architecture": "Scalability, flexibility, integration capabilities, technology stack modernity, cloud readiness, microservices adoption.",
        "Operational Risk": "System reliability, security vulnerabilities, compliance gaps, disaster recovery readiness, business continuity risk.",
        "Maintainability": "Code quality, documentation completeness, ease of updates, extensibility, technical debt levels, development velocity.",
        "Support Quality": "Vendor support availability, internal expertise depth, knowledge transfer capability, community resources, upgrade path clarity."
    }

    thi_cards = ""
    for block, description in thi_blocks_method.items():
        weight = current_weights.get(block, SYNERGY_BLOCKS[block]['Weight'])
        thi_cards += f'<div style="background:#FFFFFF;padding:1.25rem;border-radius:10px;border:1px solid #E5E7EB;margin-bottom:0.75rem;"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.5rem;"><h5 style="color:#111827;margin:0;font-weight:600;">{block}</h5><span style="background:#3B82F6;color:#FFFFFF;padding:0.25rem 0.75rem;border-radius:20px;font-size:0.75rem;font-weight:600;">{weight}%</span></div><p style="color:#6B7280;font-size:0.875rem;line-height:1.5;margin:0;">{description}</p></div>'

    st.markdown(f'<div style="background:rgba(59,130,246,0.05);padding:1.75rem;border-radius:12px;border:1px solid rgba(59,130,246,0.15);margin-bottom:2rem;"><h4 style="color:#1E3A8A;margin:0 0 1rem 0;font-weight:600;font-size:1.125rem;">Technical Health Index (THI) Components</h4>{thi_cards}</div>', unsafe_allow_html=True)

    st.markdown("<div style='margin: 2.5rem 0;'></div>", unsafe_allow_html=True)

    # Calculation Methodology
    st.markdown("### Calculation Methodology")

    # Get current weights for formulas
    curr_w = get_current_weights()

    # Build dynamic formulas
    bvi_formula = f"(Strategic Fit × {curr_w.get('Strategic Fit', 30)}%) + (Business Efficiency × {curr_w.get('Business Efficiency', 30)}%) + (User Value × {curr_w.get('User Value', 20)}%) + (Financial Value × {curr_w.get('Financial Value', 20)}%)"
    thi_formula = f"(Architecture × {curr_w.get('Architecture', 30)}%) + (Operational Risk × {curr_w.get('Operational Risk', 30)}%) + (Maintainability × {curr_w.get('Maintainability', 25)}%) + (Support Quality × {curr_w.get('Support Quality', 15)}%)"

    st.markdown(f"""
    <div style="background: #FFFFFF; padding: 2rem; border-radius: 12px; border: 1px solid #E5E7EB;">
        <div style="margin-bottom: 2rem;">
            <h4 style="color: #111827; margin: 0 0 1rem 0; font-weight: 600;">Index Calculation Formula</h4>
            <div style="background: #F9FAFB; padding: 1.5rem; border-radius: 8px; border-left: 4px solid #E87722;
                        font-family: monospace; margin-bottom: 1rem;">
                <div style="color: #374151; font-size: 0.9375rem; margin-bottom: 0.75rem;">
                    <strong>BVI</strong> = {bvi_formula}
                </div>
                <div style="color: #374151; font-size: 0.9375rem;">
                    <strong>THI</strong> = {thi_formula}
                </div>
            </div>
            <p style="color: #6B7280; font-size: 0.875rem; line-height: 1.6; margin: 0;">
                Each synergy block score (1-5) is multiplied by its weight percentage. The weighted sum produces the index value (0-100).
                Weights can be customized in the Calculator page to reflect organizational priorities.
            </p>
        </div>

    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='margin-bottom: 2rem;'></div>", unsafe_allow_html=True)
    st.markdown("#### Score Interpretation Scale")

    # Use Streamlit columns instead of CSS grid
    score_cols = st.columns(5)
    scores = [
        {"num": "1", "label": "Poor", "color": "#EF4444"},
        {"num": "2", "label": "Below Average", "color": "#F59E0B"},
        {"num": "3", "label": "Average", "color": "#6B7280"},
        {"num": "4", "label": "Good", "color": "#3B82F6"},
        {"num": "5", "label": "Excellent", "color": "#10B981"}
    ]

    for col, score in zip(score_cols, scores):
        with col:
            col.markdown(f"""
            <div style="text-align: center; padding: 1rem; border: 1px solid #E5E7EB; border-radius: 8px;">
                <div style="font-size: 1.5rem; margin-bottom: 0.5rem;">{score['num']}</div>
                <div style="color: {score['color']}; font-weight: 600; font-size: 0.75rem;">{score['label']}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("""
    <div style="background: #FFFFFF; padding: 2rem; border-radius: 12px; border: 1px solid #E5E7EB; margin-top: 2rem;">
        <h4 style="color: #111827; margin: 0 0 1rem 0; font-weight: 600;">Decision Threshold</h4>
        <p style="color: #6B7280; font-size: 0.9375rem; line-height: 1.6; margin: 0;">
            The <strong>60-point threshold</strong> divides the matrix into quadrants. Applications scoring ≥60 demonstrate
            high performance in that dimension; scores &lt;60 indicate improvement needed. This binary classification
            simplifies strategic decision-making while maintaining granularity through underlying block scores.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='margin: 2.5rem 0;'></div>", unsafe_allow_html=True)

    # Strategic Recommendations
    st.markdown("### Strategic Recommendation Logic")

    recommendations = [
        {
            "name": "EVOLVE",
            "color": "#10B981",
            "bg": "rgba(16, 185, 129, 0.08)",
            "border": "rgba(16, 185, 129, 0.2)",
            "condition": "BVI ≥ 60 AND THI ≥ 60",
            "action": "Continue investment to enhance capabilities, modernize user experience, and expand use cases. These applications are portfolio strengths.",
            "subcategories": [
                {"name": "Enhance", "priority": "P2 – Strategic", "desc": "Add new features and capabilities to maximize business value"},
                {"name": "Refactor", "priority": "P2 – Strategic", "desc": "Improve internal architecture while preserving functionality"},
                {"name": "Upgrade", "priority": "P2 – Strategic", "desc": "Update technology stack to latest versions and standards"},
                {"name": "Migrate", "priority": "P1 – Critical", "desc": "Move to new platform (e.g., cloud migration) for strategic advantage"},
            ]
        },
        {
            "name": "INVEST",
            "color": "#F59E0B",
            "bg": "rgba(251, 191, 36, 0.08)",
            "border": "rgba(251, 191, 36, 0.2)",
            "condition": "BVI ≥ 60 AND THI < 60",
            "action": "Urgent technical modernization required. High business value at risk due to technical debt. Prioritize architecture refactoring and technology stack upgrades.",
            "subcategories": [
                {"name": "Modernize", "priority": "P1 – Critical", "desc": "Comprehensive technology refresh to address critical technical debt"},
                {"name": "Absorb", "priority": "P1 – Critical", "desc": "Consolidate functionality into a stronger, healthier platform"},
            ]
        },
        {
            "name": "MAINTAIN",
            "color": "#3B82F6",
            "bg": "rgba(59, 130, 246, 0.08)",
            "border": "rgba(59, 130, 246, 0.2)",
            "condition": "BVI < 60 AND THI ≥ 60",
            "action": "Minimal investment strategy. Applications are technically sound but deliver limited business value. Maintain stability with security patches only.",
            "subcategories": [
                {"name": "Maintain", "priority": "P3 – Routine", "desc": "Keep operational with security patches and minimal updates only"},
                {"name": "Internalize", "priority": "P2 – Compliance", "desc": "Bring under internal governance for regulatory or compliance reasons"},
            ]
        },
        {
            "name": "ELIMINATE",
            "color": "#EF4444",
            "bg": "rgba(239, 68, 68, 0.08)",
            "border": "rgba(239, 68, 68, 0.2)",
            "condition": "BVI < 60 AND THI < 60",
            "action": "Strong candidates for decommissioning or replacement. Low business value combined with technical debt creates high cost, low return.",
            "subcategories": [
                {"name": "Replace", "priority": "P1 – Critical", "desc": "Substitute with a modern solution that better serves business needs"},
                {"name": "Retire", "priority": "P1 – Critical", "desc": "Decommission entirely — functionality no longer needed"},
                {"name": "Absorbed", "priority": "P2 – Tactical", "desc": "Merge functionality into another existing application"},
            ]
        }
    ]

    for rec in recommendations:
        # Build subcategory pills HTML
        sub_pills = ""
        for sub in rec["subcategories"]:
            # Color-code the priority pill
            p_color = "#DC2626" if "P1" in sub["priority"] else ("#E87722" if "P2" in sub["priority"] else "#6B7280")
            sub_pills += f'<div style="display:flex;align-items:center;gap:0.5rem;margin-bottom:0.5rem;"><span style="background:{p_color};color:#fff;padding:0.2rem 0.6rem;border-radius:12px;font-size:0.7rem;font-weight:600;white-space:nowrap;">{sub["priority"]}</span><span style="color:#374151;font-weight:600;font-size:0.85rem;min-width:80px;">{sub["name"]}</span><span style="color:#6B7280;font-size:0.8125rem;">— {sub["desc"]}</span></div>'

        st.markdown(f'<div style="background:{rec["bg"]};padding:1.75rem;border-radius:12px;border:1px solid {rec["border"]};border-left:4px solid {rec["color"]};margin-bottom:1.25rem;"><div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:1rem;"><h4 style="color:{rec["color"]};margin:0;font-weight:700;font-size:1.125rem;">{rec["name"]}</h4><span style="background:rgba(0,0,0,0.06);color:#374151;padding:0.35rem 0.85rem;border-radius:20px;font-size:0.75rem;font-weight:600;">{rec["condition"]}</span></div><p style="color:#4B5563;font-size:0.9375rem;line-height:1.6;margin:0 0 1.25rem 0;">{rec["action"]}</p><div style="background:rgba(255,255,255,0.5);padding:1rem;border-radius:8px;"><p style="color:#374151;font-size:0.8125rem;font-weight:700;margin:0 0 0.75rem 0;text-transform:uppercase;letter-spacing:0.05em;">Subcategories & Priorities</p>{sub_pills}</div></div>', unsafe_allow_html=True)

    # Priority Classification System
    st.markdown("---")
    st.markdown("### Priority Classification System")
    st.markdown("Each subcategory maps to a priority level that determines execution urgency and resource allocation:")

    priority_levels = [
        {"code": "P1 – Critical", "color": "#DC2626", "bg": "rgba(220, 38, 38, 0.06)", "border": "rgba(220, 38, 38, 0.15)", "desc": "Immediate action required. These represent the highest-urgency items due to significant business risk, technical debt threatening critical operations, or strategic decommissioning needs.", "subcats": "Replace, Retire, Absorb, Modernize, Migrate", "criteria": "Business continuity risk, security vulnerabilities, regulatory non-compliance, or end-of-life technology requiring immediate intervention."},
        {"code": "P1 – Quick Win", "color": "#D97706", "bg": "rgba(217, 119, 6, 0.06)", "border": "rgba(217, 119, 6, 0.15)", "desc": "P2 items that have been flagged as Quick Wins are automatically elevated to P1. These are strategic or tactical improvements that can be achieved rapidly with minimal effort and high impact.", "subcats": "Any P2 subcategory flagged as Quick Win", "criteria": "Low implementation effort, high visibility impact, no major dependencies, can be delivered within a single sprint cycle."},
        {"code": "P2 – Strategic", "color": "#E87722", "bg": "rgba(232, 119, 34, 0.06)", "border": "rgba(232, 119, 34, 0.15)", "desc": "Planned improvements aligned with the IT roadmap. These enhance high-value applications through feature expansion, architecture modernization, or technology upgrades.", "subcats": "Enhance, Refactor, Upgrade", "criteria": "Applications with strong business value (BVI ≥ 60) and healthy technical posture (THI ≥ 60). Focus on expanding capabilities and maintaining competitive advantage."},
        {"code": "P2 – Tactical", "color": "#8B5CF6", "bg": "rgba(139, 92, 246, 0.06)", "border": "rgba(139, 92, 246, 0.15)", "desc": "Targeted consolidation actions. Applications whose functionality can be absorbed into a larger, more capable system. Reduces portfolio complexity and maintenance overhead.", "subcats": "Absorbed", "criteria": "Redundant or overlapping functionality with another application. The target system must be identified and capable of absorbing the workload."},
        {"code": "P2 – Compliance", "color": "#0891B2", "bg": "rgba(8, 145, 178, 0.06)", "border": "rgba(8, 145, 178, 0.15)", "desc": "Governance-driven actions. Applications that must be brought under internal control for regulatory, security, or compliance requirements, even if they currently function adequately.", "subcats": "Internalize", "criteria": "Regulatory mandates, data sovereignty requirements, audit findings, or vendor risk management policies requiring internal governance."},
        {"code": "P3 – Routine", "color": "#6B7280", "bg": "rgba(107, 114, 128, 0.06)", "border": "rgba(107, 114, 128, 0.15)", "desc": "Steady-state operations with minimal investment. Applications are technically sound but offer limited strategic value. Keep stable with security patches only — no new feature development.", "subcats": "Maintain", "criteria": "Technically healthy (THI ≥ 60) but low business value (BVI < 60). No active business case for enhancement or retirement."},
    ]

    for p in priority_levels:
        st.markdown(f'<div style="background:{p["bg"]};padding:1.5rem;border-radius:12px;border:1px solid {p["border"]};border-left:4px solid {p["color"]};margin-bottom:1rem;"><div style="margin-bottom:0.75rem;"><span style="background:{p["color"]};color:#fff;padding:0.35rem 1rem;border-radius:20px;font-size:0.85rem;font-weight:700;">{p["code"]}</span></div><p style="color:#374151;font-size:0.9375rem;line-height:1.6;margin:0 0 1rem 0;">{p["desc"]}</p><div style="display:flex;gap:2rem;flex-wrap:wrap;"><div><span style="color:#6B7280;font-size:0.75rem;font-weight:700;text-transform:uppercase;letter-spacing:0.05em;">Subcategories</span><p style="color:#374151;font-size:0.875rem;margin:0.25rem 0 0 0;">{p["subcats"]}</p></div><div style="flex:1;min-width:200px;"><span style="color:#6B7280;font-size:0.75rem;font-weight:700;text-transform:uppercase;letter-spacing:0.05em;">Criteria</span><p style="color:#4B5563;font-size:0.8125rem;line-height:1.5;margin:0.25rem 0 0 0;">{p["criteria"]}</p></div></div></div>', unsafe_allow_html=True)


# ==================== PAGE: BATCH OPERATIONS (REMOVED - Replaced by auto-calculation) ====================
def page_batch_operations():
    # DEPRECATED - Batch operations replaced by automatic calculation on upload
    st.warning("⚠️ This page has been deprecated. Scores are now calculated automatically after questionnaire/transcript upload.")
    return

    st.title("⚙️ Batch Operations")
    st.markdown("### Automated Processing & Score Calculation")

    session = get_session()

    try:
        # Two main sections
        tab1, tab2 = st.tabs(["🤖 Process Transcript Queue", "🎯 Batch Calculate Scores"])

        # TAB 1: Process Transcript Queue
        with tab1:
            st.markdown("#### Process Pending Transcripts")
            st.info("📋 Automatically process all transcripts that are pending AI analysis")

            # Get pending transcripts
            pending_transcripts = session.query(MeetingTranscript).filter_by(processed=False).all()

            if not pending_transcripts:
                st.success("✅ No pending transcripts! All transcripts have been processed.")
            else:
                st.warning(f"⏳ **{len(pending_transcripts)} transcripts** waiting to be processed")

                # Show pending transcripts
                with st.expander(f"View pending transcripts ({len(pending_transcripts)})"):
                    for transcript in pending_transcripts:
                        app = session.query(Application).get(transcript.application_id)
                        st.markdown(f"- 📄 **{transcript.file_name}** → {app.name if app else 'Unknown App'}")

                st.markdown("---")

                # Process button
                if st.button("🚀 Process All Pending Transcripts", type="primary", width="stretch"):
                    import time
                    start_time = time.time()

                    st.markdown("### 🤖 Processing...")

                    # Progress UI
                    progress_container = st.container()
                    with progress_container:
                        progress_bar = st.progress(0)
                        progress_text = st.empty()
                        status_spinner = st.empty()

                    log_container = st.expander("📋 Processing Log", expanded=True)

                    processed_count = 0
                    error_count = 0

                    for idx, transcript in enumerate(pending_transcripts):
                        app = session.query(Application).get(transcript.application_id)

                        # Update progress
                        progress = idx / len(pending_transcripts)
                        progress_bar.progress(progress)

                        # Time estimate
                        if idx > 0:
                            elapsed = time.time() - start_time
                            avg_time = elapsed / idx
                            remaining = avg_time * (len(pending_transcripts) - idx)
                            progress_text.markdown(f"**Progress:** {idx}/{len(pending_transcripts)} • **Est. remaining:** {int(remaining)}s")
                        else:
                            progress_text.markdown(f"**Progress:** {idx}/{len(pending_transcripts)} • Starting...")

                        # Show spinner
                        with status_spinner:
                            with st.spinner(f"🔄 Processing: {transcript.file_name}"):
                                with log_container:
                                    try:
                                        # Extract answers using AI
                                        result = extract_answers_from_transcript(transcript.transcript_text, app.name)

                                        if result.get('answers'):
                                            answer_count = 0
                                            for answer_data in result['answers']:
                                                if answer_data.get('answer') and answer_data.get('confidence', 0) > 0.3:
                                                    # Check for duplicates
                                                    existing = session.query(TranscriptAnswer).filter_by(
                                                        transcript_id=transcript.id,
                                                        question_text=answer_data['question']
                                                    ).first()

                                                    if not existing:
                                                        ta = TranscriptAnswer(
                                                            id=str(uuid.uuid4()),
                                                            application_id=app.id,
                                                            transcript_id=transcript.id,
                                                            question_text=answer_data['question'],
                                                            answer_text=answer_data['answer'],
                                                            confidence_score=answer_data['confidence'],
                                                            synergy_block=answer_data.get('synergy_block', 'Unknown')
                                                        )
                                                        session.add(ta)
                                                        answer_count += 1

                                            transcript.processed = True
                                            session.commit()

                                            st.success(f"✅ {transcript.file_name}: Extracted {answer_count} answers")
                                            processed_count += 1
                                        else:
                                            st.warning(f"⚠️ {transcript.file_name}: No answers extracted")
                                            transcript.processed = True
                                            session.commit()
                                            processed_count += 1

                                    except Exception as e:
                                        st.error(f"❌ Error: {transcript.file_name} - {str(e)}")
                                        error_count += 1

                    # Final update
                    progress_bar.progress(1.0)
                    total_time = time.time() - start_time
                    progress_text.markdown(f"**✅ Complete!** Processed {len(pending_transcripts)} files in {int(total_time)}s")
                    status_spinner.empty()

                    # Final summary
                    st.markdown("---")
                    st.success("🎉 **PROCESSING COMPLETE!**")
                    st.markdown("### 📊 Summary")

                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("✅ Processed", processed_count)
                    with col2:
                        st.metric("❌ Errors", error_count)

                    if processed_count > 0:
                        st.balloons()
                        st.success("🎉 All pending transcripts have been processed!")

        # TAB 2: Batch Calculate Scores
        with tab2:
            st.markdown("#### Calculate Scores for All Applications")

            # Data source selection
            data_source = st.radio(
                "📊 Score calculation based on:",
                ["Both (All Applications)", "Questionnaire Only", "Questionnaire + Transcripts"],
                help="Choose data source for score calculation"
            )

            if data_source == "Both (All Applications)":
                st.info("🎯 Calculate scores for ALL apps (using Q+T when available, Q only otherwise)")
            elif data_source == "Questionnaire Only":
                st.info("🎯 Calculate scores based on questionnaire responses only (ignore transcripts)")
            else:
                st.info("🎯 Calculate scores ONLY for apps with both questionnaire AND transcript data")

            # Get all applications (exclude Questions Template)
            all_apps = session.query(Application).filter(Application.name != "Questions Template").all()

            # Filter apps based on selected data source
            eligible_apps = []
            for app in all_apps:
                qa_count = session.query(QuestionnaireAnswer).filter_by(application_id=app.id).count()
                ta_count = session.query(TranscriptAnswer).filter_by(application_id=app.id).count()

                # Include based on data source selection
                if data_source == "Both (All Applications)":
                    # Include all apps with questionnaire data
                    if qa_count > 0:
                        eligible_apps.append({
                            'app': app,
                            'qa_count': qa_count,
                            'ta_count': ta_count,
                            'data_type': 'Questionnaire' if ta_count == 0 else 'Q+T'
                        })
                elif data_source == "Questionnaire Only":
                    # Only apps with questionnaire (transcripts will be ignored even if present)
                    if qa_count > 0:
                        eligible_apps.append({
                            'app': app,
                            'qa_count': qa_count,
                            'ta_count': 0,  # Force to 0 to ignore transcripts
                            'data_type': 'Questionnaire'
                        })
                else:  # Questionnaire + Transcripts
                    # Only apps with BOTH
                    if qa_count > 0 and ta_count > 0:
                        eligible_apps.append({
                            'app': app,
                            'qa_count': qa_count,
                            'ta_count': ta_count,
                            'data_type': 'Q+T'
                        })

            if not eligible_apps:
                if data_source == "Both (All Applications)" or data_source == "Questionnaire Only":
                    st.warning("⚠️ No applications with questionnaire responses found.")
                else:
                    st.warning("⚠️ No applications with both questionnaire and transcript data found.")
                    st.info("💡 Tip: Switch to 'Both (All Applications)' to include apps without transcripts")
            else:
                # Show breakdown stats
                q_only_count = sum(1 for item in eligible_apps if item['data_type'] == 'Questionnaire')
                q_plus_t_count = sum(1 for item in eligible_apps if item['data_type'] == 'Q+T')

                if data_source == "Both (All Applications)":
                    st.success(f"✅ **{len(eligible_apps)} applications** ready for score calculation")
                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("📝 Questionnaire Only", q_only_count)
                    with col2:
                        st.metric("📝+🎙️ Questionnaire + Transcripts", q_plus_t_count)
                else:
                    st.success(f"✅ **{len(eligible_apps)} applications** ready for score calculation")

                # Show eligible apps with data source indicator
                with st.expander(f"View eligible applications ({len(eligible_apps)})"):
                    for item in eligible_apps:
                        app = item['app']
                        data_badge = "📝 Q" if item['data_type'] == 'Questionnaire' else "📝+🎙️ Q+T"
                        if item['ta_count'] > 0:
                            st.markdown(f"- {data_badge} **{app.name}** ({item['qa_count']} questionnaire, {item['ta_count']} transcript answers)")
                        else:
                            st.markdown(f"- {data_badge} **{app.name}** ({item['qa_count']} questionnaire answers)")

                st.markdown("---")

                # Calculate button
                col1, col2 = st.columns([3, 1])
                with col1:
                    overwrite = st.checkbox("🔄 Overwrite existing suggested scores", value=False)
                with col2:
                    if st.button("🎯 Calculate All Scores", type="primary", width="stretch"):
                        st.markdown("### 🤖 Calculating Scores...")

                        progress_bar = st.progress(0)
                        status_text = st.empty()

                        calculated_count = 0
                        skipped_count = 0
                        error_count = 0

                        for idx, item in enumerate(eligible_apps):
                            app = item['app']

                            status_text.text(f"🎯 Calculating {idx + 1}/{len(eligible_apps)}: {app.name}")

                            try:
                                # Check if suggested scores already exist
                                existing = session.query(SynergyScore).filter_by(
                                    application_id=app.id,
                                    approved=False
                                ).first()

                                if existing and not overwrite:
                                    st.info(f"⏭️ Skipping {app.name} (scores already exist)")
                                    skipped_count += 1
                                else:
                                    # Delete existing suggested scores if overwriting
                                    if existing and overwrite:
                                        session.query(SynergyScore).filter_by(
                                            application_id=app.id,
                                            approved=False
                                        ).delete()
                                        session.commit()

                                    # Get questionnaire answers
                                    qa_data = session.query(QuestionnaireAnswer).filter_by(
                                        application_id=app.id
                                    ).all()

                                    questionnaire_dict = {}
                                    for qa in qa_data:
                                        questionnaire_dict[qa.question_text] = {'a': qa.answer_text, 's': qa.score}

                                    # Get transcript answers
                                    transcript_answers = session.query(TranscriptAnswer).filter_by(
                                        application_id=app.id
                                    ).all()

                                    transcript_list = []
                                    for ta in transcript_answers:
                                        transcript_list.append({
                                            'question': ta.question_text,
                                            'answer': ta.answer_text,
                                            'confidence': ta.confidence_score
                                        })

                                    # Generate scores using AI
                                    if questionnaire_dict or transcript_list:
                                        result = suggest_scores(questionnaire_dict, transcript_list)

                                        if result.get('scores'):
                                            # Save suggested scores (auto-approved)
                                            for block_name, score_data in result['scores'].items():
                                                suggested_score = SynergyScore(
                                                    id=str(uuid.uuid4()),
                                                    application_id=app.id,
                                                    block_name=block_name,
                                                    score=score_data['score'],
                                                    suggested_by='ai_combined',
                                                    confidence=score_data.get('confidence', 0.8),
                                                    rationale=score_data.get('rationale', ''),
                                                    approved=True,
                                                    approved_by='auto_ai_generated',
                                                    approved_at=datetime.now(timezone.utc)
                                                )
                                                session.add(suggested_score)

                                            session.commit()
                                            st.success(f"✅ {app.name}: Scores calculated")
                                            calculated_count += 1
                                        else:
                                            st.warning(f"⚠️ {app.name}: No scores generated")
                                            error_count += 1
                                    else:
                                        st.warning(f"⚠️ {app.name}: No data available")
                                        error_count += 1

                            except Exception as e:
                                st.error(f"❌ Error: {app.name} - {str(e)}")
                                error_count += 1

                            progress_bar.progress((idx + 1) / len(eligible_apps))

                        # Final summary
                        status_text.empty()
                        st.markdown("---")
                        st.markdown("### 📊 Calculation Complete!")

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("✅ Calculated", calculated_count)
                        with col2:
                            st.metric("⏭️ Skipped", skipped_count)
                        with col3:
                            st.metric("❌ Errors", error_count)

                        if calculated_count > 0:
                            st.balloons()
                            st.success("🎉 Batch score calculation complete! Review scores in the Applications page.")

    finally:
        close_session(session)


# ==================== MAIN APPLICATION ====================
def main():
    # Top header bar - Professional, corporate style
    st.markdown("""
    <div style="background: linear-gradient(to right, #1F2937 0%, #111827 100%);
                padding: 1.25rem 2rem; margin: -6rem -6rem 1.5rem -6rem;
                border-bottom: 3px solid #E87722;
                box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);">
        <div style="display: flex; align-items: center; justify-content: space-between;">
            <div style="display: flex; align-items: center; gap: 1.25rem;">
                <div style="width: 40px; height: 40px;
                            background: linear-gradient(135deg, #E87722 0%, #FF8C42 100%);
                            border-radius: 10px; display: flex; align-items: center; justify-content: center;
                            box-shadow: 0 2px 6px rgba(232, 119, 34, 0.3);">
                    <span style="font-size: 1.5rem;">⚡</span>
                </div>
                <div>
                    <div style="margin:0;line-height:1;"><span style="color:#E87722;font-size:1.5rem;font-weight:700;letter-spacing:0.5px;">AVANGRID</span></div>
                    <div style="margin-top:0.25rem;"><span style="color:#D1D5DB;font-size:0.8125rem;font-weight:500;letter-spacing:0.3px;">Application Portfolio Management Platform</span></div>
                </div>
            </div>
            <div style="background: rgba(232, 119, 34, 0.1); color: #E87722; padding: 0.4rem 0.9rem;
                        border-radius: 16px; font-size: 0.6875rem; font-weight: 600; letter-spacing: 0.5px;
                        border: 1px solid rgba(232, 119, 34, 0.2);">
                STRATEGIC DECISION ENGINE
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Pure HTML/CSS menu with Streamlit buttons
    st.markdown("""
    <style>
        /* Style the menu buttons container */
        div[data-testid="stColumns"].menu-row {
            gap: 0 !important;
        }

        /* Style all buttons in the menu to look like nav links */
        .menu-btn-container button {
            background: transparent !important;
            border: none !important;
            color: #4B5563 !important;
            font-weight: 500 !important;
            font-size: 13px !important;
            padding: 12px 8px !important;
            width: 100% !important;
            border-radius: 0 !important;
            border-bottom: 2px solid transparent !important;
            transition: all 0.2s ease !important;
        }

        .menu-btn-container button:hover {
            color: #111827 !important;
            background: transparent !important;
            border-bottom: 2px solid #D1D5DB !important;
        }

        .menu-btn-container button:focus {
            box-shadow: none !important;
        }

        /* Active menu button styling */
        .menu-btn-active button {
            color: #E87722 !important;
            font-weight: 600 !important;
            border-bottom: 3px solid #E87722 !important;
            background: transparent !important;
        }

        .menu-btn-active button:hover {
            border-bottom: 3px solid #E87722 !important;
        }
    </style>
    """, unsafe_allow_html=True)

    # Menu items
    menu_items = [
        ("Introduction", "📄"),
        ("Methodology", "📋"),
        ("Calculator", "🔢"),
        ("Applications", "📦"),
        ("Analyses", "📊"),
        ("Uploads", "☁️"),
        ("Q&A Assistant", "💬"),
    ]

    current_page = st.session_state.get('current_page', 'Introduction')

    menu_cols = st.columns(len(menu_items))
    for idx, (label, icon) in enumerate(menu_items):
        with menu_cols[idx]:
            is_active = (current_page == label)
            css_class = "menu-btn-active" if is_active else "menu-btn-container"
            st.markdown(f'<div class="{css_class}">', unsafe_allow_html=True)
            if st.button(f"{icon}  {label}", key=f"menu_{label}", width="stretch"):
                st.session_state.current_page = label
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

    # Divider line
    st.markdown("<div style='border-bottom: 1px solid #E5E7EB; margin: 0 0 2rem 0;'></div>", unsafe_allow_html=True)

    selected = current_page

    # Old sidebar code removed - now using top horizontal menu
    if False:
        pass  # Sidebar disabled
        st.markdown("""
        <style>
        @keyframes gradient-shift {
            0%, 100% { background-position: 0% 50%; }
            50% { background-position: 100% 50%; }
        }
        .corporate-header {
            background: linear-gradient(to right, #1F2937 0%, #111827 100%);
            padding: 1.75rem 2.5rem;
            border-bottom: 3px solid #E87722;
            margin-bottom: 2rem;
            display: flex;
            align-items: center;
            justify-content: space-between;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
        }

        .header-left {
            display: flex;
            align-items: center;
            gap: 1.25rem;
        }

        .logo-corporate {
            width: 42px;
            height: 42px;
            background: linear-gradient(135deg, #E87722 0%, #FF8C42 100%);
            border-radius: 10px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 1.5rem;
            box-shadow: 0 2px 6px rgba(232, 119, 34, 0.3);
        }

        .header-branding {
            display: flex;
            flex-direction: column;
        }

        .company-name-corporate {
            color: #FFFFFF;
            margin: 0;
            font-size: 1.5rem;
            font-weight: 700;
            letter-spacing: 0.5px;
            line-height: 1;
        }

        .tagline-corporate {
            color: #9CA3AF;
            margin: 0.25rem 0 0 0;
            font-size: 0.8125rem;
            font-weight: 500;
            letter-spacing: 0.3px;
        }

        .header-right {
            display: flex;
            align-items: center;
            gap: 1rem;
        }

        .header-badge {
            background: rgba(232, 119, 34, 0.1);
            color: #E87722;
            padding: 0.4rem 0.9rem;
            border-radius: 16px;
            font-size: 0.6875rem;
            font-weight: 600;
            letter-spacing: 0.5px;
            border: 1px solid rgba(232, 119, 34, 0.2);
        }
        </style>

        <div class="corporate-header">
            <div class="header-left">
                <div class="logo-corporate">⚡</div>
                <div class="header-branding">
                    <h1 class="company-name-corporate">AVANGRID</h1>
                    <p class="tagline-corporate">Application Portfolio Management Platform</p>
                </div>
            </div>
            <div class="header-right">
                <span class="header-badge">STRATEGIC DECISION ENGINE</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='margin: 1rem 0;'></div>", unsafe_allow_html=True)

        # Menu options list
        menu_options = [
            "Introduction",
            "Methodology",
            "Calculator",
            "Applications",
            "Analyses",
            "Uploads",
            "Q&A Assistant"
        ]

        # Get default index from session state
        try:
            default_idx = menu_options.index(st.session_state.current_page)
        except (ValueError, AttributeError):
            default_idx = 0

        selected = option_menu(
            menu_title=None,
            options=menu_options,
            icons=[
                "book-half",
                "bookshelf",
                "calculator",
                "app",
                "graph-up",
                "cloud-upload",
                "chat-dots"
            ],
            menu_icon=None,
            default_index=default_idx,
            styles={
                "container": {
                    "padding": "0!important",
                    "background-color": "#FFFFFF",
                    "border-radius": "12px",
                    "border": "1px solid #E5E7EB"
                },
                "icon": {
                    "color": "#6B7280",
                    "font-size": "18px",
                    "margin-right": "10px"
                },
                "nav-link": {
                    "font-size": "14px",
                    "text-align": "left",
                    "margin": "2px 8px",
                    "padding": "12px 16px",
                    "color": "#4B5563",
                    "border-radius": "8px",
                    "font-weight": "500",
                    "transition": "all 0.2s ease",
                    "background": "transparent",
                    "border": "1px solid transparent",
                },
                "nav-link:hover": {
                    "background": "#F9FAFB",
                    "border": "1px solid #E5E7EB",
                    "color": "#111827",
                },
                "nav-link-selected": {
                    "background": "linear-gradient(to right, rgba(232, 119, 34, 0.1), rgba(255, 140, 66, 0.05))",
                    "color": "#E87722",
                    "font-weight": "600",
                    "border": "1px solid rgba(232, 119, 34, 0.3)",
                    "border-left": "3px solid #E87722",
                },
            }
        )

        # Update session state and rerun if changed
        if selected != st.session_state.current_page:
            st.session_state.current_page = selected
            st.rerun()

        st.markdown("<div style='margin: 2rem 0;'></div>", unsafe_allow_html=True)

        # Modern stats cards
        session = get_session()
        try:
            # Exclude Questions Template from count
            total_apps = session.query(Application).filter(Application.name != "Questions Template").count()
            total_transcripts = session.query(MeetingTranscript).count()
            pending_transcripts = session.query(MeetingTranscript).filter_by(processed=False).count()

            # Premium Portfolio Overview Section
            st.markdown("""
            <style>
            @keyframes counter-pop {
                0% { transform: scale(1); }
                50% { transform: scale(1.05); }
                100% { transform: scale(1); }
            }
            @keyframes card-hover-glow {
                0%, 100% { box-shadow: 0 4px 16px rgba(232, 119, 34, 0.15); }
                50% { box-shadow: 0 8px 32px rgba(232, 119, 34, 0.3); }
            }
            .stats-card-premium {
                background: linear-gradient(135deg, rgba(255, 255, 255, 0.95) 0%, rgba(248, 249, 250, 0.9) 100%);
                backdrop-filter: blur(20px);
                padding: 1.25rem;
                border-radius: 20px;
                margin-bottom: 1rem;
                border: 2px solid rgba(232, 119, 34, 0.1);
                box-shadow: 0 4px 16px rgba(0, 0, 0, 0.08);
                transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
                position: relative;
                overflow: hidden;
            }
            .stats-card-premium::before {
                content: '';
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                height: 3px;
                background: linear-gradient(90deg, #E87722 0%, #FF8C42 100%);
                opacity: 0;
                transition: opacity 0.4s ease;
            }
            .stats-card-premium:hover {
                transform: translateY(-4px) scale(1.02);
                border-color: rgba(232, 119, 34, 0.3);
                box-shadow: 0 12px 32px rgba(232, 119, 34, 0.2);
            }
            .stats-card-premium:hover::before {
                opacity: 1;
            }
            .stats-label-premium {
                color: #6B7280;
                font-size: 0.7rem;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 1.2px;
                margin-bottom: 0.5rem;
                display: flex;
                align-items: center;
                gap: 0.5rem;
            }
            .stats-icon-premium {
                font-size: 1.1rem;
                filter: grayscale(30%);
            }
            .stats-number-premium {
                color: #111827;
                font-size: 2.25rem;
                font-weight: 800;
                font-family: 'SF Pro Display', -apple-system, BlinkMacSystemFont, sans-serif;
                background: linear-gradient(135deg, #E87722 0%, #FF8C42 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                background-clip: text;
                transition: transform 0.3s ease;
            }
            .stats-card-premium:hover .stats-number-premium {
                animation: counter-pop 0.6s ease;
            }
            .pending-card-premium {
                background: linear-gradient(135deg, rgba(254, 243, 199, 0.95) 0%, rgba(253, 230, 138, 0.9) 100%);
                backdrop-filter: blur(20px);
                padding: 1.25rem;
                border-radius: 20px;
                margin-top: 1rem;
                border: 2px solid #F59E0B;
                box-shadow: 0 4px 16px rgba(245, 158, 11, 0.25);
                transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
                position: relative;
                overflow: hidden;
            }
            .pending-card-premium::before {
                content: '';
                position: absolute;
                top: 0;
                left: 0;
                right: 0;
                bottom: 0;
                background: linear-gradient(45deg, transparent 30%, rgba(255, 255, 255, 0.3) 50%, transparent 70%);
                animation: shimmer 3s infinite;
            }
            @keyframes shimmer {
                0% { transform: translateX(-100%); }
                100% { transform: translateX(100%); }
            }
            .pending-card-premium:hover {
                transform: translateY(-4px) scale(1.02);
                box-shadow: 0 12px 32px rgba(245, 158, 11, 0.4);
            }
            </style>

            <h4 style="color: #374151; font-weight: 700; margin-bottom: 1.5rem; font-size: 0.95rem;
                       text-transform: uppercase; letter-spacing: 1px;">
                📊 Portfolio Overview
            </h4>
            """, unsafe_allow_html=True)

            # Applications card with glassmorphism
            st.markdown(f"""
            <div class="stats-card-premium">
                <div class="stats-label-premium">
                    <span class="stats-icon-premium">📱</span>
                    Applications
                </div>
                <div class="stats-number-premium">{total_apps}</div>
            </div>
            """, unsafe_allow_html=True)

            # Transcripts card with glassmorphism
            st.markdown(f"""
            <div class="stats-card-premium">
                <div class="stats-label-premium">
                    <span class="stats-icon-premium">🎙️</span>
                    Transcripts
                </div>
                <div class="stats-number-premium">{total_transcripts}</div>
            </div>
            """, unsafe_allow_html=True)

            # Pending transcripts card with special effects
            if pending_transcripts > 0:
                st.markdown(f"""
                <div class="pending-card-premium">
                    <div style="color: #92400E; font-size: 0.85rem; font-weight: 700; margin-bottom: 0.5rem;
                                display: flex; align-items: center; gap: 0.5rem;">
                        <span style="font-size: 1.1rem;">⏳</span>
                        Pending Processing
                    </div>
                    <div style="color: #78350F; font-size: 1.75rem; font-weight: 800;">
                        {pending_transcripts} transcript{'s' if pending_transcripts > 1 else ''}
                    </div>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<div style='margin: 0.75rem 0;'></div>", unsafe_allow_html=True)
                if st.button("🚀 Process Queue", width="stretch", type="primary"):
                    st.session_state.current_page = "Batch Operations"
                    st.rerun()

        finally:
            close_session(session)

        # Premium footer with glassmorphism
        st.markdown("<div style='margin: 3rem 0 1.5rem 0;'></div>", unsafe_allow_html=True)
        st.markdown("""
        <style>
        @keyframes pulse-glow {
            0%, 100% { opacity: 0.6; }
            50% { opacity: 1; }
        }
        .footer-premium {
            text-align: center;
            padding: 1.5rem 1rem;
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.7) 0%, rgba(248, 249, 250, 0.6) 100%);
            backdrop-filter: blur(12px);
            border-radius: 16px;
            border: 1px solid rgba(232, 119, 34, 0.15);
            box-shadow: 0 4px 16px rgba(0, 0, 0, 0.05);
            margin: 0 0.5rem;
        }
        .footer-version {
            color: #374151;
            font-size: 0.8rem;
            font-weight: 700;
            margin-bottom: 0.5rem;
            letter-spacing: 0.5px;
        }
        .footer-powered {
            color: #6B7280;
            font-size: 0.7rem;
            font-weight: 500;
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 0.5rem;
        }
        .footer-ai-icon {
            background: linear-gradient(135deg, #E87722 0%, #FF8C42 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            font-weight: 800;
            animation: pulse-glow 2s ease-in-out infinite;
        }
        </style>
        <div class="footer-premium">
            <div class="footer-version">
                APM Platform v1.0
            </div>
            <div class="footer-powered">
                Powered by <span class="footer-ai-icon">OpenAI GPT-4</span>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Route to page
    if st.session_state.current_page == "Introduction":
        page_introduction()
    elif st.session_state.current_page == "Methodology":
        page_methodology()
    elif st.session_state.current_page == "Calculator":
        page_calculator()
    elif st.session_state.current_page == "Applications":
        page_applications()
    elif st.session_state.current_page == "Analyses":
        page_analyses()
    elif st.session_state.current_page == "Uploads":
        page_uploads()
    elif st.session_state.current_page == "Q&A Assistant":
        page_qa_assistant()


if __name__ == "__main__":
    main()
