"""
Excel Generator for Avangrid APM Platform
Generates a full Excel workbook matching the original APM format with:
Calculator, Dashboard, Strategic Roadmap, Application Groups, Value Chain,
and individual application sheets.
"""

import io
import os
import re
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import DataPoint
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from database import (
    get_session, close_session,
    Application, QuestionnaireAnswer, TranscriptAnswer, DavidNote, SynergyScore
)

EXCEL_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template_excel.xlsx")

# ============================================================
# Constants
# ============================================================

SYNERGY_BLOCKS = {
    'Strategic Fit': {'Type': 'Business', 'Weight': 30},
    'Business Efficiency': {'Type': 'Business', 'Weight': 30},
    'User Value': {'Type': 'Business', 'Weight': 20},
    'Financial Value': {'Type': 'Business', 'Weight': 20},
    'Architecture': {'Type': 'Technical', 'Weight': 30},
    'Operational Risk': {'Type': 'Technical', 'Weight': 30},
    'Maintainability': {'Type': 'Technical', 'Weight': 25},
    'Support Quality': {'Type': 'Technical', 'Weight': 15},
}

BLOCK_DEFINITIONS = {
    'Strategic Fit': {1: 'Completely misaligned', 2: 'Partially aligned', 3: 'Neutral', 4: 'Well-aligned', 5: 'Strategic driver'},
    'Business Efficiency': {1: 'Manual', 2: 'Low efficiency', 3: 'Average', 4: 'High', 5: 'Optimized'},
    'User Value': {1: 'Rejected', 2: 'Low satisfaction', 3: 'Acceptable', 4: 'Good', 5: 'Delightful'},
    'Financial Value': {1: 'Negative', 2: 'Poor', 3: 'Neutral', 4: 'Positive', 5: 'Exceptional'},
    'Architecture': {1: 'Obsolete', 2: 'Aging', 3: 'Stable', 4: 'Modern', 5: 'Future-proof'},
    'Operational Risk': {1: 'Critical', 2: 'High', 3: 'Managed', 4: 'Low', 5: 'Fortified'},
    'Maintainability': {1: 'Impossible', 2: 'Hard', 3: 'Standard', 4: 'Good', 5: 'Excellent'},
    'Support Quality': {1: 'Non-existent', 2: 'Reactive', 3: 'Defined', 4: 'Proactive', 5: 'World-class'},
}

REC_COLORS = {
    'EVOLVE': {'fill': 'C6EFCE', 'font': '006100'},
    'INVEST': {'fill': 'FFEB9C', 'font': '9C5700'},
    'MAINTAIN': {'fill': 'BDD7EE', 'font': '0066CC'},
    'ELIMINATE': {'fill': 'FFC7CE', 'font': '9C0006'},
}

MATRIX_CONFIG = [
    ('ELIMINATE', 'Replace', 'P1 - Critical', 'High Risk / EOL.'),
    ('ELIMINATE', 'Retire', 'P1 - Critical', 'Decommission'),
    ('ELIMINATE', 'Absorbed', 'P2 - Tactical', 'Consolidation'),
    ('INVEST', 'Absorb', 'P1 - Critical', 'High Value Opportunity'),
    ('INVEST', 'Modernize', 'P1 - Critical', 'Transformation'),
    ('EVOLVE', 'Migrate', 'P1 - Critical', 'Platform shift'),
    ('EVOLVE', 'Enhance', 'P2 - Strategic', 'Expansion'),
    ('EVOLVE', 'Refactor', 'P2 - Strategic', 'Code quality'),
    ('EVOLVE', 'Upgrade', 'P2 - Strategic', 'Version'),
    ('MAINTAIN', 'Internalize', 'P2 - Compliance', 'Governance'),
    ('MAINTAIN', 'Maintain', 'P3 - Routine', 'Keep lights on'),
]

APP_GROUP_CATEGORIES = {
    'Task Management / User Alignment': {
        'color': 'E87722',
        'keywords': ['task', 'schedule', 'track', 'assign', 'user', 'alignment', 'collab', 'arcos', 'jums', 'switching', 'ppe']
    },
    'Maintenance & Asset Mgmt': {
        'color': '0066B3',
        'keywords': ['maintain', 'asset', 'work order', 'inspection', 'repair', 'lifecycle', 'bentley', 'cimplicity', 'cathodic', 'poleforeman', 'esosr']
    },
    'Grid Operations / Engineering': {
        'color': '00B050',
        'keywords': ['scada', 'dms', 'oms', 'real-time', 'grid', 'voltage', 'design', 'model', 'calculate', 'engineer', 'kaffa', 'scal', 'dsd']
    },
    'Document / Info Management': {
        'color': '7030A0',
        'keywords': ['document', 'file', 'map', 'drawing', 'view', 'repository', 'knowledge', 'projectwise', 'mapping']
    },
    'Corporate / Administrative': {
        'color': 'FFC000',
        'keywords': ['finance', 'hr', 'legal', 'compliance', 'security', 'supply chain', 'admin', 'erp', 'sap']
    },
}

VALUE_CHAIN_STAGES = {
    'Generation (Renewables)': {
        'color': '00B0F0',
        'keywords': ['generation', 'renewable', 'wind', 'solar', 'hydro', 'offshore', 'onshore', 'plant', 'turbine', 'energy source', 'production']
    },
    'Transmission (Transport)': {
        'color': '0070C0',
        'keywords': ['transmission', 'high voltage', 'substation', 'interconnection', 'control center', 'ecc', 'tcc', 'line', 'relay', 'protection']
    },
    'Distribution (Delivery)': {
        'color': '00B050',
        'keywords': ['distribution', 'medium voltage', 'low voltage', 'ami', 'smart grid', 'meter', 'outage', 'oms', 'dms', 'scada', 'field', 'pole', 'circuit']
    },
    'Customer Solutions': {
        'color': 'FFC000',
        'keywords': ['customer', 'billing', 'crm', 'call center', 'service', 'payment', 'meter to cash', 'der', 'ev charging', 'portal', 'account']
    },
    'Corporate / Shared Services': {
        'color': '7030A0',
        'keywords': ['finance', 'hr', 'legal', 'compliance', 'security', 'supply chain', 'admin', 'it', 'procurement', 'cybersecurity', 'ehs']
    },
}

# Common styles
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL_DARK = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
HEADER_FILL_ORANGE = PatternFill(start_color='E87722', end_color='E87722', fill_type='solid')
HEADER_FILL_BLUE = PatternFill(start_color='0066B3', end_color='0066B3', fill_type='solid')
CONTENT_FONT = Font(name='Calibri', size=10)
CONTENT_FILL = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')


def sanitize_sheet_name(name):
    """Make a valid Excel sheet name (max 31 chars, no special chars)."""
    clean = re.sub(r'[\[\]:*?/\\\'"]', '', name)
    return clean[:31]


def get_recommendation(bvi, thi):
    if bvi >= 60 and thi >= 60:
        return 'EVOLVE'
    elif bvi >= 60 and thi < 60:
        return 'INVEST'
    elif bvi < 60 and thi >= 60:
        return 'MAINTAIN'
    else:
        return 'ELIMINATE'


def calculate_bvi_thi(scores, custom_weights=None):
    """Calculate BVI and THI from block scores."""
    weights = custom_weights or SYNERGY_BLOCKS
    bvi_blocks = ['Strategic Fit', 'Business Efficiency', 'User Value', 'Financial Value']
    thi_blocks = ['Architecture', 'Operational Risk', 'Maintainability', 'Support Quality']

    bvi_total = sum(scores.get(b, 0) * weights.get(b, {}).get('Weight', SYNERGY_BLOCKS[b]['Weight']) for b in bvi_blocks)
    bvi_weight = sum(weights.get(b, {}).get('Weight', SYNERGY_BLOCKS[b]['Weight']) for b in bvi_blocks)
    bvi = (bvi_total / bvi_weight * 20) if bvi_weight > 0 else 0

    thi_total = sum(scores.get(b, 0) * weights.get(b, {}).get('Weight', SYNERGY_BLOCKS[b]['Weight']) for b in thi_blocks)
    thi_weight = sum(weights.get(b, {}).get('Weight', SYNERGY_BLOCKS[b]['Weight']) for b in thi_blocks)
    thi = (thi_total / thi_weight * 20) if thi_weight > 0 else 0

    return round(bvi, 1), round(thi, 1)


def categorize_app(app_name, qa_texts):
    """Categorize an application into groups based on name and Q&A content."""
    search_text = (app_name + " " + " ".join(qa_texts)).lower()
    for cat_name, cat_info in APP_GROUP_CATEGORIES.items():
        for kw in cat_info['keywords']:
            if kw in search_text:
                return cat_name
    return 'Uncategorized'


def categorize_value_chain(app_name, qa_texts):
    """Categorize an application into value chain stage."""
    search_text = (app_name + " " + " ".join(qa_texts)).lower()
    for stage_name, stage_info in VALUE_CHAIN_STAGES.items():
        for kw in stage_info['keywords']:
            if kw in search_text:
                return stage_name
    return 'Cross-Cutting'


# ============================================================
# Sheet builders
# ============================================================

def build_calculator_sheet(wb, apps_data, custom_weights):
    """Build the Calculator sheet with scores, BVI, THI, and recommendations."""
    ws = wb.create_sheet("Calculator")
    blocks = list(SYNERGY_BLOCKS.keys())

    # Column widths
    ws.column_dimensions['A'].width = 30
    for i in range(len(blocks)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 18
    ws.column_dimensions[get_column_letter(len(blocks) + 2)].width = 22
    ws.column_dimensions[get_column_letter(len(blocks) + 3)].width = 22
    ws.column_dimensions[get_column_letter(len(blocks) + 4)].width = 18

    # Row 1: Headers
    ws.cell(row=1, column=1, value="Application Name").font = Font(name='Calibri', size=11, bold=True)
    ws['A1'].fill = HEADER_FILL_DARK
    ws['A1'].font = HEADER_FONT
    ws['A1'].border = THIN_BORDER

    for i, block in enumerate(blocks):
        col = i + 2
        cell = ws.cell(row=1, column=col, value=block)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        if SYNERGY_BLOCKS[block]['Type'] == 'Business':
            cell.fill = HEADER_FILL_ORANGE
        else:
            cell.fill = HEADER_FILL_BLUE

    bvi_col = len(blocks) + 2
    thi_col = len(blocks) + 3
    rec_col = len(blocks) + 4

    for col, label, fill in [(bvi_col, 'BVI', HEADER_FILL_ORANGE),
                              (thi_col, 'THI', HEADER_FILL_BLUE),
                              (rec_col, 'Recommendation', HEADER_FILL_DARK)]:
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    # Row 2: Weights
    ws.cell(row=2, column=1, value="Weight").font = Font(name='Calibri', size=9, italic=True, color='666666')
    for i, block in enumerate(blocks):
        w = custom_weights.get(block, SYNERGY_BLOCKS[block]['Weight'])
        cell = ws.cell(row=2, column=i + 2, value=f"{w}%")
        cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
        cell.alignment = Alignment(horizontal='center')

    # Freeze panes
    ws.freeze_panes = 'B4'

    # Data rows
    for row_idx, app in enumerate(apps_data, start=3):
        # App name
        cell = ws.cell(row=row_idx, column=1, value=app['name'])
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.border = THIN_BORDER

        # Block scores
        for i, block in enumerate(blocks):
            score = app['scores'].get(block, 0)
            cell = ws.cell(row=row_idx, column=i + 2, value=score)
            cell.font = CONTENT_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            # Conditional coloring
            if score <= 2:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            elif score == 3:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif score >= 4:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        # BVI, THI
        cell = ws.cell(row=row_idx, column=bvi_col, value=app['bvi'])
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0'

        cell = ws.cell(row=row_idx, column=thi_col, value=app['thi'])
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0'

        # Recommendation
        rec = app['recommendation']
        cell = ws.cell(row=row_idx, column=rec_col, value=rec)
        cell.font = Font(name='Calibri', size=10, bold=True, color=REC_COLORS.get(rec, {}).get('font', '000000'))
        cell.fill = PatternFill(start_color=REC_COLORS.get(rec, {}).get('fill', 'FFFFFF'),
                                end_color=REC_COLORS.get(rec, {}).get('fill', 'FFFFFF'), fill_type='solid')
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

    return ws


def _compute_label_positions(apps_data):
    """Compute optimal label positions ('t','b','l','r') for each app to minimize
    text collisions on the scatter chart. Uses a greedy algorithm that considers
    nearby neighbors and already-assigned label positions."""
    import math

    coords = [(app['thi'], app['bvi'], app['name']) for app in apps_data]
    # Sort by density (most crowded points first) so they get priority in placement
    neighbor_counts = []
    for i, (x1, y1, _) in enumerate(coords):
        count = sum(1 for j, (x2, y2, _) in enumerate(coords)
                    if i != j and math.hypot(x2 - x1, y2 - y1) < 15)
        neighbor_counts.append(count)
    order = sorted(range(len(coords)), key=lambda i: -neighbor_counts[i])

    positions = {}  # name -> 't'/'b'/'l'/'r'
    # Track occupied label regions: list of (lx, ly, direction)
    placed = []

    # Label approximate dimensions in chart units (0-100 scale)
    # Horizontal labels (~15 chars avg) are wider than tall
    LABEL_W = 12  # width in x-units
    LABEL_H = 5   # height in y-units

    def label_rect(px, py, direction):
        """Return (x1, y1, x2, y2) bounding box for a label placed at (px,py)."""
        if direction == 't':
            return (px - LABEL_W / 2, py + 1, px + LABEL_W / 2, py + 1 + LABEL_H)
        elif direction == 'b':
            return (px - LABEL_W / 2, py - 1 - LABEL_H, px + LABEL_W / 2, py - 1)
        elif direction == 'r':
            return (px + 2, py - LABEL_H / 2, px + 2 + LABEL_W, py + LABEL_H / 2)
        else:  # 'l'
            return (px - 2 - LABEL_W, py - LABEL_H / 2, px - 2, py + LABEL_H / 2)

    def rects_overlap(r1, r2):
        """Check if two rectangles overlap."""
        return not (r1[2] <= r2[0] or r2[2] <= r1[0] or r1[3] <= r2[1] or r2[3] <= r1[1])

    def count_overlaps(px, py, direction):
        """Count how many already-placed labels this position would overlap."""
        rect = label_rect(px, py, direction)
        overlaps = 0
        for (ox, oy, odir) in placed:
            other_rect = label_rect(ox, oy, odir)
            if rects_overlap(rect, other_rect):
                overlaps += 1
        # Also penalize labels that go out of bounds (0-100)
        if rect[0] < 0 or rect[2] > 100 or rect[1] < 0 or rect[3] > 100:
            overlaps += 2
        return overlaps

    for idx in order:
        x, y, name = coords[idx]
        # Try all 4 directions, pick the one with fewest overlaps
        candidates = ['t', 'r', 'b', 'l']
        best_dir = 't'
        best_overlaps = float('inf')
        for d in candidates:
            ov = count_overlaps(x, y, d)
            if ov < best_overlaps:
                best_overlaps = ov
                best_dir = d
                if ov == 0:
                    break  # perfect placement found
        positions[name] = best_dir
        placed.append((x, y, best_dir))

    return positions


def build_dashboard_sheet(wb, apps_data):
    """Build the Dashboard sheet with scatter chart matching the Streamlit app's chart.
    Features: app name labels on dots, no side legend, dashed gray lines at 60,60,
    grid numbers, large colored quadrant labels (EVOLVE/INVEST/MAINTAIN/ELIMINATE),
    white background with light gray gridlines."""
    ws = wb.create_sheet("Dashboard")

    # Data table for chart (hidden reference area)
    ws.cell(row=1, column=1, value="Application").font = Font(name='Calibri', size=9, color='999999')
    ws.cell(row=1, column=2, value="THI").font = Font(name='Calibri', size=9, color='999999')
    ws.cell(row=1, column=3, value="BVI").font = Font(name='Calibri', size=9, color='999999')
    ws.cell(row=1, column=4, value="Recommendation").font = Font(name='Calibri', size=9, color='999999')

    for i, app in enumerate(apps_data, start=2):
        ws.cell(row=i, column=1, value=app['name'])
        ws.cell(row=i, column=2, value=app['thi'])
        ws.cell(row=i, column=3, value=app['bvi'])
        ws.cell(row=i, column=4, value=app['recommendation'])

    n = len(apps_data)
    aux_start = n + 4

    # Quadrant label data points (invisible markers, visible text labels)
    quadrant_labels = [
        ("EVOLVE", 80, 80),
        ("INVEST", 30, 80),
        ("MAINTAIN", 80, 30),
        ("ELIMINATE", 30, 30),
    ]
    for i, (name, x, y) in enumerate(quadrant_labels):
        row = aux_start + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=x)
        ws.cell(row=row, column=3, value=y)

    # Create scatter chart
    chart = ScatterChart()
    chart.title = "Application Portfolio - Strategic Positioning"
    chart.x_axis.title = "Technical Health Index (THI)"
    chart.y_axis.title = "Business Value Index (BVI)"
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 100
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.width = 35
    chart.height = 22
    chart.style = 2

    # Remove legend (app names shown directly on dots)
    chart.legend = None

    # Explicitly show axes
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    # Axis tick marks and labels at edges (low = bottom/left)
    chart.x_axis.majorUnit = 20
    chart.y_axis.majorUnit = 20
    chart.x_axis.tickLblPos = 'low'
    chart.y_axis.tickLblPos = 'low'

    # Light gray major gridlines (matching Streamlit's gridcolor='lightgray')
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties

    grid_line_props = GraphicalProperties()
    grid_line_props.ln = LineProperties(solidFill='D3D3D3', w=6350)  # light gray, 0.5pt
    chart.x_axis.majorGridlines = ChartLines(spPr=grid_line_props)

    grid_line_props2 = GraphicalProperties()
    grid_line_props2.ln = LineProperties(solidFill='D3D3D3', w=6350)
    chart.y_axis.majorGridlines = ChartLines(spPr=grid_line_props2)

    # Number format for axis labels
    chart.x_axis.numFmt = '0'
    chart.y_axis.numFmt = '0'

    # Axes at min (0) - labels and tick marks at the edges
    chart.x_axis.crosses = 'min'
    chart.y_axis.crosses = 'min'

    # Marker colors matching Streamlit: EVOLVE=#10B981, INVEST=#F59E0B, MAINTAIN=#3B82F6, ELIMINATE=#EF4444
    marker_colors = {
        'EVOLVE': '10B981',
        'INVEST': 'F59E0B',
        'MAINTAIN': '3B82F6',
        'ELIMINATE': 'EF4444',
    }

    # Pre-compute optimal label positions to avoid collisions
    # For each point, find the direction with most clearance from neighbors
    label_positions = _compute_label_positions(apps_data)

    # Add each app as individual series with name label
    for i, app in enumerate(apps_data):
        row = i + 2
        rec = app['recommendation']
        x_vals = Reference(ws, min_col=2, min_row=row, max_row=row)
        y_vals = Reference(ws, min_col=3, min_row=row, max_row=row)
        series = Series(y_vals, x_vals, title=app['name'])
        series.marker = Marker(symbol='circle', size=8)
        color = marker_colors.get(rec, '999999')
        series.marker.graphicalProperties.solidFill = color
        series.graphicalProperties.line.noFill = True

        # Show app name as data label with smart positioning
        series.dLbls = DataLabelList()
        series.dLbls.showSerName = True
        series.dLbls.showVal = False
        series.dLbls.showCatName = False
        series.dLbls.showPercent = False
        series.dLbls.showLegendKey = False
        series.dLbls.dLblPos = label_positions.get(app['name'], 't')

        chart.series.append(series)

    # Add quadrant name labels as invisible data points with text
    for i, (name, x, y) in enumerate(quadrant_labels):
        row = aux_start + i
        x_ref = Reference(ws, min_col=2, min_row=row, max_row=row)
        y_ref = Reference(ws, min_col=3, min_row=row, max_row=row)
        s = Series(y_ref, x_ref, title=name)
        s.marker = Marker(symbol='none')
        s.graphicalProperties.line.noFill = True
        s.dLbls = DataLabelList()
        s.dLbls.showSerName = True
        s.dLbls.showVal = False
        s.dLbls.showCatName = False
        s.dLbls.showPercent = False
        s.dLbls.showLegendKey = False
        chart.series.append(s)

    # Dashed gray threshold lines at 60,60 (matching Streamlit's hline/vline)
    line_start = aux_start + len(quadrant_labels)

    # Horizontal dashed line at BVI=60
    ws.cell(row=line_start, column=2, value=0)
    ws.cell(row=line_start, column=3, value=60)
    ws.cell(row=line_start + 1, column=2, value=100)
    ws.cell(row=line_start + 1, column=3, value=60)

    h_x = Reference(ws, min_col=2, min_row=line_start, max_row=line_start + 1)
    h_y = Reference(ws, min_col=3, min_row=line_start, max_row=line_start + 1)
    h_series = Series(h_y, h_x, title=None)
    h_series.graphicalProperties.line.solidFill = '808080'
    h_series.graphicalProperties.line.dashStyle = 'dash'
    h_series.graphicalProperties.line.width = 19050  # 1.5pt
    h_series.marker = Marker(symbol='none')
    chart.series.append(h_series)

    # Vertical dashed line at THI=60
    ws.cell(row=line_start + 2, column=2, value=60)
    ws.cell(row=line_start + 2, column=3, value=0)
    ws.cell(row=line_start + 3, column=2, value=60)
    ws.cell(row=line_start + 3, column=3, value=100)

    v_x = Reference(ws, min_col=2, min_row=line_start + 2, max_row=line_start + 3)
    v_y = Reference(ws, min_col=3, min_row=line_start + 2, max_row=line_start + 3)
    v_series = Series(v_y, v_x, title=None)
    v_series.graphicalProperties.line.solidFill = '808080'
    v_series.graphicalProperties.line.dashStyle = 'dash'
    v_series.graphicalProperties.line.width = 19050  # 1.5pt
    v_series.marker = Marker(symbol='none')
    chart.series.append(v_series)

    # Place chart below data
    chart_row = line_start + 6
    ws.add_chart(chart, "A" + str(chart_row))

    return ws


def build_roadmap_sheet(wb, apps_data):
    """Build the Strategic Roadmap sheet."""
    ws = wb.create_sheet("Strategic Roadmap")

    headers = ['Application Name', 'BVI', 'THI', 'Recommendation', 'Subcategory', 'Quick Win?', 'Priority', 'Rationale', 'Comments']
    widths = [35, 15, 15, 18, 18, 12, 18, 25, 30]

    for i, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_DARK
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'

    for row_idx, app in enumerate(apps_data, start=2):
        rec = app['recommendation']
        rec_colors = REC_COLORS.get(rec, {'fill': 'FFFFFF', 'font': '000000'})

        # App name
        cell = ws.cell(row=row_idx, column=1, value=app['name'])
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.border = THIN_BORDER

        # BVI
        cell = ws.cell(row=row_idx, column=2, value=app['bvi'])
        cell.font = CONTENT_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0'

        # THI
        cell = ws.cell(row=row_idx, column=3, value=app['thi'])
        cell.font = CONTENT_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = '0.0'

        # Recommendation
        cell = ws.cell(row=row_idx, column=4, value=rec)
        cell.font = Font(name='Calibri', size=10, bold=True, color=rec_colors['font'])
        cell.fill = PatternFill(start_color=rec_colors['fill'], end_color=rec_colors['fill'], fill_type='solid')
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

        # Subcategory
        cell = ws.cell(row=row_idx, column=5, value=app.get('subcategory', ''))
        cell.font = CONTENT_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

        # Quick Win
        cell = ws.cell(row=row_idx, column=6, value='Yes' if app.get('quick_win') else 'No')
        cell.font = CONTENT_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

        # Priority
        priority = app.get('priority', '')
        cell = ws.cell(row=row_idx, column=7, value=priority)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')

        # Rationale - lookup from matrix
        rationale = ''
        for decision, subcat, prio, rat in MATRIX_CONFIG:
            if decision == rec and subcat == app.get('subcategory', ''):
                rationale = rat
                break
        cell = ws.cell(row=row_idx, column=8, value=rationale)
        cell.font = CONTENT_FONT
        cell.border = THIN_BORDER

        # Comments
        cell = ws.cell(row=row_idx, column=9, value='')
        cell.font = CONTENT_FONT
        cell.border = THIN_BORDER

    return ws


def build_app_groups_sheet(wb, apps_data):
    """Build the Application Groups sheet."""
    ws = wb.create_sheet("Application Groups")

    # Group apps
    groups = {cat: [] for cat in APP_GROUP_CATEGORIES}
    groups['Uncategorized'] = []

    for app in apps_data:
        qa_texts = [a for a in app.get('qa_answers', {}).values() if a]
        cat = categorize_app(app['name'], qa_texts)
        groups[cat].append(app)

    # Layout: 2 rows x 3 cols grid
    positions = [
        ('Task Management / User Alignment', 3, 2),
        ('Maintenance & Asset Mgmt', 3, 7),
        ('Grid Operations / Engineering', 3, 12),
        ('Document / Info Management', 25, 2),
        ('Corporate / Administrative', 25, 7),
        ('Uncategorized', 25, 12),
    ]

    # Title
    ws.merge_cells('B1:P1')
    cell = ws.cell(row=1, column=2, value="APPLICATION GROUPS - Functional Categorization")
    cell.font = Font(name='Calibri', size=14, bold=True, color='333333')
    cell.alignment = Alignment(horizontal='center')

    for cat_name, start_row, start_col in positions:
        color = APP_GROUP_CATEGORIES.get(cat_name, {}).get('color', '999999')
        apps_in_cat = groups.get(cat_name, [])

        # Header
        end_col = start_col + 3
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=cat_name)
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

        # App items
        for i, app in enumerate(apps_in_cat[:15]):
            row = start_row + 1 + i
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            cell = ws.cell(row=row, column=start_col, value=app['name'])
            cell.font = Font(name='Calibri', size=10)
            rec = app.get('recommendation', '')
            rec_colors = REC_COLORS.get(rec, {'fill': 'F2F2F2'})
            cell.fill = PatternFill(start_color=rec_colors['fill'], end_color=rec_colors['fill'], fill_type='solid')
            cell.border = THIN_BORDER

    # Column widths
    for col in [2, 7, 12]:
        for offset in range(4):
            ws.column_dimensions[get_column_letter(col + offset)].width = 10

    return ws


def build_value_chain_sheet(wb, apps_data):
    """Build the Value Chain sheet with improved layout."""
    ws = wb.create_sheet("Value Chain")

    # Title
    ws.merge_cells('B2:V2')
    cell = ws.cell(row=2, column=2, value="AVANGRID INTEGRATED UTILITY VALUE CHAIN")
    cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

    # Group apps
    stages = {s: [] for s in VALUE_CHAIN_STAGES}
    stages['Cross-Cutting'] = []

    for app in apps_data:
        qa_texts = [a for a in app.get('qa_answers', {}).values() if a]
        stage = categorize_value_chain(app['name'], qa_texts)
        stages[stage].append(app)

    # Wider columns (5 cols per stage instead of 4) and more spacing
    stage_positions = [
        ('Generation (Renewables)', 4, 2),
        ('Transmission (Transport)', 4, 8),
        ('Distribution (Delivery)', 4, 14),
        ('Customer Solutions', 4, 20),
        ('Corporate / Shared Services', 22, 2),
        ('Cross-Cutting', 22, 14),
    ]

    span = 4  # columns per stage box

    # Set column widths for stage areas
    for col in range(2, 26):
        ws.column_dimensions[get_column_letter(col)].width = 8

    for stage_name, start_row, start_col in stage_positions:
        color = VALUE_CHAIN_STAGES.get(stage_name, {}).get('color', '999999')
        apps_in_stage = stages.get(stage_name, [])
        end_col = start_col + span

        # Header
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
        cell = ws.cell(row=start_row, column=start_col, value=f"{stage_name} ({len(apps_in_stage)})")
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

        for i, app in enumerate(apps_in_stage[:15]):
            row = start_row + 1 + i
            rec = app.get('recommendation', '')
            rec_colors = REC_COLORS.get(rec, {'fill': 'F2F2F2', 'font': '000000'})

            # App name with BVI/THI
            display = f"{app['name']}  ({app.get('bvi', 0):.0f}/{app.get('thi', 0):.0f})"
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            cell = ws.cell(row=row, column=start_col, value=display)
            cell.font = Font(name='Calibri', size=9, color=rec_colors.get('font', '000000'))
            cell.fill = PatternFill(start_color=rec_colors['fill'], end_color=rec_colors['fill'], fill_type='solid')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')

    # Arrows between top-row stages
    for arrow_col in [7, 13, 19]:
        cell = ws.cell(row=4, column=arrow_col, value="\u27A1")
        cell.font = Font(name='Calibri', size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    return ws


def build_app_sheet(wb, app_data, session):
    """Build an individual application assessment sheet.
    4-column layout: Question | Questionnaire Answer | Transcript Answer | David's Comments
    Includes David's insights summary table near the scorecard."""
    sheet_name = sanitize_sheet_name(app_data['name'])
    ws = wb.create_sheet(sheet_name)

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 45

    # Title
    ws.merge_cells('A1:D1')
    cell = ws.cell(row=1, column=1, value=f"Assessment: {app_data['name']}")
    cell.font = Font(name='Calibri', size=14, bold=True, color='E87722')

    # Scorecard header
    ws.merge_cells('A3:B3')
    ws.cell(row=3, column=1, value="EXECUTIVE SCORECARD").font = HEADER_FONT
    ws['A3'].fill = HEADER_FILL_DARK
    ws['A3'].border = THIN_BORDER
    ws['B3'].fill = HEADER_FILL_DARK
    ws['B3'].border = THIN_BORDER
    ws.merge_cells('C3:D3')
    ws.cell(row=3, column=3, value="SCORE (0-5)").font = HEADER_FONT
    ws['C3'].fill = HEADER_FILL_DARK
    ws['C3'].border = THIN_BORDER
    ws['D3'].fill = HEADER_FILL_DARK
    ws['D3'].border = THIN_BORDER

    # Score rows
    blocks = list(SYNERGY_BLOCKS.keys())
    dv = DataValidation(type="whole", operator="between", formula1="0", formula2="5")
    ws.add_data_validation(dv)

    for i, block in enumerate(blocks, start=4):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        cell_a = ws.cell(row=i, column=1, value=block)
        cell_a.font = Font(name='Calibri', size=10, bold=True)
        cell_a.fill = CONTENT_FILL
        cell_a.border = THIN_BORDER
        cell_a.alignment = Alignment(horizontal='right')
        ws.cell(row=i, column=2).fill = CONTENT_FILL
        ws.cell(row=i, column=2).border = THIN_BORDER

        score = app_data['scores'].get(block, 0)
        ws.merge_cells(start_row=i, start_column=3, end_row=i, end_column=4)
        cell_b = ws.cell(row=i, column=3, value=score)
        cell_b.font = Font(name='Calibri', size=11)
        cell_b.border = THIN_BORDER
        cell_b.alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=4).border = THIN_BORDER
        dv.add(cell_b)

    # Get all data sources
    qa_answers = session.query(QuestionnaireAnswer).filter_by(application_id=app_data['id']).all()
    ta_answers = session.query(TranscriptAnswer).filter_by(application_id=app_data['id']).all()
    david_notes_answers = session.query(DavidNote).filter_by(application_id=app_data['id'], note_type='answer').all()
    david_notes_insights = session.query(DavidNote).filter_by(application_id=app_data['id'], note_type='insight').all()
    all_david_notes = session.query(DavidNote).filter_by(application_id=app_data['id']).all()

    # Build lookup dicts by question text
    qa_by_question = {}
    for qa in qa_answers:
        if qa.question_text:
            qa_by_question[qa.question_text] = qa.answer_text or '-'

    ta_by_question = {}
    for ta in ta_answers:
        if ta.question_text and ta.answer_text:
            ta_by_question[ta.question_text] = ta.answer_text

    dn_by_question = {}
    for dn in david_notes_answers:
        if dn.question_text and dn.answer_text:
            dn_by_question[dn.question_text] = dn.answer_text

    # ── David's Notes & Insights section (after scorecard) ──
    current_row = 13

    if all_david_notes:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        cell = ws.cell(row=current_row, column=1, value="DAVID'S NOTES & INSIGHTS")
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        cell.border = THIN_BORDER
        for c in range(2, 5):
            ws.cell(row=current_row, column=c).fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        # Column headers
        ws.cell(row=current_row, column=1, value="Topic").font = HEADER_FONT
        ws['A' + str(current_row)].fill = HEADER_FILL_DARK
        ws['A' + str(current_row)].border = THIN_BORDER
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        ws.cell(row=current_row, column=2, value="Insight / Comment").font = HEADER_FONT
        ws['B' + str(current_row)].fill = HEADER_FILL_DARK
        ws['B' + str(current_row)].border = THIN_BORDER
        for c in range(3, 5):
            ws.cell(row=current_row, column=c).fill = HEADER_FILL_DARK
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        for dn in all_david_notes:
            topic = dn.question_text or ('General Insights' if dn.note_type == 'insight' else 'Note')
            cell_topic = ws.cell(row=current_row, column=1, value=topic)
            cell_topic.font = Font(name='Calibri', size=9, bold=True)
            cell_topic.fill = CONTENT_FILL
            cell_topic.border = THIN_BORDER
            cell_topic.alignment = WRAP_ALIGN

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            cell_insight = ws.cell(row=current_row, column=2, value=dn.answer_text or '')
            cell_insight.font = Font(name='Calibri', size=9)
            cell_insight.fill = CONTENT_FILL
            cell_insight.border = THIN_BORDER
            cell_insight.alignment = WRAP_ALIGN
            for c in range(3, 5):
                ws.cell(row=current_row, column=c).fill = CONTENT_FILL
                ws.cell(row=current_row, column=c).border = THIN_BORDER
            current_row += 1

        current_row += 1  # Blank separator

    # ── Detailed Q&A sections per synergy block ──
    for block in blocks:
        # Block header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        cell = ws.cell(row=current_row, column=1, value=block.upper())
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='444444', end_color='444444', fill_type='solid')
        cell.border = THIN_BORDER
        for c in range(2, 5):
            ws.cell(row=current_row, column=c).fill = PatternFill(start_color='444444', end_color='444444', fill_type='solid')
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        # Definitions
        cell = ws.cell(row=current_row, column=1, value="Definitions")
        cell.font = Font(name='Calibri', size=10, bold=True)
        current_row += 1

        for score_val, definition in BLOCK_DEFINITIONS.get(block, {}).items():
            cell = ws.cell(row=current_row, column=1, value=f"{score_val} - {definition}")
            cell.font = Font(name='Calibri', size=9, color='666666')
            current_row += 1

        current_row += 1

        # Q&A header (4 columns)
        qa_headers = [
            ('Q', HEADER_FILL_DARK),
            ('A (Questionnaire)', HEADER_FILL_ORANGE),
            ('A (Transcript)', HEADER_FILL_BLUE),
            ("David's Comments", PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')),
        ]
        for col_idx, (header, fill) in enumerate(qa_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
        current_row += 1

        # Collect all unique questions for this block
        block_questions = set()
        for qa in qa_answers:
            if qa.synergy_block == block and qa.question_text:
                block_questions.add(qa.question_text)
        for ta in ta_answers:
            if ta.synergy_block == block and ta.question_text:
                block_questions.add(ta.question_text)
        for dn in david_notes_answers:
            if dn.synergy_block == block and dn.question_text:
                block_questions.add(dn.question_text)

        if block_questions:
            for question in sorted(block_questions):
                # Column A: Question
                cell_q = ws.cell(row=current_row, column=1, value=question)
                cell_q.font = Font(name='Calibri', size=9)
                cell_q.fill = CONTENT_FILL
                cell_q.border = THIN_BORDER
                cell_q.alignment = WRAP_ALIGN

                # Column B: Questionnaire Answer
                cell_qa = ws.cell(row=current_row, column=2, value=qa_by_question.get(question, '-'))
                cell_qa.font = Font(name='Calibri', size=9)
                cell_qa.fill = CONTENT_FILL
                cell_qa.border = THIN_BORDER
                cell_qa.alignment = WRAP_ALIGN

                # Column C: Transcript Answer
                cell_ta = ws.cell(row=current_row, column=3, value=ta_by_question.get(question, '-'))
                cell_ta.font = Font(name='Calibri', size=9)
                cell_ta.fill = CONTENT_FILL
                cell_ta.border = THIN_BORDER
                cell_ta.alignment = WRAP_ALIGN

                # Column D: David's Comments
                cell_dn = ws.cell(row=current_row, column=4, value=dn_by_question.get(question, '-'))
                cell_dn.font = Font(name='Calibri', size=9)
                cell_dn.fill = CONTENT_FILL
                cell_dn.border = THIN_BORDER
                cell_dn.alignment = WRAP_ALIGN

                current_row += 1
        else:
            cell = ws.cell(row=current_row, column=1, value="No data available")
            cell.font = Font(name='Calibri', size=9, italic=True, color='999999')
            current_row += 1

        current_row += 1  # Blank separator

    # Tab color based on recommendation
    rec = app_data.get('recommendation', '')
    tab_colors = {
        'EVOLVE': '00B050', 'INVEST': 'FFC000',
        'MAINTAIN': '4472C4', 'ELIMINATE': 'FF0000',
    }
    ws.sheet_properties.tabColor = tab_colors.get(rec, 'D9D9D9')

    return ws


# ============================================================
# Main generation function
# ============================================================

def copy_template_sheet(wb, source_ws, tab_name, position):
    """Copy a sheet from the template workbook preserving styles, merges, and dimensions."""
    target_ws = wb.create_sheet(tab_name, position)

    # Copy all cells and styles
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws[cell.coordinate]
            target_cell.value = cell.value
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)

    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Copy column dimensions
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width

    # Copy row dimensions
    for row_num, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dim.height

    # Copy images
    try:
        if hasattr(source_ws, '_images') and source_ws._images:
            from openpyxl.drawing.image import Image as XLImage
            for img in source_ws._images:
                new_image = XLImage(img.ref)
                if hasattr(img, 'anchor'):
                    new_image.anchor = img.anchor
                target_ws.add_image(new_image)
    except Exception:
        pass  # Images may fail in some environments

    return target_ws


def populate_index_sheet(ws, apps_data):
    """Update the Index sheet with a unified table of all sheets with descriptions and hyperlinks."""
    # Unmerge all merged cells in the dynamic area first (template may have merges)
    merges_to_remove = [str(m) for m in ws.merged_cells.ranges
                        if m.min_row >= 5]
    for merge_range in merges_to_remove:
        ws.unmerge_cells(merge_range)

    # Clear old dynamic content
    last_row = ws.max_row
    for row in range(5, last_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.hyperlink = None

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 18

    # Table header (row 5)
    header_row = 5
    headers = [('#', 6), ('Sheet', 30), ('Description', 55), ('Recommendation', 18)]
    for col_idx, (header, _) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_DARK
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fixed sheets with descriptions
    fixed_sheets = [
        ('Index', 'Table of contents and navigation'),
        ('Introduction', 'APM program overview, mission, and scope'),
        ('Methodology', 'Scoring framework and synergy block definitions'),
        ('Calculator', 'Application scores with BVI/THI calculations'),
        ('Dashboard', 'Portfolio scatter chart (BVI vs THI)'),
        ('Strategic Roadmap', 'Prioritized strategic action plan'),
        ('Value Chain', 'Applications mapped to utility value chain stages'),
    ]

    row = header_row + 1

    for idx, (sheet_name, description) in enumerate(fixed_sheets, start=1):
        # Number
        cell_num = ws.cell(row=row, column=1, value=idx)
        cell_num.font = CONTENT_FONT
        cell_num.border = THIN_BORDER
        cell_num.alignment = Alignment(horizontal='center')

        # Sheet name with hyperlink
        cell_name = ws.cell(row=row, column=2, value=sheet_name)
        cell_name.font = Font(name='Calibri', size=10, color='0066CC', underline='single')
        cell_name.border = THIN_BORDER
        cell_name.hyperlink = f"#'{sheet_name}'!A1"

        # Description
        cell_desc = ws.cell(row=row, column=3, value=description)
        cell_desc.font = CONTENT_FONT
        cell_desc.border = THIN_BORDER

        # Recommendation (empty for fixed sheets)
        cell_rec = ws.cell(row=row, column=4, value='')
        cell_rec.border = THIN_BORDER

        # Alternate row shading
        if idx % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = CONTENT_FILL

        row += 1

    # Application sheets
    for app_idx, app in enumerate(apps_data):
        idx = len(fixed_sheets) + app_idx + 1
        sheet_name = sanitize_sheet_name(app['name'])
        rec = app.get('recommendation', '')
        rec_colors = REC_COLORS.get(rec, {'fill': 'FFFFFF', 'font': '000000'})

        # Number
        cell_num = ws.cell(row=row, column=1, value=idx)
        cell_num.font = CONTENT_FONT
        cell_num.border = THIN_BORDER
        cell_num.alignment = Alignment(horizontal='center')

        # Sheet name with hyperlink
        cell_name = ws.cell(row=row, column=2, value=app['name'])
        cell_name.font = Font(name='Calibri', size=10, color='0066CC', underline='single')
        cell_name.border = THIN_BORDER
        cell_name.hyperlink = f"#'{sheet_name}'!A1"

        # Description
        desc = f"Assessment: {app['name']}"
        cell_desc = ws.cell(row=row, column=3, value=desc)
        cell_desc.font = CONTENT_FONT
        cell_desc.border = THIN_BORDER

        # Recommendation badge
        cell_rec = ws.cell(row=row, column=4, value=rec)
        cell_rec.font = Font(name='Calibri', size=10, bold=True, color=rec_colors.get('font', '000000'))
        cell_rec.fill = PatternFill(start_color=rec_colors.get('fill', 'FFFFFF'),
                                     end_color=rec_colors.get('fill', 'FFFFFF'), fill_type='solid')
        cell_rec.border = THIN_BORDER
        cell_rec.alignment = Alignment(horizontal='center')

        row += 1


def generate_portfolio_excel(custom_weights=None):
    """Generate the full portfolio Excel workbook.
    Returns bytes of the generated .xlsx file."""

    session = get_session()
    try:
        apps = session.query(Application).order_by(Application.name).all()
        if not apps:
            return None

        # Build apps_data list (exclude "Questions Template" which is not a real application)
        apps_data = []
        for app in apps:
            if app.name.strip().lower() == 'questions template':
                continue
            scores_data = session.query(SynergyScore).filter_by(
                application_id=app.id, approved=True
            ).all()

            scores = {s.block_name: s.score for s in scores_data}

            # Use custom weights
            w = custom_weights or {b: SYNERGY_BLOCKS[b]['Weight'] for b in SYNERGY_BLOCKS}
            weight_dict = {b: {'Weight': w.get(b, SYNERGY_BLOCKS[b]['Weight'])} for b in SYNERGY_BLOCKS}
            bvi, thi = calculate_bvi_thi(scores, weight_dict)
            rec = get_recommendation(bvi, thi)

            # Priority
            priority = ''
            if app.subcategory:
                priority_map = {m[1]: m[2] for m in MATRIX_CONFIG if m[0] == rec}
                base_priority = priority_map.get(app.subcategory, '')
                if not base_priority:
                    for decision, subcat, prio, rat in MATRIX_CONFIG:
                        if subcat == app.subcategory:
                            base_priority = prio
                            break
                if app.quick_win and base_priority.startswith('P2'):
                    priority = 'P1 - Quick Win'
                elif app.quick_win and base_priority.startswith('P3'):
                    priority = 'P2 - Quick Win'
                else:
                    priority = base_priority

            # Collect Q&A answers for grouping
            qa_answers = session.query(QuestionnaireAnswer).filter_by(application_id=app.id).all()
            qa_dict = {}
            for qa in qa_answers:
                if qa.answer_text:
                    qa_dict[qa.question_text] = qa.answer_text

            apps_data.append({
                'id': app.id,
                'name': app.name,
                'scores': scores,
                'bvi': bvi,
                'thi': thi,
                'recommendation': rec,
                'subcategory': app.subcategory or '',
                'quick_win': app.quick_win,
                'priority': priority,
                'qa_answers': qa_dict,
            })

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # 1. Copy Index, Introduction, Methodology from template
        try:
            wb_template = load_workbook(EXCEL_TEMPLATE_PATH)
            template_tabs = [("Index", 0), ("Introduction", 1), ("Methodology", 2)]
            for tab_name, position in template_tabs:
                if tab_name in wb_template.sheetnames:
                    copy_template_sheet(wb, wb_template[tab_name], tab_name, position)
            wb_template.close()
        except Exception:
            pass  # Template may not exist in all environments

        # 2. Build data sheets (positions after Index=0, Introduction=1, Methodology=2)
        build_calculator_sheet(wb, apps_data, custom_weights or {b: SYNERGY_BLOCKS[b]['Weight'] for b in SYNERGY_BLOCKS})
        build_dashboard_sheet(wb, apps_data)
        build_roadmap_sheet(wb, apps_data)
        build_value_chain_sheet(wb, apps_data)

        # 3. Individual app sheets
        for app in apps_data:
            build_app_sheet(wb, app, session)

        # 4. Update Index with dynamic hyperlinks
        if "Index" in wb.sheetnames:
            populate_index_sheet(wb["Index"], apps_data)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    finally:
        close_session(session)
